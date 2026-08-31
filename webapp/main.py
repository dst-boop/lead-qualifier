"""Lead Qualifier web app.

FastAPI backend serving the browser-based qualifier UI with two sign-in
providers. Email and calendar invites are sent from whichever account the
user signs in with:

  - Microsoft 365 (MSAL auth-code flow -> Microsoft Graph sendMail / events)
  - Google (OAuth 2.0 -> Gmail API send / Google Calendar API events)

Configure one provider or both; the UI only offers the ones configured.

Environment variables:
  MS_CLIENT_ID / MS_CLIENT_SECRET / MS_TENANT_ID   (see SETUP-microsoft.md)
  GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET          (see SETUP-google.md)
  WHITEPAGES_API_KEY [/ WHITEPAGES_BASE_URL]       (see SETUP-whitepages.md)
  ZI_CLIENT_ID / ZI_CLIENT_SECRET                  per-user ZoomInfo (see SETUP-zoominfo.md)
  EDGAR_USER_AGENT                                 required for SEC lookups (see SETUP-edgar.md)
  ZI_AUTH_URL / ZI_TOKEN_URL / ZI_API_BASE / ZI_SCOPES   override the defaults
  ANTHROPIC_API_KEY [/ CLAUDE_MODEL]               enables the AI QC button
  USE_FIRESTORE=0                                  force memory mode (see SETUP-firestore.md)
  KMS_KEY_NAME                                     envelope-encrypt stored tokens (see SETUP-firestore.md)
  DRIVE_LEADS_FILE                                 default Drive filename to look for
  APP_BASE_URL   Public URL of this app, no trailing slash
"""

import asyncio
import base64
import hashlib
import json
import os
import re
import secrets
import time
from email.message import EmailMessage
from pathlib import Path
from typing import Optional
from urllib.parse import urlencode

import anthropic
import httpx
import msal
from fastapi import Depends, FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse

from webapp import freesources, harvest, prospecting, signals
from pydantic import BaseModel, EmailStr

# --- Microsoft ---
MS_CLIENT_ID = os.environ.get("MS_CLIENT_ID", "")
MS_CLIENT_SECRET = os.environ.get("MS_CLIENT_SECRET", "")
MS_TENANT_ID = os.environ.get("MS_TENANT_ID", "common")
MS_AUTHORITY = f"https://login.microsoftonline.com/{MS_TENANT_ID}"
MS_SCOPES = ["User.Read", "Mail.Send", "Calendars.ReadWrite"]
# The employer-mailbox scope set, learned from a working example. Magic List
# connects to corporate tenants (Equitable's included — proven by a live
# connection) with calendar scopes and no Mail.Send at all: creating an event
# with attendees makes Exchange dispatch the invitation from the mailbox
# itself, through the firm's normal transport — journaled like any other
# outbound mail, no mail permission requested. Mail.Send is precisely the
# scope a corporate tenant is most likely to refuse, so the calendar-only
# footprint is what gets an employer account connected at all.
# (Their example also requests Calendars.*.Shared; adding an attendee needs
# only ReadWrite, so the narrower set is kept deliberately.)
MS_CAL_SCOPES = ["User.Read", "Calendars.ReadWrite"]
GRAPH = "https://graph.microsoft.com/v1.0"

# --- Money-in-motion sources (WARN notices, Form 5500) ---
# Both are free public data published for the purpose. Neither is reachable from
# the environment this was written in, but Cloud Run has ordinary egress, so the
# app fetches them at run time and /api/sources/probe reports what actually came
# back — the parser gets pinned to a real response rather than to a guess about
# column names.
#
# WARN_FEEDS is JSON: [{"id","state","format":"csv"|"json","url"}]. Empty by
# default, because a wrong URL that quietly 404s is worse than no feed at all.
WARN_FEEDS = os.environ.get("WARN_FEEDS", "")
FORM5500_URL = os.environ.get("FORM5500_URL", "")
FORM5500_CSV_IN_ZIP = os.environ.get("FORM5500_CSV_IN_ZIP", "f_5500")
# Schedule H and Schedule I, comma-separated. The 5500 file has no assets on
# it; these are where the money is, joined back on ACK_ID.
FORM5500_SCHEDULE_URLS = [u.strip() for u in
                          os.environ.get("FORM5500_SCHEDULE_URLS", "").split(",") if u.strip()]
SOURCE_STATES = os.environ.get("SOURCE_STATES", "")     # e.g. "NY,NJ,CT,PA"
SOURCE_COUNTIES = os.environ.get("SOURCE_COUNTIES", "")  # e.g. "Nassau,Suffolk"
SOURCE_MIN_WORKERS = int(os.environ.get("SOURCE_MIN_WORKERS", "25"))
FS_OPPS = os.environ.get("FIRESTORE_OPPS_COLLECTION", "opportunities")

# --- ZoomInfo via the MCP connector ---
# The DevPortal REST API and the MCP server are different doors. The REST API
# needs an entitlement this subscription does not carry; the MCP server takes a
# token of its own. Anthropic's MCP connector will hold that token and make the
# connection server-side, so the app never speaks ZoomInfo's protocol at all —
# it asks Claude, and Claude calls ZoomInfo with the user's own credentials.
ZI_MCP_URL = os.environ.get("ZI_MCP_URL", "https://mcp.zoominfo.com/mcp")
ZI_MCP_MODEL = os.environ.get("ZI_MCP_MODEL", "")   # defaults to CLAUDE_MODEL at call time

# --- SEC EDGAR (public filings) ---
# The one automated source in this app that is explicitly permitted rather than
# merely tolerated: the SEC asks for a descriptive User-Agent carrying a contact
# address and a ceiling of 10 requests a second, and grants access on that basis.
# Both obligations are honoured here — EDGAR_USER_AGENT has no default, because
# sending a made-up one is how a firm gets its IP range blocked.
EDGAR_USER_AGENT = os.environ.get("EDGAR_USER_AGENT", "")
EDGAR_DATA = os.environ.get("EDGAR_DATA", "https://data.sec.gov").rstrip("/")
EDGAR_WWW = os.environ.get("EDGAR_WWW", "https://www.sec.gov").rstrip("/")
EDGAR_FTS = os.environ.get("EDGAR_FTS", "https://efts.sec.gov").rstrip("/")
EDGAR_MAX_RPS = float(os.environ.get("EDGAR_MAX_RPS", "8"))   # SEC ceiling is 10
# FEC's demo key works out of the box at 40 requests/hour shared per IP. A
# personal key (free, instant, api.open.fec.gov/developers) raises that to
# 1,000/hour, so the default is "works today, upgrade when it matters".
FEC_API_KEY = os.environ.get("FEC_API_KEY", "DEMO_KEY")
FEC_API_BASE = os.environ.get("FEC_API_BASE", "https://api.open.fec.gov/v1")
EFTS_URL = os.environ.get("EFTS_URL", "https://efts.sec.gov/LATEST/search-index")

# --- ZoomInfo (per-user OAuth) ---
# Each advisor connects their own ZoomInfo seat, so searches and enrichment are
# attributed and billed to them rather than to one shared service account.
# A Standard app in the ZoomInfo DevPortal covers a single org, which is what a
# firm needs; distributing to other firms would require a Partner app instead.
ZI_CLIENT_ID = os.environ.get("ZI_CLIENT_ID", "")
ZI_CLIENT_SECRET = os.environ.get("ZI_CLIENT_SECRET", "")
# Endpoints are configurable because they are the one part of this that has not
# been exercised against a live tenant — see SETUP-zoominfo.md before trusting
# the defaults.
ZI_AUTH_URL = os.environ.get("ZI_AUTH_URL", "https://auth.zoominfo.com/authorize")
ZI_TOKEN_URL = os.environ.get("ZI_TOKEN_URL", "https://auth.zoominfo.com/oauth/token")
ZI_API_BASE = os.environ.get("ZI_API_BASE", "https://api.zoominfo.com").rstrip("/")
ZI_SCOPES = os.environ.get("ZI_SCOPES", "openid profile email offline_access")

# --- Google ---
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_USERINFO_URL = "https://openidconnect.googleapis.com/v1/userinfo"
GMAIL_SEND_URL = "https://gmail.googleapis.com/gmail/v1/users/me/messages/send"
GCAL_EVENTS_URL = "https://www.googleapis.com/calendar/v3/calendars/primary/events"
GOOGLE_SCOPES = (
    "openid email profile "
    "https://www.googleapis.com/auth/gmail.send "
    "https://www.googleapis.com/auth/calendar.events "
    # Read-only, and used only to fetch a named export by filename. drive.file
    # would be narrower but requires running Google's Picker in the browser
    # with the access token, which would put a third-party token in the page.
    "https://www.googleapis.com/auth/drive.readonly "
    # Lists the addresses Gmail will let this account send as — a work alias, a
    # shared team address. Read-only and settings-only; it cannot change them.
    "https://www.googleapis.com/auth/gmail.settings.basic"
)
GMAIL_SENDAS_URL = "https://gmail.googleapis.com/gmail/v1/users/me/settings/sendAs"
DRIVE_FILES_URL = "https://www.googleapis.com/drive/v3/files"
SHEET_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# --- WhitePages (phone verification) ---
# Two flavours of this API exist and they are not compatible:
#   Whitepages Pro   https://api.whitepages.com   /v2/phone   header X-Api-Key
#   Trestle          https://api.trestleiq.com    /3.1/phone  header x-api-key
# The path is inferred from the base URL; WHITEPAGES_PHONE_PATH overrides it.
WHITEPAGES_API_KEY = os.environ.get("WHITEPAGES_API_KEY", "")
WHITEPAGES_BASE_URL = os.environ.get("WHITEPAGES_BASE_URL", "https://api.whitepages.com").rstrip("/")
WHITEPAGES_PHONE_PATH = os.environ.get("WHITEPAGES_PHONE_PATH", "")
WHITEPAGES_PERSON_PATH = os.environ.get("WHITEPAGES_PERSON_PATH", "")
WHITEPAGES_PROPERTY_PATH = os.environ.get("WHITEPAGES_PROPERTY_PATH", "")

# --- Claude (AI lead QC) ---
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-opus-5")

BASE_URL = os.environ.get("APP_BASE_URL", "http://localhost:8000").rstrip("/")
STATIC_DIR = Path(__file__).parent / "static"
SESSION_COOKIE = "lq_session"
# A month, not a workday: with sessions KMS-sealed server-side and every
# attachment vaulted by identity, the only thing an 8-hour cookie bought was
# a daily sign-in ritual. Override with SESSION_TTL_SECONDS.
SESSION_TTL = int(os.environ.get("SESSION_TTL_SECONDS", str(30 * 24 * 3600)))

# ------------------------- Storage -------------------------
# Sessions and saved lead lists live in Firestore so they survive the container.
# Cloud Run load-balances across instances and recycles them freely, so an
# in-process dict means sign-in breaks the moment a second instance appears.
# Falls back to memory when Firestore is unreachable — fine for local runs,
# and /api/me reports which backend is live so a misconfigured deploy is
# visible rather than silently ephemeral.

USE_FIRESTORE = os.environ.get("USE_FIRESTORE", "1") != "0"
FS_SESSIONS = os.environ.get("FIRESTORE_SESSIONS_COLLECTION", "sessions")
# Refresh tokens keyed by who the person is, not by which browser cookie they
# happen to hold. A session dies with its cookie; the grant the user made does
# not, and pretending it did is why the consent screen kept coming back.
FS_VAULT = os.environ.get("FIRESTORE_VAULT_COLLECTION", "token_vault")
# Password accounts: identity for people whose email is not a Google or
# Microsoft login. The account is the identity; providers are attachments.
FS_ACCOUNTS = os.environ.get("FIRESTORE_ACCOUNTS_COLLECTION", "accounts")
FS_STATE = os.environ.get("FIRESTORE_STATE_COLLECTION", "lead_state")
# Paid-lookup answers, shared and durable. The in-process cache saved a credit
# only until Cloud Run recycled the instance — which it does whenever traffic
# stops — so the "already asked this" saving evaporated overnight, every night,
# and the same lead cost a second credit in the morning. Sealed, because these
# documents hold dates of birth and home addresses.
FS_CACHE = os.environ.get("FIRESTORE_CACHE_COLLECTION", "lookup_cache")
# What has been spent this calendar month, against a monthly allowance. One
# document per month, incremented atomically, because the process-lifetime
# counter it replaces reported a fraction of the truth after every recycle.
FS_LEDGER = os.environ.get("FIRESTORE_LEDGER_COLLECTION", "credit_ledger")
# One document per list rather than one per user. An advisor working four
# campaigns keeps four independent documents, so a five-thousand-row import
# cannot push the others towards Firestore's per-document limit, and switching
# lists loads only the one being opened.
FS_LISTS = os.environ.get("FIRESTORE_LISTS_COLLECTION", "lead_lists")
# Lists shared with you, keyed by your email: a reverse index, because
# Firestore cannot ask "which of everyone's lists name me" without one.
FS_SHARED = os.environ.get("FIRESTORE_SHARED_COLLECTION", "lead_shares")
# Everyone who has ever signed in, so an admin can be shown the firm rather
# than only the people who happen to have shared something with them. Firestore
# cannot answer "who exists" without an index somebody wrote.
FS_DIRECTORY = os.environ.get("FIRESTORE_DIRECTORY_COLLECTION", "directory")
# Admins are named in the environment, not promoted inside the app. An admin
# can read and reclaim every lead in the firm, so the grant lives where only a
# deploy can change it: a compromised account cannot promote itself, and there
# is no bootstrap problem where the first admin has to be created by hand.
ADMIN_EMAILS = os.environ.get("ADMIN_EMAILS", "")
_ADMINS = {e.strip().lower() for e in ADMIN_EMAILS.split(",") if e.strip()}


def _is_admin(email: str) -> bool:
    return bool(email) and email.lower() in _ADMINS
# One document per advisor per day of counted activity. Aggregating the team
# from everyone's lead documents would mean reading every lead in the firm to
# draw a leaderboard; a daily counter is two numbers and a date.
FS_STATS = os.environ.get("FIRESTORE_STATS_COLLECTION", "advisor_stats")
FS_BATTLES = os.environ.get("FIRESTORE_BATTLES_COLLECTION", "battles")
# Who counts as a teammate. Everyone at the same email domain is the rule a
# single firm actually wants, and it needs no invitations to administer.
TEAM_BY_DOMAIN = os.environ.get("TEAM_BY_DOMAIN", "1") not in ("0", "false", "no")

_MEM_SESSIONS: dict[str, dict] = {}
_MEM_STATE: dict[str, dict] = {}
_MEM_DIRECTORY: dict[str, dict] = {}
_fs_client = None
_fs_failed = False

# ------------------------- Token encryption -------------------------
# Session documents hold OAuth refresh tokens. Firestore encrypts at rest with
# Google-managed keys, but anyone who can read the database can read those
# tokens and use them. Envelope encryption puts a second lock on: a random
# data key per write encrypts the payload, KMS wraps the data key, and only
# the ciphertext and the wrapped key are stored. A Firestore reader without
# KMS decrypt permission then holds nothing usable.
#
# Unset KMS_KEY_NAME keeps the previous behaviour, so this is opt-in and the
# app still runs locally and in any deploy that has not configured a key.
# /api/me reports which is live.

KMS_KEY_NAME = os.environ.get("KMS_KEY_NAME", "")
_kms_client = None
_kms_failed = False


def _kms():
    global _kms_client, _kms_failed
    if not KMS_KEY_NAME or _kms_failed:
        return None
    if _kms_client is None:
        try:
            from google.cloud import kms
            _kms_client = kms.KeyManagementServiceClient()
        except Exception as e:
            print(f"[crypto] KMS unavailable, storing tokens unwrapped: {e}")
            _kms_failed = True
            return None
    return _kms_client


def encryption_backend() -> str:
    return "kms" if _kms() is not None else "google-managed"


def _seal(plaintext: str) -> dict:
    """{'data': ...} unwrapped, or {'ct','dek'} envelope-encrypted."""
    client = _kms()
    if client is None:
        return {"data": plaintext}
    try:
        from cryptography.fernet import Fernet
        dek = Fernet.generate_key()
        ct = Fernet(dek).encrypt(plaintext.encode()).decode()
        wrapped = client.encrypt(request={"name": KMS_KEY_NAME, "plaintext": dek})
        return {"ct": ct, "dek": base64.b64encode(wrapped.ciphertext).decode()}
    except Exception as e:
        # Never lose a session because encryption broke.
        print(f"[crypto] seal failed, storing unwrapped: {e}")
        return {"data": plaintext}


def _unseal(doc: dict) -> Optional[str]:
    """Read either shape, so existing documents keep working."""
    if doc.get("data") is not None:
        return doc["data"]
    if not doc.get("ct") or not doc.get("dek"):
        return None
    client = _kms()
    if client is None:
        print("[crypto] document is sealed but KMS is unavailable")
        return None
    try:
        from cryptography.fernet import Fernet
        unwrapped = client.decrypt(
            request={"name": KMS_KEY_NAME,
                     "ciphertext": base64.b64decode(doc["dek"])})
        return Fernet(unwrapped.plaintext).decrypt(doc["ct"].encode()).decode()
    except Exception as e:
        print(f"[crypto] unseal failed: {e}")
        return None


def _firestore():
    """Async Firestore client, or None when unavailable."""
    global _fs_client, _fs_failed
    if not USE_FIRESTORE or _fs_failed:
        return None
    if _fs_client is None:
        try:
            from google.cloud import firestore
            _fs_client = firestore.AsyncClient()
        except Exception as e:            # no SDK, no credentials, no project
            print(f"[storage] Firestore unavailable, using memory: {e}")
            _fs_failed = True
            return None
    return _fs_client


def storage_backend() -> str:
    return "firestore" if _firestore() is not None else "memory"


async def _fs_get(collection: str, key: str) -> Optional[dict]:
    global _fs_failed
    db = _firestore()
    if db is None:
        return None
    try:
        snap = await db.collection(collection).document(key).get()
        return snap.to_dict() if snap.exists else None
    except Exception as e:
        print(f"[storage] Firestore read failed, using memory: {e}")
        _fs_failed = True
        return None


async def _fs_set(collection: str, key: str, value: dict) -> bool:
    global _fs_failed
    db = _firestore()
    if db is None:
        return False
    try:
        await db.collection(collection).document(key).set(value)
        return True
    except Exception as e:
        print(f"[storage] Firestore write failed, using memory: {e}")
        _fs_failed = True
        return False


async def _fs_inc(collection: str, key: str, deltas: dict) -> bool:
    """Add to numeric fields atomically, creating the document if needed.

    Read-modify-write would lose an increment whenever two lookups overlap, and
    a lost increment means an under-count, which means over-spending against an
    allowance the user set precisely so that could not happen.
    """
    global _fs_failed
    db = _firestore()
    if db is None:
        return False
    try:
        from google.cloud import firestore
        await db.collection(collection).document(key).set(
            {k: firestore.Increment(v) for k, v in deltas.items()}, merge=True)
        return True
    except Exception as e:
        print(f"[storage] Firestore increment failed, using memory: {e}")
        _fs_failed = True
        return False


# ------------------------- the credit ledger -------------------------
#
# "Costs MUST be minimized when using paid enrichment. ZoomInfo has 2000
# credits per month. White pages has 1000 credits per month."
#
# Two allowances, one rule: nothing bills twice for the same question, and
# nothing bills at all once the month's allowance is gone. The vendors enforce
# their own caps eventually — with a 429 and a wasted round trip — but a cap
# discovered at the vendor is a cap discovered too late to choose differently.

# Two ceilings, and both are real. The account-wide one is what the vendor
# will actually bill; the per-user one is what stops the first advisor awake
# on the first of the month from spending everyone else's share by lunchtime.
WP_MONTHLY = int(os.environ.get("WHITEPAGES_MONTHLY_CREDITS", "1000"))
WP_PER_USER = int(os.environ.get("WHITEPAGES_USER_CREDITS", "100"))
# ZoomInfo is deliberately NOT here. Every user brings their own subscription:
# the app never holds a ZoomInfo seat of its own, and enrichment runs through
# the user's own saved token or their own Claude connector. Counting a shared
# pool would have been counting a pool that does not exist. What is counted is
# usage, further down — how many enrichments this app ran for them — because
# that is the number they reconcile against their own dashboard.
_MEM_LEDGER: dict = {}


def _month() -> str:
    return time.strftime("%Y-%m", time.gmtime())


def _month_resets() -> str:
    """First day of next month, ISO, so the UI can say when the pool refills."""
    y, m = (int(x) for x in _month().split("-"))
    return f"{y + 1:04d}-01-01" if m == 12 else f"{y:04d}-{m + 1:02d}-01"


def _ledger_key(who: str = "") -> str:
    """The firm's month, or one person's month within it.

    Two documents rather than one map of everybody: a single document
    incremented by every lookup from every advisor is a hot key, and Firestore
    would rather have the writes spread out.
    """
    return f"{_month()}|{who.lower()}" if who else _month()


async def _ledger_read(who: str = "") -> dict:
    key = _ledger_key(who)
    doc = await _fs_get(FS_LEDGER, key) or _MEM_LEDGER.get(key) or {}
    return {"wp": int(doc.get("wp") or 0), "zi": int(doc.get("zi") or 0)}


async def _ledger_add(kind: str, n: int = 1, who: str = "") -> None:
    if n <= 0:
        return
    keys = [_ledger_key(), _ledger_key(who or "unattributed")]
    for key in keys:
        if not await _fs_inc(FS_LEDGER, key, {kind: n}):
            cur = _MEM_LEDGER.setdefault(key, {})
            cur[kind] = int(cur.get(kind) or 0) + n


async def _wp_left(who: str = "") -> dict:
    """What is left for this person, and why — theirs, or the firm's.

    Both are returned rather than just the smaller, because they are different
    problems with different fixes: your own allowance resets on the first, and
    the firm's needs the admin to buy more.
    """
    firm = max(0, WP_MONTHLY - (await _ledger_read())["wp"])
    # An unnameable spender is still one spender. Falling back to the firm's
    # pool here handed them all thousand — the per-user cap applied to everyone
    # except the sessions too old to carry a name, which is precisely backwards.
    # They share one unattributed hundred instead, so the guarantee holds:
    # nobody spends more than WP_PER_USER without being named.
    mine = max(0, WP_PER_USER - (await _ledger_read(who or "unattributed"))["wp"])
    return {"mine": mine, "firm": firm, "left": min(mine, firm),
            "named": bool(who),
            "capped_by": "you" if mine <= firm else "firm"}


async def _fs_del(collection: str, key: str) -> bool:
    db = _firestore()
    if db is None:
        return False
    try:
        await db.collection(collection).document(key).delete()
        return True
    except Exception as e:
        print(f"[storage] Firestore delete failed: {e}")
        return False


async def load_session(sid: str) -> Optional[dict]:
    doc = await _fs_get(FS_SESSIONS, sid)
    if doc is not None:
        if doc.get("expires_at", 0) < time.time():
            return None
        raw = _unseal(doc)
        if raw is None:
            return None
        try:
            return json.loads(raw)
        except ValueError:
            return None
    return _MEM_SESSIONS.get(sid)


async def save_session(sid: str, session: dict) -> None:
    payload = _seal(json.dumps(session))
    payload["expires_at"] = time.time() + SESSION_TTL
    if not await _fs_set(FS_SESSIONS, sid, payload):
        _MEM_SESSIONS[sid] = session


async def drop_session(sid: str) -> None:
    _MEM_SESSIONS.pop(sid, None)
    db = _firestore()
    if db is not None:
        try:
            await db.collection(FS_SESSIONS).document(sid).delete()
        except Exception as e:
            print(f"[storage] Firestore delete failed: {e}")


app = FastAPI(title="Lead Qualifier")
from fastapi.staticfiles import StaticFiles  # noqa: E402
app.mount("/static", StaticFiles(directory=Path(__file__).parent / "static"), name="static")


@app.middleware("http")
async def https_middleware(request: Request, call_next):
    """Send plain-HTTP visitors to HTTPS, and keep them there.

    Cloud Run terminates TLS and passes the original scheme in
    X-Forwarded-Proto. The header is absent for local runs and for the
    platform's own container probes, so only an explicit "http" redirects.
    """
    if not BASE_URL.startswith("https"):
        return await call_next(request)

    if request.headers.get("x-forwarded-proto") == "http":
        return RedirectResponse(
            str(request.url.replace(scheme="https")), status_code=301
        )

    response = await call_next(request)
    # Without this, a browser that has once reached the site over HTTP will
    # keep guessing HTTP from its own autocomplete.
    response.headers["Strict-Transport-Security"] = "max-age=31536000"
    return response


@app.middleware("http")
async def session_middleware(request: Request, call_next):
    sid = request.cookies.get(SESSION_COOKIE)
    session = await load_session(sid) if sid else None
    is_new = session is None
    if is_new:
        sid = secrets.token_urlsafe(32)
        session = {}
    request.state.sid = sid
    request.state.session = session

    # Handlers mutate the session dict in place, so compare against a snapshot
    # rather than writing on every request — most requests change nothing.
    before = json.dumps(session, sort_keys=True)
    response = await call_next(request)
    if not getattr(request.state, "session_dropped", False):
        if is_new or json.dumps(session, sort_keys=True) != before:
            await save_session(sid, session)

    if is_new:
        response.set_cookie(
            SESSION_COOKIE, sid,
            httponly=True, samesite="lax",
            secure=BASE_URL.startswith("https"),
            max_age=SESSION_TTL,
        )
    return response


# ------------------------- Microsoft (MSAL) -------------------------

def _ms_load_cache(session: dict) -> msal.SerializableTokenCache:
    cache = msal.SerializableTokenCache()
    if session.get("ms_token_cache"):
        cache.deserialize(session["ms_token_cache"])
    return cache


def _ms_save_cache(session: dict, cache: msal.SerializableTokenCache) -> None:
    if cache.has_state_changed:
        session["ms_token_cache"] = cache.serialize()


def _ms_app(cache: Optional[msal.SerializableTokenCache] = None) -> msal.ConfidentialClientApplication:
    return msal.ConfidentialClientApplication(
        MS_CLIENT_ID, authority=MS_AUTHORITY,
        client_credential=MS_CLIENT_SECRET, token_cache=cache,
    )


def _ms_mail_ok(session: dict) -> bool:
    """Whether this Microsoft connection may send mail at all.

    A calendar-only connection is a deliberate choice, not a failure: an
    employer tenant that refuses Mail.Send will still approve calendar
    scopes, and an invite created with attendees goes out from the mailbox
    via Exchange itself — no mail permission involved.
    """
    return session.get("ms_mode", "full") != "calendar"


def _ms_token(session: dict) -> Optional[str]:
    if not MS_CLIENT_ID:
        return None
    cache = _ms_load_cache(session)
    client = _ms_app(cache)
    accounts = client.get_accounts()
    if not accounts:
        return None
    scopes = MS_CAL_SCOPES if not _ms_mail_ok(session) else MS_SCOPES
    result = client.acquire_token_silent(scopes, account=accounts[0])
    _ms_save_cache(session, cache)
    if result and "access_token" in result:
        return result["access_token"]
    return None


@app.get("/auth/login")
async def ms_login(request: Request):
    if not MS_CLIENT_ID or not MS_CLIENT_SECRET:
        raise HTTPException(
            status_code=500,
            detail="Microsoft app credentials not configured — set MS_CLIENT_ID / "
                   "MS_CLIENT_SECRET / MS_TENANT_ID (see SETUP-microsoft.md).",
        )
    cache = _ms_load_cache(request.state.session)
    # mode=calendar is for an employer mailbox (an advisor's wirehouse or
    # insurer account): invites only, no mail permission requested, because
    # the narrow ask is what such tenants actually approve.
    calendar_only = request.query_params.get("mode") == "calendar"
    scopes = MS_CAL_SCOPES if calendar_only else MS_SCOPES
    flow = _ms_app(cache).initiate_auth_code_flow(
        scopes, redirect_uri=BASE_URL + "/auth/callback"
    )
    request.state.session["ms_flow"] = flow
    request.state.session["ms_mode"] = "calendar" if calendar_only else "full"
    return RedirectResponse(flow["auth_uri"])


@app.get("/auth/callback")
async def ms_callback(request: Request):
    session = request.state.session
    flow = session.pop("ms_flow", None)
    if not flow:
        return RedirectResponse("/")
    cache = _ms_load_cache(session)
    try:
        result = _ms_app(cache).acquire_token_by_auth_code_flow(
            flow, dict(request.query_params)
        )
    except ValueError:
        return RedirectResponse("/")
    _ms_save_cache(session, cache)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result.get("error_description", "Sign-in failed"))
    if session.get("provider") != "password":
        session["provider"] = "microsoft"
        claims = result.get("id_token_claims") or {}
        email = (claims.get("preferred_username") or claims.get("email") or "").lower()
        if email:
            session["identity"] = email
    # Same deal as the Google callback: vault everything under the person so
    # a fresh browser next month starts with this connection already made.
    await _save_all_attachments(session)
    await _restore_attachments(session, _identity(session))
    return RedirectResponse("/")


# ------------------------- Google (OAuth 2.0) -------------------------

async def _vault_put(provider: str, email: str, refresh_token: str) -> None:
    """Store one user's refresh token, sealed with the same KMS envelope as
    sessions. Best-effort: without Firestore this is a no-op and sign-in
    degrades to the per-session behaviour it had before."""
    if not (email and refresh_token):
        return
    doc = _seal(refresh_token)
    doc["updated_at"] = time.time()
    await _fs_set(FS_VAULT, f"{provider}:{email.lower()}", doc)


async def _vault_get(provider: str, email: str) -> Optional[str]:
    if not email:
        return None
    doc = await _fs_get(FS_VAULT, f"{provider}:{email.lower()}")
    return _unseal(doc) if doc else None


async def _vault_del(provider: str, email: str) -> None:
    if email:
        await _fs_del(FS_VAULT, f"{provider}:{email.lower()}")


# ---------------------- attachments follow the person ----------------------
# Connecting an account is a fact about the user, not about the browser tab
# they happened to do it in. Every attachment (Google, Microsoft, ZoomInfo,
# the ZoomInfo MCP token) is sealed into the vault under the user's identity,
# and signing in — any browser, any day — rehydrates all of them. Without
# this, only Google survived a new cookie, and only for Google-primary users.

def _identity(session: dict) -> str:
    """The email a person's data is keyed under, if this session knows it.

    The stamped identity is the answer for anything created since §31. Older
    sessions predate it and last thirty days, so the Google link's own recorded
    address is asked for second — it is already on the session, and without it
    a signed-in user of a month-old cookie reads as nobody, which is how the
    per-user allowance came to be skippable by simply not signing in again.
    """
    if session.get("provider") == "password":
        return (session.get("account_email") or "").lower()
    stamped = (session.get("identity") or "").lower()
    if stamped:
        return stamped
    return ((session.get("google") or {}).get("email") or "").lower()


async def _save_attachment(identity: str, kind: str, data) -> None:
    if not identity:
        return
    doc = _seal(json.dumps(data))
    doc["updated_at"] = time.time()
    await _fs_set(FS_VAULT, f"att-{kind}:{identity.lower()}", doc)


async def _load_attachment(identity: str, kind: str):
    if not identity:
        return None
    doc = await _fs_get(FS_VAULT, f"att-{kind}:{identity.lower()}")
    if not doc:
        return None
    raw = _unseal(doc)
    if raw is None:
        return None
    try:
        return json.loads(raw)
    except ValueError:
        return None


async def _del_attachment(identity: str, kind: str) -> None:
    # Disconnect means disconnect: without this the vault would quietly
    # resurrect the connection at the next sign-in.
    if identity:
        await _fs_del(FS_VAULT, f"att-{kind}:{identity.lower()}")


async def _restore_attachments(session: dict, identity: str) -> None:
    """Rehydrate every vaulted connection into a fresh session.

    Only fills gaps: a connection already live in this session is newer than
    anything the vault holds, and must not be replaced by it.
    """
    if not identity:
        return
    session["identity"] = identity.lower()
    if not (session.get("google") or {}).get("refresh_token"):
        g = await _load_attachment(identity, "google")
        if g and g.get("refresh_token"):
            session["google"] = {"access_token": "", "refresh_token": g["refresh_token"],
                                 "expires_at": 0, "email": g.get("email", "")}
    if not session.get("ms_token_cache"):
        m = await _load_attachment(identity, "msal")
        if m and m.get("cache"):
            session["ms_token_cache"] = m["cache"]
            session["ms_mode"] = m.get("mode", "full")
    if not session.get("zoominfo"):
        z = await _load_attachment(identity, "zoominfo")
        if z and z.get("refresh_token"):
            session["zoominfo"] = dict(z, access_token="", expires_at=0)
    if not session.get("zi_mcp"):
        t = await _load_attachment(identity, "zimcp")
        if t and t.get("token"):
            session["zi_mcp"] = t


async def _save_all_attachments(session: dict) -> None:
    """Push this session's live connections into the vault, best-effort."""
    identity = _identity(session)
    if not identity:
        return
    g = session.get("google") or {}
    if g.get("refresh_token"):
        await _save_attachment(identity, "google",
                               {"refresh_token": g["refresh_token"],
                                "email": g.get("email", "")})
    if session.get("ms_token_cache"):
        await _save_attachment(identity, "msal",
                               {"cache": session["ms_token_cache"],
                                "mode": session.get("ms_mode", "full")})
    z = session.get("zoominfo") or {}
    if z.get("refresh_token"):
        await _save_attachment(identity, "zoominfo",
                               {k: v for k, v in z.items() if k != "access_token"})
    if session.get("zi_mcp"):
        await _save_attachment(identity, "zimcp", session["zi_mcp"])


async def _google_token(session: dict) -> Optional[str]:
    g = session.get("google")
    if not g:
        return None
    if time.time() < g.get("expires_at", 0):
        return g["access_token"]
    if not g.get("refresh_token"):
        session.pop("google", None)
        return None
    async with httpx.AsyncClient(timeout=20) as cx:
        r = await cx.post(GOOGLE_TOKEN_URL, data={
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "refresh_token": g["refresh_token"],
            "grant_type": "refresh_token",
        })
    if r.status_code != 200:
        # Revoked or expired. Clear the vault too: a dead token left there
        # would be handed to the next sign-in, which would fail the same way
        # instead of showing the one consent screen that fixes it.
        await _vault_del("google", (g or {}).get("email") or "")
        await _del_attachment(_identity(session), "google")
        session.pop("google", None)
        return None
    tok = r.json()
    g["access_token"] = tok["access_token"]
    g["expires_at"] = time.time() + tok.get("expires_in", 3600) - 60
    return g["access_token"]


@app.get("/auth/google/login")
async def google_login(request: Request):
    if not GOOGLE_CLIENT_ID or not GOOGLE_CLIENT_SECRET:
        raise HTTPException(
            status_code=500,
            detail="Google app credentials not configured — set GOOGLE_CLIENT_ID / "
                   "GOOGLE_CLIENT_SECRET (see SETUP-google.md).",
        )
    state = secrets.token_urlsafe(24)
    session = request.state.session
    session["g_state"] = state
    # Google issues a refresh token only on a grant that shows the consent
    # screen. Asking for consent every time guarantees one — and makes the user
    # approve Drive, Calendar and Gmail on every single sign-in, which is what
    # this used to do. They read the same screen so often it stops being a
    # decision, which is the opposite of what a consent screen is for.
    #
    # So consent is asked for once. If the grant comes back without a refresh
    # token and we do not already hold one, the callback sends them round once
    # more with force=1, and only that trip shows the screen.
    force = request.query_params.get("force") == "1"
    if force:
        session["g_forced"] = True      # so the callback cannot bounce twice
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": BASE_URL + "/auth/google/callback",
        "response_type": "code",
        "scope": GOOGLE_SCOPES,
        "access_type": "offline",   # refresh token, so sign-in survives the hour
        # select_account still lets someone switch identity — the thing a
        # sign-in button is actually for — without re-approving every scope.
        # Consent is never requested on the first pass: at this point we do
        # not know who the user is, so we cannot know whether they already
        # granted it. The callback decides, with the vault in hand.
        "prompt": "consent" if force else "select_account",
        "include_granted_scopes": "true",
        "state": state,
    }
    return RedirectResponse(f"{GOOGLE_AUTH_URL}?{urlencode(params)}")


@app.get("/auth/google/callback")
async def google_callback(request: Request):
    session = request.state.session
    if request.query_params.get("state") != session.pop("g_state", None):
        return RedirectResponse("/")
    # Guard against a loop: if the forced trip also comes back without a
    # refresh token, accept the hour-long session rather than bouncing forever.
    already_forced = session.pop("g_forced", False)
    code = request.query_params.get("code")
    if not code:
        return RedirectResponse("/")
    async with httpx.AsyncClient(timeout=20) as cx:
        r = await cx.post(GOOGLE_TOKEN_URL, data={
            "code": code,
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "redirect_uri": BASE_URL + "/auth/google/callback",
            "grant_type": "authorization_code",
        })
    if r.status_code != 200:
        raise HTTPException(status_code=400, detail=f"Google sign-in failed: {r.text[:200]}")
    tok = r.json()
    # Who just signed in? The vault is keyed by the person, so a returning
    # user on a brand-new browser still finds the grant they already made.
    email = ""
    try:
        async with httpx.AsyncClient(timeout=20) as cx:
            u = await cx.get(GOOGLE_USERINFO_URL,
                             headers={"Authorization": f"Bearer {tok['access_token']}"})
        if u.status_code == 200:
            email = (u.json().get("email") or "").lower()
    except httpx.HTTPError:
        pass
    # A re-authorisation returns no refresh token, because the existing one is
    # still valid. Writing tok.get("refresh_token") straight in would replace a
    # working token with None and silently end the session an hour later — a
    # bug that could not bite while consent was forced on every sign-in, and
    # would have started biting the moment it stopped.
    keep = (session.get("google") or {}).get("refresh_token")
    refresh = tok.get("refresh_token") or keep
    if not refresh:
        # The cookie is new but the person may not be: the consent they gave
        # last month is in the vault. This lookup is the difference between
        # "approve Drive, Calendar and Gmail once" and "approve them on every
        # sign-in", which is what was actually happening — the old check
        # looked in the session, and a fresh sign-in never has one.
        refresh = await _vault_get("google", email)
    if tok.get("refresh_token"):
        await _vault_put("google", email, tok["refresh_token"])
    session["google"] = {
        "access_token": tok["access_token"],
        "refresh_token": refresh,
        "expires_at": time.time() + tok.get("expires_in", 3600) - 60,
        "email": email,
    }
    if session.get("provider") != "password":
        session["provider"] = "google"
        session["identity"] = email
    # Vault the link under the person, and pull back whatever else they have
    # connected before — signing in with Google on a fresh browser must bring
    # the Microsoft and ZoomInfo attachments with it.
    await _save_all_attachments(session)
    await _restore_attachments(session, _identity(session))
    # No refresh token anywhere — grant, session, vault. This person has
    # genuinely never granted offline access, so ask for consent exactly
    # once, here, where we finally know it is needed.
    if not refresh and not already_forced:
        return RedirectResponse("/auth/google/login?force=1")
    return RedirectResponse("/")


# ------------------------- ZoomInfo (per-user OAuth) -------------------------
# Deliberately separate from app sign-in: a user signs in with Google or
# Microsoft to use the app, then connects ZoomInfo on top. One app account can
# therefore exist without a ZoomInfo seat, and the seat that is connected is
# unambiguously that person's own.

def _zi_configured() -> bool:
    return bool(ZI_CLIENT_ID and ZI_CLIENT_SECRET)


async def _zi_token(session: dict) -> Optional[str]:
    """This user's ZoomInfo access token, refreshed if it has aged out."""
    z = session.get("zoominfo")
    if not z:
        return None
    if time.time() < z.get("expires_at", 0):
        return z["access_token"]
    if not z.get("refresh_token"):
        session.pop("zoominfo", None)     # expired and unrenewable: reconnect
        await _del_attachment(_identity(session), "zoominfo")
        return None
    async with httpx.AsyncClient(timeout=20) as cx:
        r = await cx.post(ZI_TOKEN_URL, data={
            "grant_type": "refresh_token",
            "refresh_token": z["refresh_token"],
            "client_id": ZI_CLIENT_ID,
            "client_secret": ZI_CLIENT_SECRET,
        })
    if r.status_code != 200:
        session.pop("zoominfo", None)
        await _del_attachment(_identity(session), "zoominfo")
        return None
    tok = r.json()
    z["access_token"] = tok["access_token"]
    z["expires_at"] = time.time() + tok.get("expires_in", 3600) - 60
    if tok.get("refresh_token"):
        z["refresh_token"] = tok["refresh_token"]
    session["zoominfo"] = z
    # The refresh may have rotated the refresh token; re-vault so the next
    # fresh sign-in gets the live one, not a revoked ancestor.
    await _save_all_attachments(session)
    return z["access_token"]


@app.get("/auth/zoominfo/login")
async def zi_login(request: Request):
    # App sign-in first. Without this a direct hit on the URL would park a
    # ZoomInfo token on an anonymous session — the same leak that namespacing
    # local storage closed for lead lists, and worse: this one is a billable seat.
    try:
        await _active_token(request)
    except HTTPException:
        return RedirectResponse("/?zi=signin")
    if not _zi_configured():
        raise HTTPException(
            status_code=500,
            detail="ZoomInfo app credentials not configured — set ZI_CLIENT_ID / "
                   "ZI_CLIENT_SECRET (see SETUP-zoominfo.md).",
        )
    # PKCE throughout. A Standard app may use it and a Partner app must, so the
    # same flow survives a later move to multi-org distribution.
    verifier = secrets.token_urlsafe(64)[:96]
    challenge = base64.urlsafe_b64encode(
        hashlib.sha256(verifier.encode()).digest()
    ).decode().rstrip("=")
    state = secrets.token_urlsafe(24)
    request.state.session["zi_state"] = state
    request.state.session["zi_verifier"] = verifier
    params = {
        "client_id": ZI_CLIENT_ID,
        "redirect_uri": BASE_URL + "/auth/zoominfo/callback",
        "response_type": "code",
        "scope": ZI_SCOPES,
        "state": state,
        "code_challenge": challenge,
        "code_challenge_method": "S256",
    }
    return RedirectResponse(f"{ZI_AUTH_URL}?{urlencode(params)}")


@app.get("/auth/zoominfo/callback")
async def zi_callback(request: Request):
    try:
        await _active_token(request)
    except HTTPException:
        return RedirectResponse("/?zi=signin")
    session = request.state.session
    state = session.pop("zi_state", None)
    verifier = session.pop("zi_verifier", None)
    if not state or request.query_params.get("state") != state:
        return RedirectResponse("/?zi=state")
    code = request.query_params.get("code")
    if not code:
        return RedirectResponse("/?zi=denied")
    async with httpx.AsyncClient(timeout=20) as cx:
        r = await cx.post(ZI_TOKEN_URL, data={
            "grant_type": "authorization_code",
            "code": code,
            "client_id": ZI_CLIENT_ID,
            "client_secret": ZI_CLIENT_SECRET,
            "redirect_uri": BASE_URL + "/auth/zoominfo/callback",
            "code_verifier": verifier or "",
        })
    if r.status_code != 200:
        # Surfaced rather than swallowed: a token-endpoint mismatch is the most
        # likely first failure here, and a silent redirect hides it.
        raise HTTPException(status_code=400,
                            detail=f"ZoomInfo sign-in failed ({r.status_code}): {r.text[:300]}")
    tok = r.json()
    session["zoominfo"] = {
        "access_token": tok["access_token"],
        "refresh_token": tok.get("refresh_token"),
        "expires_at": time.time() + tok.get("expires_in", 3600) - 60,
        "connected_at": int(time.time()),
    }
    await _save_all_attachments(session)
    return RedirectResponse("/?zi=ok")


@app.get("/auth/zoominfo/disconnect")
async def zi_disconnect(request: Request):
    request.state.session.pop("zoominfo", None)
    await _del_attachment(_identity(request.state.session), "zoominfo")
    return RedirectResponse("/")


@app.get("/auth/logout")
async def logout(request: Request):
    await drop_session(request.state.sid)
    request.state.session_dropped = True     # stop the middleware re-saving it
    response = RedirectResponse("/")
    response.delete_cookie(SESSION_COOKIE)
    return response


# ------------------------- Shared helpers -------------------------

_MEM_ACCOUNTS: dict[str, dict] = {}


def _hash_password(password: str, salt: bytes = None) -> tuple[str, str]:
    """scrypt from the standard library — no new dependency, memory-hard, and
    the parameters are the library documentation's interactive-login set."""
    salt = salt or secrets.token_bytes(16)
    h = hashlib.scrypt(password.encode("utf-8"), salt=salt, n=2**14, r=8, p=1, dklen=32)
    return base64.b64encode(salt).decode(), base64.b64encode(h).decode()


def _password_ok(password: str, salt_b64: str, hash_b64: str) -> bool:
    import hmac
    try:
        salt = base64.b64decode(salt_b64)
        want = base64.b64decode(hash_b64)
    except Exception:
        return False
    _, got = _hash_password(password, salt)
    return hmac.compare_digest(base64.b64decode(got), want)


async def _account_get(email: str) -> Optional[dict]:
    doc = await _fs_get(FS_ACCOUNTS, email)
    return doc if doc is not None else _MEM_ACCOUNTS.get(email)


async def _account_put(email: str, doc: dict) -> None:
    if not await _fs_set(FS_ACCOUNTS, email, doc):
        _MEM_ACCOUNTS[email] = doc


class AccountRequest(BaseModel):
    email: EmailStr
    password: str


@app.post("/auth/signup")
async def signup(req: AccountRequest, request: Request):
    """An account is an email and a password — any email.

    No verification mail goes out yet, because the server has no mailbox of
    its own to send from; the gap is documented rather than papered over.
    What an unverified account gets is only its own empty workspace, so the
    thing verification protects is the email's owner being impersonated in
    shares — worth closing before strangers are invited, said out loud here.
    """
    email = str(req.email).lower().strip()
    if len(req.password) < 10:
        raise HTTPException(status_code=400,
                            detail="Use at least 10 characters — a short phrase beats a complex word.")
    if await _account_get(email):
        raise HTTPException(status_code=409,
                            detail="That email already has an account — sign in instead.")
    salt, h = _hash_password(req.password)
    await _account_put(email, {"salt": salt, "hash": h, "created_at": time.time()})
    session = request.state.session
    session["provider"] = "password"
    session["account_email"] = email
    session["identity"] = email
    return {"ok": True, "email": email}


@app.post("/auth/password-login")
async def password_login(req: AccountRequest, request: Request):
    email = str(req.email).lower().strip()
    acct = await _account_get(email)
    # One message for both failures: which half was wrong is exactly what a
    # guessing attacker wants confirmed.
    if not acct or not _password_ok(req.password, acct.get("salt", ""), acct.get("hash", "")):
        raise HTTPException(status_code=401, detail="Wrong email or password.")
    session = request.state.session
    session["provider"] = "password"
    session["account_email"] = email
    # The whole point of an account: everything they connected comes back.
    await _restore_attachments(session, email)
    return {"ok": True, "email": email}


async def _signed_in_email(request: Request) -> str:
    """Email of the signed-in user — the key their saved list is stored under.

    For a password account this is the account email, deliberately even after
    a Google or Microsoft address is linked: linking a sender must never move
    someone's leads to a different key.
    """
    session = request.state.session
    if session.get("provider") == "password" and session.get("account_email"):
        return session["account_email"]
    if session.get("provider") == "google":
        token = await _google_token(session)
        if token:
            async with httpx.AsyncClient(timeout=15) as cx:
                r = await cx.get(GOOGLE_USERINFO_URL, headers={"Authorization": f"Bearer {token}"})
            if r.status_code == 200:
                return (r.json().get("email") or "").lower()
    token = _ms_token(session)
    if token:
        async with httpx.AsyncClient(timeout=15) as cx:
            r = await cx.get(f"{GRAPH}/me", headers={"Authorization": f"Bearer {token}"})
        if r.status_code == 200:
            d = r.json()
            return ((d.get("mail") or d.get("userPrincipalName")) or "").lower()
    return ""


async def _active_token(request: Request) -> tuple[str, str]:
    """Return (provider, access_token) for the signed-in account, else 401.

    A password session is signed in with no provider token at all; callers
    that only use this as a gate work unchanged, and the ones that need a
    real Google or Microsoft token already handle its absence with their own
    sentence.
    """
    session = request.state.session
    if session.get("provider") == "password" and session.get("account_email"):
        token = await _google_token(session)
        if token:
            return "google", token
        token = _ms_token(session)
        if token:
            return "microsoft", token
        return "password", ""
    if session.get("provider") == "google":
        token = await _google_token(session)
        if token:
            return "google", token
    token = _ms_token(session)
    if token:
        return "microsoft", token
    raise HTTPException(status_code=401, detail="Not signed in")


def _gmail_raw(to: str, subject: str, body: str, sender: str = "") -> str:
    msg = EmailMessage()
    msg["To"] = to
    msg["Subject"] = subject
    # Gmail rejects a From it has not been told this account may send as, so
    # this is only ever set to an address /api/senders read back from Gmail.
    if sender:
        msg["From"] = sender
    msg.set_content(body)
    return base64.urlsafe_b64encode(msg.as_bytes()).decode()


# ------------------------- who the mail goes out as -------------------------
# One person, several work addresses: a personal domain on Google, an employer
# mailbox on Microsoft, an alias or shared box hanging off either. All of them
# are the user's own accounts — the app never sends as anyone who has not
# signed in here — and which one to use is a per-message decision, because the
# right From for a prospect at their old employer is not the right one for a
# calendar invite from the advisory firm.

async def _google_senders(session: dict) -> list[dict]:
    token = await _google_token(session)
    if not token:
        return []
    headers = {"Authorization": f"Bearer {token}"}
    primary = ""
    async with httpx.AsyncClient(timeout=15) as cx:
        r = await cx.get(GOOGLE_USERINFO_URL, headers=headers)
        if r.status_code == 200:
            primary = (r.json().get("email") or "").lower()
        # Aliases need a scope granted after some accounts signed in, so a 403
        # here is ordinary and means "primary only", not an error.
        try:
            a = await cx.get(GMAIL_SENDAS_URL, headers=headers)
            rows = a.json().get("sendAs", []) if a.status_code == 200 else []
        except Exception:
            rows = []
    out, seen = [], set()
    for row in rows:
        addr = (row.get("sendAsEmail") or "").lower()
        if not addr or addr in seen:
            continue
        # An unverified alias is one Gmail will refuse at send time.
        if not row.get("isPrimary") and row.get("verificationStatus") not in ("accepted", None, ""):
            continue
        seen.add(addr)
        out.append({"id": "google:" + addr, "provider": "google", "address": addr,
                    "name": row.get("displayName") or "",
                    "primary": bool(row.get("isPrimary")),
                    "kind": "primary" if row.get("isPrimary") else "alias",
                    # An alias can front an email, but an invite is owned by a
                    # calendar and an alias does not have one.
                    "mail": True, "calendar": bool(row.get("isPrimary"))})
    if primary and primary not in seen:
        out.insert(0, {"id": "google:" + primary, "provider": "google", "address": primary,
                       "name": "", "primary": True, "kind": "primary",
                       "mail": True, "calendar": True})
    return out


async def _ms_senders(session: dict) -> list[dict]:
    token = _ms_token(session)
    if not token:
        return []
    async with httpx.AsyncClient(timeout=15) as cx:
        r = await cx.get(f"{GRAPH}/me", headers={"Authorization": f"Bearer {token}"})
    if r.status_code != 200:
        return []
    d = r.json()
    addr = ((d.get("mail") or d.get("userPrincipalName")) or "").lower()
    if not addr:
        return []
    return [{"id": "microsoft:" + addr, "provider": "microsoft", "address": addr,
             "name": d.get("displayName") or "", "primary": True, "kind": "primary",
             # A calendar-only employer connection can invite but not mail;
             # the client uses this to keep the address out of the email
             # dialog rather than let a send fail at Graph.
             "mail": _ms_mail_ok(session), "calendar": True}]


async def _senders(request: Request) -> list[dict]:
    session = request.state.session
    out = await _google_senders(session)
    out += await _ms_senders(session)
    return out


async def _sender_token(request: Request, sender_id: str) -> tuple[str, str, str]:
    """(provider, token, from_address) for a chosen sender, or the default one.

    An unknown id is refused rather than quietly falling back: sending from the
    wrong address is the kind of mistake that is only noticed by the recipient.
    """
    session = request.state.session
    if not sender_id:
        provider, token = await _active_token(request)
        if provider == "password" and not token:
            raise HTTPException(status_code=400,
                                detail="No sending address yet — link a Google or Microsoft "
                                       "account from the bar at the top, then choose it here.")
        return provider, token, ""
    available = await _senders(request)
    match = next((s for s in available if s["id"] == sender_id), None)
    if not match:
        raise HTTPException(
            status_code=400,
            detail=f"{sender_id.split(':', 1)[-1]} is not one of your connected sending "
                   f"addresses. Connect that account, or pick another address.")
    if match["provider"] == "google":
        token = await _google_token(session)
        if not token:
            raise HTTPException(status_code=401, detail="Google account is not connected.")
        return "google", token, ("" if match["primary"] else match["address"])
    token = _ms_token(session)
    if not token:
        raise HTTPException(status_code=401, detail="Microsoft account is not connected.")
    return "microsoft", token, ""


@app.get("/api/senders")
async def senders(request: Request):
    await _active_token(request)
    out = await _senders(request)
    return {"senders": out, "default": (out[0]["id"] if out else "")}


# ------------------------- API routes -------------------------

@app.get("/api/me")
async def me(request: Request):
    session = request.state.session
    providers = {
        "microsoft": bool(MS_CLIENT_ID and MS_CLIENT_SECRET),
        "google": bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET),
    }
    features = {
        "whitepages": bool(WHITEPAGES_API_KEY),
        "ai_qc": bool(ANTHROPIC_API_KEY),
        "server_state": storage_backend() == "firestore",
        "zoominfo": _zi_configured(),
        "edgar": bool(EDGAR_USER_AGENT and ANTHROPIC_API_KEY),
        # Free public-record lookups. FEC ships with the demo key, so the only
        # half that can be dark is the SEC one, and it says so per-source.
        "free_sources": True,
        # So the sweep can be honest about pace before it starts: the shared
        # demo key allows 40 FEC lookups an hour across everyone using it.
        "fec_personal_key": FEC_API_KEY != "DEMO_KEY",
        "zi_mcp": bool(ANTHROPIC_API_KEY),
        "opportunities": bool((await _warn_feeds_value())[0].strip()),
        # A private company's About page, read on the publisher's terms.
        "site_reader": bool(HARVEST_USER_AGENT),
        # Per-lead web research through the Claude API's search tool — a
        # licensed API, never scraping, and never run as a sweep.
        "web_research": bool(ANTHROPIC_API_KEY),
        # TRIAL: Attom property data. Lives and dies with the env var — when
        # the 10-day trial key is removed, the buttons disappear and the
        # already-fetched answers stay on their leads.
        "attom": bool(ATTOM_API_KEY),
        "drive": False,          # set per-session below when Google is signed in
    }
    encryption = encryption_backend()
    storage = storage_backend()
    credits = await _credit_block(_identity(session))
    if session.get("provider") == "password" and session.get("account_email"):
        email = session["account_email"]
        # Linked providers light their features up exactly as a native
        # sign-in would; the identity just stays the account's own.
        if session.get("google") and await _google_token(session):
            features["drive"] = True
        return {"signed_in": True, "provider": "password",
                "name": email.split("@")[0], "email": email,
                "providers": providers, "features": features,
                "linked_google": bool(session.get("google")),
                "linked_microsoft": bool(session.get("ms_token_cache")),
                "zi_connected": bool(session.get("zoominfo")),
                "zi_mcp_connected": bool(_zi_mcp_token(session)),
                "storage": storage, "encryption": encryption, "credits": credits,
                "is_admin": _is_admin(email)}
    if session.get("provider") == "google":
        token = await _google_token(session)
        if token:
            async with httpx.AsyncClient(timeout=15) as cx:
                r = await cx.get(GOOGLE_USERINFO_URL, headers={"Authorization": f"Bearer {token}"})
            if r.status_code == 200:
                d = r.json()
                features["drive"] = True
                return {"signed_in": True, "provider": "google",
                        "name": d.get("name"), "email": d.get("email"),
                        "providers": providers, "features": features,
                        "zi_connected": bool(session.get("zoominfo")),
                        "zi_mcp_connected": bool(_zi_mcp_token(session)),
                        "storage": storage, "encryption": encryption, "credits": credits,
                        "is_admin": _is_admin(d.get("email") or "")}
    token = _ms_token(session)
    if token:
        async with httpx.AsyncClient(timeout=15) as cx:
            r = await cx.get(f"{GRAPH}/me", headers={"Authorization": f"Bearer {token}"})
        if r.status_code == 200:
            d = r.json()
            return {"signed_in": True, "provider": "microsoft",
                    "name": d.get("displayName"),
                    "email": d.get("mail") or d.get("userPrincipalName"),
                    "providers": providers, "features": features,
                    # Which accounts actually have tokens on this session, as
                    # opposed to which the deployment is configured for.
                    "linked_google": bool(session.get("google")),
                    "linked_microsoft": bool(session.get("ms_token_cache")),
                    "zi_connected": bool(session.get("zoominfo")),
                    "storage": storage, "encryption": encryption, "credits": credits,
                    "is_admin": _is_admin(d.get("mail") or d.get("userPrincipalName") or "")}
    return JSONResponse({"signed_in": False, "providers": providers,
                         "features": features, "storage": storage,
                         "encryption": encryption})


class EmailRequest(BaseModel):
    to: EmailStr
    subject: str
    body: str
    sender: str = ""          # a /api/senders id; empty means the signed-in account


@app.post("/api/send-email")
async def send_email(req: EmailRequest, request: Request):
    provider, token, from_addr = await _sender_token(request, req.sender)
    if provider == "microsoft" and not _ms_mail_ok(request.state.session):
        raise HTTPException(
            status_code=400,
            detail="This Microsoft account is connected for calendar invites only — "
                   "its tenant was asked for no mail permission. Send the email from "
                   "a Gmail address (an alias of this address works), or reconnect "
                   "the account without calendar-only mode.")
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=30) as cx:
        if provider == "google":
            r = await cx.post(GMAIL_SEND_URL,
                              json={"raw": _gmail_raw(str(req.to), req.subject, req.body, from_addr)},
                              headers=headers)
            ok = r.status_code == 200
        else:
            payload = {
                "message": {
                    "subject": req.subject,
                    "body": {"contentType": "Text", "content": req.body},
                    "toRecipients": [{"emailAddress": {"address": str(req.to)}}],
                },
                "saveToSentItems": True,
            }
            r = await cx.post(f"{GRAPH}/me/sendMail", json=payload, headers=headers)
            ok = r.status_code in (200, 202)
    if not ok:
        raise HTTPException(status_code=502, detail=f"{provider} send error {r.status_code}: {r.text[:300]}")
    return {"ok": True, "provider": provider, "from": from_addr}


class EventRequest(BaseModel):
    attendee: EmailStr
    subject: str
    body: str = ""
    start: str          # "YYYY-MM-DDTHH:MM:SS" local wall time
    end: str
    timezone: str = "America/New_York"
    sender: str = ""    # which connected calendar the invite comes from


@app.post("/api/create-event")
async def create_event(req: EventRequest, request: Request):
    # A calendar invite is owned by the calendar it is created on, so unlike
    # mail there is no alias to set: picking a sender picks whose calendar it
    # lands in and therefore what address the attendee sees.
    provider, token, _from = await _sender_token(request, req.sender)
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=30) as cx:
        if provider == "google":
            payload = {
                "summary": req.subject,
                "description": req.body,
                "start": {"dateTime": req.start, "timeZone": req.timezone},
                "end": {"dateTime": req.end, "timeZone": req.timezone},
                "attendees": [{"email": str(req.attendee)}],
            }
            r = await cx.post(f"{GCAL_EVENTS_URL}?sendUpdates=all", json=payload, headers=headers)
            ok = r.status_code == 200
        else:
            payload = {
                "subject": req.subject,
                "body": {"contentType": "Text", "content": req.body},
                "start": {"dateTime": req.start, "timeZone": req.timezone},
                "end": {"dateTime": req.end, "timeZone": req.timezone},
                "attendees": [
                    {"emailAddress": {"address": str(req.attendee)}, "type": "required"}
                ],
            }
            r = await cx.post(f"{GRAPH}/me/events", json=payload, headers=headers)
            ok = r.status_code == 201
    if not ok:
        raise HTTPException(status_code=502, detail=f"{provider} calendar error {r.status_code}: {r.text[:300]}")
    return {"ok": True, "provider": provider}


class VerifyPhoneRequest(BaseModel):
    phone: str
    first_name: str = ""
    last_name: str = ""
    # The lead's address, so a different owner at the same address can be told
    # apart from a wrong number. One is a spouse; the other is bad data.
    street: str = ""
    city: str = ""
    state: str = ""
    zip: str = ""


# ---------------------------------------------------------------------------
# Spending credits on purpose
#
# What is billed, from the response-code table on the endpoint reference:
#
#   200 OK   billable      404 by id  billable
#   400/403  NOT billable  429/5xx    NOT billable
#
# Note what that means: **a 200 with zero results is billed**. "No such person"
# costs exactly what finding them costs. The expensive mistake here is not the
# malformed query — that one is free — it is the well-formed query that was
# never going to identify anybody, and the well-formed query asked twice.
#
# (An earlier draft of this module was written against the integration guide's
# looser wording, "successful (2xx) and client-error (4xx) responses are
# billed", and justified the validator below as a way to avoid paying for 400s.
# The per-endpoint table is more specific and says otherwise. The validator
# stays — a request that cannot succeed should fail instantly with a reason a
# person can act on, rather than after a round trip — but it is worth being
# clear that it buys clarity and latency, not credits.)
#
# The two things that genuinely save money:
#
#   1. Never ask a question that cannot identify anyone. A surname with no
#      location returns a stranger, at full price.
#   2. Never ask the same question twice. A person's record does not change
#      between two clicks, so the answer is cached on the exact query that
#      produced it — including "nobody", which was paid for like anything else.
# ---------------------------------------------------------------------------

# Documented on the endpoint. Kept as the API states them rather than loosened,
# because a value this rejects is a value the API charges for rejecting.
WP_PHONE_RE = re.compile(r"^(\+?1[-.\s]?)?\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})$")
WP_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
WP_ZIP_RE = re.compile(r"^\d{5}$")
WP_STATES = {
    "AL", "AK", "AZ", "AR", "AS", "CA", "CO", "CT", "DE", "DC", "FL", "GA",
    "GU", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA",
    "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC",
    "ND", "MP", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX",
    "TT", "UT", "VT", "VA", "VI", "WA", "WV", "WI", "WY",
}
# Any one of these is enough to make the query a search rather than a listing.
WP_SEARCH_KEYS = ("phone", "email", "name", "first_name", "middle_name",
                  "last_name", "street", "city", "state_code", "zipcode")


class WPRefused(Exception):
    """A request that cannot succeed. Never sent.

    A 400 is not billed, so this is not primarily about money — it is about
    failing in a hundredth of a second with a reason the user can act on,
    instead of after a round trip with one they cannot.
    """


def wp_validate(params: dict) -> dict:
    """Params as they will be sent, or WPRefused with the reason.

    Empty values are dropped rather than sent: an empty state_code is not a
    wildcard, it is a 400.
    """
    p = {k: v for k, v in params.items()
         if v is not None and str(v).strip() != ""}

    if "phone" in p and not WP_PHONE_RE.match(str(p["phone"]).strip()):
        raise WPRefused(f"{p['phone']!r} is not a phone number the API accepts")
    if "email" in p and not WP_EMAIL_RE.match(str(p["email"]).strip()):
        raise WPRefused(f"{p['email']!r} is not an email address the API accepts")
    if "zipcode" in p and not WP_ZIP_RE.match(str(p["zipcode"]).strip()):
        raise WPRefused(f"{p['zipcode']!r} is not a five-digit ZIP")
    if "state_code" in p and str(p["state_code"]).strip().upper() not in WP_STATES:
        raise WPRefused(f"{p['state_code']!r} is not a two-letter state code")

    # "first_name / middle_name / last_name cannot be used with 'name'."
    if "name" in p and any(k in p for k in ("first_name", "middle_name", "last_name")):
        raise WPRefused("name cannot be combined with first_name/last_name")
    # "Setting both on the same request returns HTTP 400."
    if _truthy_param(p.get("strict_match")) and _truthy_param(p.get("include_fuzzy_matching")):
        raise WPRefused("strict_match and include_fuzzy_matching contradict each other")

    for key, lo, hi in (("min_age", 18, 65), ("max_age", 18, 65),
                        ("page", 1, 10), ("page_size", 1, 15)):
        if key in p:
            try:
                v = int(p[key])
            except (TypeError, ValueError):
                raise WPRefused(f"{key} must be a whole number")
            if not lo <= v <= hi:
                raise WPRefused(f"{key} must be between {lo} and {hi}, got {v}")
            p[key] = v
    if "min_age" in p and "max_age" in p and p["min_age"] > p["max_age"]:
        raise WPRefused("min_age is above max_age")
    if "radius" in p:
        try:
            r = float(p["radius"])
        except (TypeError, ValueError):
            raise WPRefused("radius must be a number")
        if not 0.1 <= r <= 100:
            raise WPRefused(f"radius must be between 0.1 and 100, got {r}")

    if not any(k in p for k in WP_SEARCH_KEYS):
        raise WPRefused("nothing to search on")
    return p


def _truthy_param(v) -> bool:
    return str(v).strip().lower() in ("true", "1", "yes")


# How long an answer stands. A person's record does not change between two
# clicks, or between this morning and this afternoon. Long by default because
# the scarce resource is credits, not freshness, and any lookup can be forced.
WP_TTL = int(os.environ.get("WHITEPAGES_CACHE_SECONDS", str(30 * 86400)))
_WP_CACHE: dict = {}
# What has actually been spent, so the number shown to the user is counted
# rather than estimated.
WP_SPEND = {"calls": 0, "served_from_cache": 0, "refused": 0}


def _wp_key(kind: str, params: dict) -> str:
    return kind + "|" + "&".join(f"{k}={params[k]}" for k in sorted(params))


async def _wp_cached(key: str):
    """A previous answer to this exact question, from memory or from Firestore.

    Returns (found, value). `found` is separate from `value` because None is a
    real answer here — "WhitePages has no record of this person" is worth a
    credit and worth remembering, and re-asking it is the most wasteful lookup
    the app can make.
    """
    hit = _WP_CACHE.get(key)
    if hit and time.time() - hit[0] < WP_TTL:
        return True, hit[1]
    doc = await _fs_get(FS_CACHE, "wp:" + hashlib.sha256(key.encode()).hexdigest())
    if not doc or time.time() - float(doc.get("at") or 0) >= WP_TTL:
        return False, None
    raw = _unseal(doc)
    if raw is None:
        return False, None
    try:
        value = json.loads(raw)
    except ValueError:
        return False, None
    _WP_CACHE[key] = (float(doc["at"]), value)          # warm this instance too
    return True, value


async def _wp_remember(key: str, value) -> None:
    _WP_CACHE[key] = (time.time(), value)
    doc = _seal(json.dumps(value))
    doc["at"] = time.time()
    await _fs_set(FS_CACHE, "wp:" + hashlib.sha256(key.encode()).hexdigest(), doc)


def _wp_path(kind: str) -> str:
    """Endpoint path for a lookup, per API flavour.

    There is no /v2/phone on the Pro API. "Reverse phone lookup" is documented
    as a mode of person search — GET /v2/person?phone= — and the responses
    prove it: they come back as person records tagged matched_by:["phone"].
    Calling /v2/phone 404s, which the app reported as "no record found" while
    the account showed no usage at all, because nothing was ever billed.
    """
    override = {"phone": WHITEPAGES_PHONE_PATH, "person": WHITEPAGES_PERSON_PATH,
                "property": WHITEPAGES_PROPERTY_PATH}[kind]
    if override:
        return "/" + override.lstrip("/")
    if "trestle" in WHITEPAGES_BASE_URL:
        return {"phone": "/3.1/phone", "person": "/3.1/person", "property": "/3.1/property"}[kind]
    if kind == "property":
        return "/v2/property/"     # documented with the trailing slash
    return "/v2/person"            # both person search and reverse phone


async def signed_in(request: Request) -> str:
    """The signed-in address, or a 401. Declared, rather than re-checked.

    Twenty routes opened with the same three lines. As a dependency the
    requirement moves into the signature, where it is visible without reading
    the body, and a route cannot forget it by being written in a hurry.
    """
    email = await _signed_in_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="Not signed in")
    return email


async def _who(request: Request) -> str:
    """Who is spending, without paying a round trip to find out.

    The session already knows: a password account carries its own address, and
    an OAuth sign-in stamps one at the callback. _signed_in_email is the
    fallback for sessions predating that, and it costs a userinfo call — which
    is why it is the fallback and not the answer.
    """
    who = _identity(request.state.session)
    return who or await _signed_in_email(request)


async def _wp_phone(digits: str, fresh: bool = False, who: str = "") -> tuple:
    """(person, line) for a number. The single door to a reverse-phone query.

    Both endpoints came here by different routes and built the same query by
    hand, which meant the cache saw two spellings of one question. Now there is
    one spelling.
    """
    return _phone_owner(await _wp_get("phone", {"phone": digits}, fresh=fresh, who=who), digits)


async def _wp_get(kind: str, params: dict, fresh: bool = False, who: str = ""):
    """One WhitePages call, or none at all.

    Three things stand between a caller and a charge:

    1. The parameters are checked against the documented constraints. A request
       that cannot succeed is refused here for nothing, rather than there for
       the price of a 400.
    2. An identical question already asked is answered from memory.
    3. Only then does anything go out.

    Returns parsed JSON, or None when nothing matched.
    """
    if not WHITEPAGES_API_KEY:
        raise HTTPException(
            status_code=500,
            detail="WhitePages not configured — set WHITEPAGES_API_KEY "
                   "(see SETUP-whitepages.md).",
        )
    try:
        params = wp_validate(params)
    except WPRefused as e:
        WP_SPEND["refused"] += 1
        raise HTTPException(
            status_code=400,
            detail=f"Not sent — {e}. (A rejected query is not billed either way; "
                   f"this just tells you sooner.)",
        )

    key = _wp_key(kind, params)
    if not fresh:
        found, value = await _wp_cached(key)
        if found:
            WP_SPEND["served_from_cache"] += 1
            return value
        # Single-flight: two identical questions in the air at once — a
        # double-click, or two advisors on the same lead — both miss the
        # cache and both bill. The second caller waits for the first answer
        # instead of buying its own copy.
        if key in _WP_INFLIGHT:
            value = await asyncio.shield(_WP_INFLIGHT[key])
            WP_SPEND["served_from_cache"] += 1
            return value
        fut = asyncio.get_running_loop().create_future()
        _WP_INFLIGHT[key] = fut
        try:
            value = await _wp_fetch(kind, params, key, who)
            if not fut.done():
                fut.set_result(value)
            return value
        except BaseException as e:
            # A failure is not an answer: waiters re-raise rather than being
            # handed a corpse, and the key is freed for the next attempt.
            if not fut.done():
                fut.set_exception(e)
            raise
        finally:
            _WP_INFLIGHT.pop(key, None)
    return await _wp_fetch(kind, params, key, who)


_WP_INFLIGHT: dict = {}


async def _wp_fetch(kind: str, params: dict, key: str, who: str = ""):
    """The half of _wp_get that actually spends a credit."""

    # The allowance is checked here, at the one place a credit is actually
    # spent, so every caller is covered by construction and none of them has to
    # remember to ask.
    room = await _wp_left(who)
    if room["left"] <= 0:
        WP_SPEND["refused"] += 1
        # Which ceiling was hit decides what the user does next: wait for the
        # first of the month, or go and ask the admin. Saying "the allowance is
        # spent" without saying whose sends half of them to the wrong place.
        if room["mine"] <= 0 and who:
            detail = (f"You have used all {WP_PER_USER} of your WhitePages lookups this "
                      f"month. They reset {_month_resets()}. Nothing was looked up. "
                      f"The free sources are unaffected — press Enrich all (free).")
        else:
            detail = (f"The firm's WhitePages allowance ({WP_MONTHLY} lookups) is spent for "
                      f"this month, so nothing was looked up even though you have "
                      f"{room['mine']} of your own left. It resets {_month_resets()}, or an "
                      f"admin can raise it. The free sources are unaffected — press "
                      f"Enrich all (free).")
        raise HTTPException(status_code=400, detail=detail)

    url = WHITEPAGES_BASE_URL + _wp_path(kind)
    WP_SPEND["calls"] += 1
    await _ledger_add("wp", 1, who)
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as cx:
        r = await cx.get(url, params=params, headers={"X-Api-Key": WHITEPAGES_API_KEY})
    if r.status_code in (401, 403):
        raise HTTPException(status_code=502, detail="WhitePages rejected the API key — check WHITEPAGES_API_KEY.")
    if r.status_code == 404:
        await _wp_remember(key, None)
        return None          # documented as "no matching record"
    if r.status_code == 429:
        # Two very different things share this code. Ordinary throttling clears
        # in seconds; a usage cap does not clear until the billing period
        # resets, and telling someone to "try again shortly" when their
        # allowance is gone until the first of the month is useless advice.
        body = {}
        try:
            body = r.json()
        except ValueError:
            pass
        if body.get("error") == "usage_cap_exceeded":
            used, limit = body.get("used"), body.get("limit")
            when = body.get("reset_at") or ""
            raise HTTPException(
                status_code=502,
                detail=("WhitePages allowance is used up"
                        + (f" — {used} of {limit} queries this period" if limit else "")
                        + (f", resets {when[:10]}" if when else "")
                        + ". Nothing further will be looked up until then."),
            )
        raise HTTPException(status_code=502, detail="WhitePages rate limit hit — try again shortly.")
    if r.status_code == 400:
        try:
            msg = (r.json().get("error") or {}).get("long_message") or r.text
        except ValueError:
            msg = r.text
        raise HTTPException(status_code=502, detail=f"WhitePages rejected the query: {msg[:300]}")
    if r.status_code != 200:
        raise HTTPException(
            status_code=502,
            detail=f"WhitePages error {r.status_code} from {url}: {r.text[:300]}",
        )
    try:
        out = r.json()
    except ValueError:
        raise HTTPException(status_code=502, detail=f"WhitePages returned non-JSON from {url}.")
    # A zero-result answer is cached like any other. It was paid for, and the
    # same query will keep returning it.
    await _wp_remember(key, out)
    return out


def _best_person(payload, require_last: str = "") -> dict:
    """Highest-confidence person record out of a search response.

    Live responses wrap candidates in {"results": [...], "metadata": {...}};
    the docs show a bare array. Candidates carry an explicit match_score and
    partial records come back with it null, so pick on score rather than
    trusting position.

    require_last guards against the API's fuzzy fallback: a search for a name
    that isn't in the data happily returns near-spellings — "Tracy" for
    "Treacy" — and attributing a stranger's home address and phone numbers to
    a lead is worse than returning nothing.
    """
    if isinstance(payload, dict):
        for key in ("results", "result"):
            v = payload.get(key)
            if isinstance(v, list):
                payload = v
                break
            if isinstance(v, dict):
                return v
        else:
            return payload if "name" in payload else {}
    if not isinstance(payload, list):
        return {}
    people = [p for p in payload if isinstance(p, dict)]
    last = require_last.strip().lower()
    if last:
        people = [
            p for p in people
            if last in (p.get("name") or "").lower()
            or any(last in (a or "").lower() for a in (p.get("aliases") or []))
        ]
    return max(people, key=lambda p: p.get("match_score") or 0) if people else {}


def _first_str(d: dict, *keys) -> str:
    """First non-empty string among keys, ignoring nested containers."""
    for k in keys:
        v = d.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return str(v)
    return ""


def _digits(s: str) -> str:
    return "".join(c for c in str(s) if c.isdigit())[-10:]


US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "district of columbia": "DC", "florida": "FL", "georgia": "GA", "hawaii": "HI",
    "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI",
    "south carolina": "SC", "south dakota": "SD", "tennessee": "TN", "texas": "TX",
    "utah": "UT", "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
}


def _state_code(s: str) -> str:
    """Two-letter code, or "" when it can't be resolved.

    state_code is validated server-side and a full name like "New York" is a
    400, so an unrecognisable value is dropped from the query rather than sent.
    """
    s = (s or "").strip()
    if len(s) == 2 and s.isalpha():
        return s.upper()
    return US_STATES.get(s.lower(), "")


def _results(payload) -> list:
    """Candidate records out of a search response, whatever the envelope."""
    if isinstance(payload, dict):
        for key in ("results", "result"):
            v = payload.get(key)
            if isinstance(v, list):
                return [p for p in v if isinstance(p, dict)]
            if isinstance(v, dict):
                return [v]
        return [payload] if "name" in payload else []
    if isinstance(payload, list):
        return [p for p in payload if isinstance(p, dict)]
    return []


def _phone_owner(payload, want: str) -> tuple:
    """(record, phone_entry) for whoever actually holds this number.

    A reverse-phone search returns everyone associated with the number —
    current holder, prior holder, household. Ranking by match_score picks the
    strongest name match, which is not the same question. Rank by the API's
    own confidence in the phone-to-person link instead, and ignore records
    that do not list the queried number at all.
    """
    best = (None, None, -1)
    for person in _results(payload):
        for p in _phone_list(person):
            if _digits(p["number"]) != want:
                continue
            score = p.get("score")
            score = score if isinstance(score, (int, float)) else -1
            if score > best[2]:
                best = (person, p, score)
    return best[0], best[1]


def _phone_list(person: dict) -> list:
    """Every line on the record, with what the record says about each.

    The carrier, the prepaid flag, the do-not-call flag and the spam label all
    came back in the same response as the line type; the app read the type and
    dropped the rest, and a comment in verify_phone asserted the carrier was
    "not offered by this API". Whether it is offered depends on the account,
    so it is read where present and left empty where absent, which is a fact
    the user can see rather than an assertion in a comment they cannot.
    """
    out = []
    for p in person.get("phones") or []:
        if not isinstance(p, dict):
            continue
        num = _first_str(p, "number", "phone_number", "phone")
        if num:
            out.append({"number": num,
                        "type": _first_str(p, "type", "line_type", "phone_type"),
                        "score": p.get("score"),
                        "carrier": _first_str(p, "carrier", "carrier_name",
                                              "company", "provider"),
                        "prepaid": _flag(p, "is_prepaid", "prepaid"),
                        "dnc": _flag(p, "do_not_call", "is_do_not_call", "dnc"),
                        "valid": _flag(p, "is_valid", "valid"),
                        "spam": _spam(p)})
    return out


def _mobiles(person: dict) -> list:
    return [p["number"] for p in _phone_list(person) if "mobile" in p["type"].lower()]


# ---------------------------------------------------------------------------
# Reading the record we already paid for
#
# A reverse-phone lookup on this API is a person search — GET /v2/person?phone=
# — so it comes back as a full person record, the same one a name search
# returns. Until now the app read three fields off it (the owner's name, the
# matching line's type, the address) and threw the rest away, then charged a
# second lookup when the user pressed Enrich to fetch the same record again.
#
# The consumer site shows what is in that record: month and year of birth,
# other names, the carrier and spam status of each line, several typed email
# addresses, employer and title, and an address history. All of it arrives in
# the response the app was already receiving.
#
# Every reader below is optional and additive. WhitePages returns what it
# holds and the Pro and Trestle dialects name things differently, so each one
# takes a list of plausible keys and returns nothing when none is present.
# Nothing here fills a gap with a guess: an absent date of birth stays absent
# rather than becoming an estimate, because the entire value of a date of
# birth is that it is not one.
# ---------------------------------------------------------------------------

MONTH_NUM = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"])}
MONTH_NAME = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

DOB_KEYS = ("date_of_birth", "dob", "birth_date", "birthdate", "born")


def _int_or_none(v):
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None


def _month_num(v) -> int:
    """A month as 1-12, from a number or a name. 0 when it is neither."""
    n = _int_or_none(v)
    if n and 1 <= n <= 12:
        return n
    s = str(v or "").strip().lower()[:3]
    return MONTH_NUM.get(s, 0)


def _parse_dob_string(s: str) -> tuple:
    """(year, month, day) out of a written date. Zeros for the parts absent.

    The site prints "Aug 1970" — a month and a year, no day — and the API
    field follows suit, but not in one single format, so the shapes that
    actually occur are each matched explicitly rather than handed to a
    permissive date parser that would read "55-59" as a date.
    """
    s = (s or "").strip()
    if not s:
        return 0, 0, 0
    m = re.match(r"^(\d{4})[-/](\d{1,2})(?:[-/](\d{1,2}))?$", s)      # ISO-ish
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)                 # m/d/Y
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        # Day and month are only distinguishable when one of them is over 12.
        if a > 12 and b <= 12:
            a, b = b, a
        return int(m.group(3)), a, b
    m = re.match(r"^(\d{1,2})/(\d{4})$", s)                           # m/Y
    if m:
        return int(m.group(2)), int(m.group(1)), 0
    year = re.search(r"\b(1[89]\d\d|20\d\d)\b", s)
    if not year:
        return 0, 0, 0
    name = re.search(r"[A-Za-z]{3,}", s)
    mon = _month_num(name.group(0)) if name else 0
    day = 0
    if mon:
        rest = s.replace(year.group(1), " ")
        d = re.search(r"\b(\d{1,2})\b", rest)
        if d:
            day = int(d.group(1))
    return int(year.group(1)), mon, day


def _dob(person: dict) -> dict:
    """Month and year of birth, when the record carries one.

    This is the single most valuable field in the response. Every other age in
    this app is either an integer as of some filing date or a guess from a
    graduation year; a month and a year give the exact month a lead reaches
    59 1/2, which is the whole question the campaign turns on.

    An age_range like "55-59" is deliberately NOT read here. It is a bucket,
    not a date, and the app already has a place for uncertain ages that says
    so out loud.
    """
    raw = None
    for k in DOB_KEYS:
        v = person.get(k)
        if v not in (None, "", [], {}):
            raw = v
            break
    if raw is None:
        return {}
    if isinstance(raw, dict):
        year = _int_or_none(raw.get("year")) or 0
        month = _month_num(raw.get("month"))
        day = _int_or_none(raw.get("day")) or 0
    elif isinstance(raw, (str, int)):
        year, month, day = _parse_dob_string(str(raw))
    else:
        return {}
    this_year = time.gmtime().tm_year
    if not (1900 <= year <= this_year):
        return {}
    if not 1 <= month <= 12:
        month = 0
    if not 1 <= day <= 31:
        day = 0
    out = {"year": year, "month": month, "day": day}
    if month and day:
        out["text"] = f"{MONTH_NAME[month]} {day}, {year}"
    elif month:
        out["text"] = f"{MONTH_NAME[month]} {year}"
    else:
        out["text"] = str(year)
    return out


def _flag(d: dict, *keys):
    """True/False from whichever key is present, None when none of them is.

    None and False are different answers here — "this API did not tell us
    whether the line is on the do-not-call list" must not read as "it is not".
    """
    for k in keys:
        v = d.get(k)
        if isinstance(v, bool):
            return v
        if isinstance(v, str) and v.strip().lower() in ("true", "yes", "y", "1"):
            return True
        if isinstance(v, str) and v.strip().lower() in ("false", "no", "n", "0"):
            return False
    return None


def _spam(d: dict) -> str:
    """Spam or risk label on a line, as the record words it.

    Reported verbatim rather than thresholded into a yes/no, because the
    scales differ between fields and a number invented here would look
    looked-up.
    """
    for k in ("spam_risk", "spam_status", "risk_level", "reputation"):
        v = d.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, dict):
            s = _first_str(v, "level", "risk", "status", "label")
            if s:
                return s
    for k in ("spam_score", "risk_score"):
        v = d.get(k)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return str(v)
    v = _flag(d, "is_spam", "spam")
    return "spam" if v is True else ""


def _aliases(person: dict) -> list:
    """Other names on the record — married, maiden, misspelt.

    Worth having for its own sake: an email to a name the lead no longer uses
    is a worse opening than no email, and a maiden name explains a household
    record that otherwise looks like a stranger's.
    """
    out = []
    for key in ("aliases", "akas", "also_known_as", "other_names", "names"):
        v = person.get(key)
        if not isinstance(v, list):
            continue
        for a in v:
            s = a.strip() if isinstance(a, str) else (
                _first_str(a, "name", "full_name") if isinstance(a, dict) else "")
            if s and s not in out:
                out.append(s)
    return out[:8]


def _email_list(person: dict) -> list:
    """Every email on the record, with what the record says about each.

    The type is the point. A professional address at the employer confirms the
    employer; a personal one is the address a 59-year-old actually reads, and
    a "recently used" flag says which of four old ones is live. Flattening
    them to a list of strings, which is what the app did, throws away the only
    thing that distinguishes them.
    """
    out = []
    for e in person.get("emails") or []:
        if isinstance(e, str) and e.strip():
            out.append({"email": e.strip(), "type": "", "recent": None, "valid": None})
            continue
        if not isinstance(e, dict):
            continue
        addr = _first_str(e, "email", "email_address", "address")
        if not addr:
            continue
        out.append({
            "email": addr,
            "type": _first_str(e, "type", "email_type", "category").lower(),
            "recent": _flag(e, "is_recently_used", "recently_used", "is_recent"),
            "valid": _flag(e, "is_valid", "valid", "is_deliverable"),
        })
    return out[:10]


def _jobs(person: dict) -> list:
    """Employer and title, when the record carries them.

    Cross-checking rather than sourcing: the lead already has an employer from
    the list it was built from, and a second source agreeing is worth showing.
    A second source disagreeing is worth showing more.
    """
    out = []
    title = _first_str(person, "job_title", "occupation", "title")
    emp = _first_str(person, "company", "employer", "company_name", "organization")
    if title or emp:
        out.append({"title": title, "employer": emp})
    for key in ("jobs", "employments", "job_history", "occupations", "work"):
        v = person.get(key)
        if not isinstance(v, list):
            continue
        for j in v:
            if not isinstance(j, dict):
                continue
            t = _first_str(j, "title", "job_title", "position", "occupation")
            c = _first_str(j, "company", "employer", "company_name", "organization")
            if (t or c) and not any(x["title"] == t and x["employer"] == c for x in out):
                out.append({"title": t, "employer": c})
    return out[:4]


def _person_facts(person: dict) -> dict:
    """Everything useful in one person record.

    Shared by Enrich and by Verify, because they receive the same record and
    there is no reason the one that arrived first should be read less
    thoroughly than the one that arrived second.
    """
    addr = _home_address(person)
    phones = _phone_list(person)
    emails = _email_list(person)
    return {
        "owner": _first_str(person, "name", "full_name"),
        "aliases": _aliases(person),
        "age": person.get("age"),
        "dob": _dob(person),
        "home_street": addr["street"],
        "home_city": addr["city"],
        "home_state": addr["state"],
        "home_zip": addr["zip"],
        "mobiles": [p["number"] for p in phones if "mobile" in p["type"].lower()],
        "phone_records": phones,
        "phones_total": _list_total(person, "phones"),
        "emails": [e["email"] for e in emails],
        "email_records": emails,
        "jobs": _jobs(person),
        "linkedin_url": _first_str(person, "linkedin_url"),
        "properties": _property_addresses(person),
        "properties_owned": _list_total(person, "owned_properties"),
        "prior_places": _prior_places(person),
        "addresses_total": _list_total(person, "historical_addresses")
                           or _list_total(person, "addresses"),
        "relatives": [_first_str(r, "name") for r in (person.get("relatives") or [])
                      if isinstance(r, dict) and _first_str(r, "name")][:8],
    }


EMPTY_ADDR = {"street": "", "city": "", "state": "", "zip": ""}


def _parse_one_line(s: str) -> dict:
    """'123 Main St, Seattle, WA 98101' -> parts."""
    bits = [b.strip() for b in s.split(",") if b.strip()]
    if not bits:
        return dict(EMPTY_ADDR)
    tail = bits[-1].split() if len(bits) >= 3 else []
    return {
        "street": bits[0],
        "city": bits[1] if len(bits) >= 3 else "",
        "state": tail[0] if tail else "",
        "zip": tail[1] if len(tail) > 1 else "",
    }


def _addr(raw) -> dict:
    """Address parts from a structured record or a one-line string."""
    if isinstance(raw, str):
        return _parse_one_line(raw)
    if not isinstance(raw, dict):
        return dict(EMPTY_ADDR)
    parts = {
        "street": _first_str(raw, "line1", "street_line_1", "street_line", "street"),
        "city": _first_str(raw, "city", "city_name", "locality"),
        "state": _first_str(raw, "state", "state_code", "region"),
        "zip": _first_str(raw, "zip", "postal_code", "zip_code"),
    }
    if not parts["city"]:
        one = _first_str(raw, "full_address", "address", "formatted_address")
        if one:
            return _parse_one_line(one)
    return parts


def _home_address(person: dict) -> dict:
    for key in ("current_addresses", "addresses", "current_address", "address"):
        v = person.get(key)
        if isinstance(v, list) and v:
            return _addr(v[0])
        if isinstance(v, (dict, str)) and v:
            return _addr(v)
    return dict(EMPTY_ADDR)


def _list_total(person: dict, name: str) -> int:
    """Full count of a capped list: what came back plus what was withheld."""
    shown = person.get(name)
    shown = len(shown) if isinstance(shown, list) else 0
    meta = ((person.get("result_metadata") or {}).get(name)) or {}
    return shown + (meta.get("additional") or 0)


@app.post("/api/verify-phone")
async def verify_phone(req: VerifyPhoneRequest, request: Request):
    """Who a number belongs to, and everything else that came back with it.

    The Pro API answers a reverse-phone query with a full person record, so
    this call has always been buying far more than the three fields it read.
    It now returns the record too, under "record", so pressing the phone check
    fills the household panel at no extra cost and the Enrich button is a
    second lookup only when the first one found nothing.

    The name-match verdict is unchanged: it is what stops someone dialling a
    stranger, and nothing added here should be able to weaken it.
    """
    await _active_token(request)
    want = _digits(req.phone)
    person, hit = await _wp_phone(want, who=await _who(request))
    if not person or not hit:
        return {"valid": None, "line_type": "", "carrier": "", "prepaid": None,
                "owner": "", "name_match": None, "note": "no record found"}

    owner = _first_str(person, "name", "full_name")
    facts = _person_facts(person)

    # A married name is still the same person. The surname test used to look
    # only at the primary name, so a record filed under a maiden name read as
    # a wrong number and the lead was flagged undialable.
    name_match = None
    last = req.last_name.strip().lower()
    if last and (owner or facts["aliases"]):
        name_match = (last in owner.lower()
                      or any(last in a.lower() for a in facts["aliases"]))

    owner_addr = _home_address(person)
    hit = hit or {}
    # The record listing the number is itself the evidence it is real, unless
    # the record explicitly says otherwise.
    stated_valid = hit.get("valid")
    return {
        "valid": stated_valid if stated_valid is not None else True,
        "line_type": hit.get("type", ""),
        "carrier": hit.get("carrier", ""),
        "prepaid": hit.get("prepaid"),
        "dnc": hit.get("dnc"),
        "spam": hit.get("spam", ""),
        "owner": owner,
        "name_match": name_match,
        "matched_alias": next((a for a in facts["aliases"]
                               if last and last in a.lower()
                               and last not in owner.lower()), ""),
        "owner_city": owner_addr["city"],
        "owner_state": owner_addr["state"],
        "owner_street": owner_addr["street"],
        "owner_zip": owner_addr["zip"],
        "same_household": _same_household(owner_addr, req),
        # Everything else the lookup already returned. Named "record" rather
        # than "household" because the front end uses "household" for a
        # different person living at the same address.
        "record": facts,
    }


def _same_household(addr: dict, req: "VerifyPhoneRequest") -> Optional[bool]:
    """Whether the number's owner lives at the lead's address.

    A different name at the same address is usually a spouse — often the
    better prospect for a retirement conversation. A different name at a
    different address is a wrong number. Unknown when we have nothing to
    compare against.
    """
    lead_zip = "".join(c for c in req.zip if c.isdigit())[:5]
    got_zip = "".join(c for c in addr["zip"] if c.isdigit())[:5]
    if req.street and addr["street"]:
        same_street = _norm_street(req.street) == _norm_street(addr["street"])
        if lead_zip and got_zip:
            return same_street and lead_zip == got_zip
        return same_street
    if lead_zip and got_zip:
        return lead_zip == got_zip
    return None


UNIT_WORDS = {"apt", "apartment", "unit", "suite", "ste", "fl", "floor", "rm", "room"}
STREET_WORDS = {"street", "st", "avenue", "ave", "road", "rd", "drive", "dr",
                "lane", "ln", "court", "ct", "boulevard", "blvd", "place", "pl",
                "terrace", "ter", "circle", "cir", "highway", "hwy", "way",
                "north", "south", "east", "west", "n", "s", "e", "w"}


def _norm_street(s: str) -> str:
    """Loose street comparison: case, punctuation, suffixes and unit numbers.

    Two people in the same building are not the same household, but "14
    Alexander Ave Apt B" and "14 Alexander Avenue" are the same record written
    two ways — the unit is dropped by one source, not by the resident. Cut at
    the unit designator so the comparison is house number plus street name.
    """
    s = "".join(c.lower() if c.isalnum() else " " for c in s).split()
    out = []
    for w in s:
        if w in UNIT_WORDS:
            break               # everything after this is a unit, not a street
        if w not in STREET_WORDS:
            out.append(w)
    return " ".join(out)


class ZIRequest(BaseModel):
    path: str                       # e.g. "search/contact"
    body: dict = {}
    method: str = "POST"


@app.post("/api/zi/search")
async def zi_search(req: ZIRequest, request: Request):
    """Thin passthrough to ZoomInfo using *this user's* token.

    Deliberately not a parser. §12's rule is to see a live response before
    writing one, and no response from this endpoint has been seen yet, so the
    raw JSON is handed back for the client to shape. That also means a change at
    ZoomInfo's end surfaces as odd data rather than as a silent empty list.
    """
    await _active_token(request)                 # app sign-in first
    token = await _zi_token(request.state.session)
    if not token:
        raise HTTPException(status_code=401,
                            detail="Connect your ZoomInfo account first.")
    url = f"{ZI_API_BASE}/{req.path.lstrip('/')}"
    async with httpx.AsyncClient(timeout=60) as cx:
        if req.method.upper() == "GET":
            r = await cx.get(url, params=req.body,
                             headers={"Authorization": f"Bearer {token}"})
        else:
            r = await cx.post(url, json=req.body,
                              headers={"Authorization": f"Bearer {token}"})
    if r.status_code != 200:
        raise HTTPException(status_code=502,
                            detail=f"ZoomInfo {r.status_code}: {r.text[:300]}")
    try:
        return r.json()
    except Exception:
        raise HTTPException(status_code=502,
                            detail=f"ZoomInfo returned non-JSON: {r.text[:200]}")


@app.get("/api/zi-debug")
async def zi_debug(request: Request, path: str = "lookup/inputfields/contact/search"):
    """Raw ZoomInfo round-trip for this user's token — URL, status, body.

    The same probe that eventually diagnosed WhitePages, added up front this
    time rather than after two wrong parsers.
    """
    await _active_token(request)
    if not _zi_configured():
        return {"error": "ZI_CLIENT_ID / ZI_CLIENT_SECRET are not set on this service."}
    token = await _zi_token(request.state.session)
    if not token:
        return {"error": "This account has not connected ZoomInfo.",
                "connect": "/auth/zoominfo/login"}
    url = f"{ZI_API_BASE}/{path.lstrip('/')}"
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as cx:
        r = await cx.get(url, headers={"Authorization": f"Bearer {token}"})
    return {"url": url, "status": r.status_code, "body": r.text[:4000]}


# ------------------- Money in motion: WARN x Form 5500 -------------------
# The rest of the app scores a list somebody else built. This finds the events
# that make the list: a dated mass separation at a named employer, priced by the
# size of that employer's retirement plan.

def _source_states() -> set:
    return {s.strip().upper() for s in SOURCE_STATES.split(",") if s.strip()}


def _source_counties() -> set:
    # Normalised here so "Nassau County" in the env var matches "Nassau" in a
    # feed, and the comparison inside build_opportunities stays a set lookup.
    return {prospecting.norm_county(c) for c in SOURCE_COUNTIES.split(",") if c.strip()}


# The feed source an admin saved in the app wins over the env var. Config is
# data (CLAUDE.md principle 3), and three shredded gcloud attempts on this one
# variable earned it a settings field: a URL or a JSON list, saved in
# Firestore, no shell quoting anywhere in the loop.
_FEEDS_CFG = {"at": 0.0, "value": None}
_MEM_FEEDS: dict = {}                     # local-dev fallback, like every _MEM_*


async def _warn_feeds_value() -> tuple:
    """(value, source_label). Cached a minute so /api/me stays cheap."""
    now = time.time()
    if _FEEDS_CFG["value"] is None or now - _FEEDS_CFG["at"] > 60:
        doc = await _fs_get(FS_STATE, "warn_feeds") or _MEM_FEEDS.get("v") or {}
        _FEEDS_CFG.update(at=now, value=(doc.get("value") or "").strip())
    if _FEEDS_CFG["value"]:
        return _FEEDS_CFG["value"], "The feed source saved in the app"
    return WARN_FEEDS, "WARN_FEEDS"


class FeedsCfg(BaseModel):
    value: str = ""


@app.post("/api/sources/feeds")
async def set_feeds(body: FeedsCfg, request: Request, email: str = Depends(signed_in)):
    """Save the WARN feed source in the app — a URL to a JSON list, or the
    list itself. Validated before it is stored, because storing a broken
    value would only move the diagnostic, not the problem. Admin-only: the
    feed list is firm-wide configuration. An empty value clears the stored
    one and the env var applies again."""
    if not _is_admin(email):
        raise HTTPException(status_code=403, detail="Only an admin can set the firm's feed source.")
    value = (body.value or "").strip()
    count = 0
    if value:
        try:
            raw = value
            if not raw.lstrip().startswith("["):
                raw = await _get(value, drive_token="")
            feeds = json.loads(raw)
            good = [f for f in feeds if isinstance(f, dict) and f.get("url")] \
                if isinstance(feeds, list) else []
            if not good:
                raise ValueError("no feed objects carrying a url in it")
            count = len(good)
        except HTTPException as e:
            raise HTTPException(status_code=400,
                                detail=f"That did not yield a usable feed list "
                                       f"({str(e.detail)[:140]}). Nothing was saved.")
        except Exception as e:
            raise HTTPException(status_code=400,
                                detail=f"That did not yield a usable feed list "
                                       f"({type(e).__name__}: {str(e)[:140]}). Nothing was saved.")
    doc = {"value": value, "by": email, "at": time.time()}
    if not await _fs_set(FS_STATE, "warn_feeds", doc):
        _MEM_FEEDS["v"] = doc
    _FEEDS_CFG.update(at=0.0, value=None)
    return {"ok": True, "feeds": count, "cleared": not value}


async def _warn_feeds(drive_token: str = "") -> list:
    return (await _warn_feeds_checked(drive_token))[0]


async def _warn_feeds_checked(drive_token: str = "") -> tuple:
    """(feeds, complaint). A complaint means the setting is wrong, not empty.

    Three silences used to look identical: nothing set, JSON that will not
    parse, and entries with no url in them. Only the first is "you have not
    configured this yet"; the other two are a variable that was typed and is
    not doing anything, which is the worst of the three to leave unnamed —
    it looks configured from the outside and reports no layoffs from the
    inside.
    """
    raw, src = await _warn_feeds_value()
    raw = raw.strip()
    if not raw:
        return [], ""
    # The value may also be a single URL (or Drive link) to a JSON file holding
    # the list. One token, no commas — immune to gcloud's env-var splitting,
    # which is exactly the mistake that shredded this variable in the field.
    # tools/warn_sync.py publishes precisely that file (warn_feeds.json).
    if raw.startswith(("http://", "https://", "drive:")) or _DRIVE_ID.search(raw):
        try:
            raw = await _get(raw, drive_token=drive_token)
        except Exception as e:
            return [], (f"{src} points at {raw[:80]} but fetching it "
                        f"failed ({type(e).__name__}: {str(e)[:120]}).")
    try:
        feeds = json.loads(raw)
    except ValueError as e:
        return [], (f"{src} is set but is not valid JSON ({e}). It must be a list "
                    f'like [{{"id":"nj","state":"NJ","format":"csv","url":"https://..."}}] '
                    f"— or one URL to a JSON file holding that list, which survives "
                    f"gcloud's comma-splitting because it has no commas in it.")
    if not isinstance(feeds, list):
        return [], (f"{src} parsed as "
                    f"{type(feeds).__name__}, but it must be a list of feed objects.")
    good = [f for f in feeds if isinstance(f, dict) and f.get("url")]
    if feeds and not good:
        return [], (f"{src} has {len(feeds)} entr"
                    f"{'y' if len(feeds) == 1 else 'ies'} but none of them carries a "
                    f'"url", so there is nothing to fetch.')
    return good, ""


# A source can live in the advisor's own Drive rather than on a government
# host. That is often the better arrangement: the DOL publishes a zip behind a
# path that moves each year, whereas a file dropped in Drive is stable, already
# unpacked, and under the control of whoever is going to notice when it is
# stale. Accepts a Drive share link, a bare file id, or drive:<id>.
_DRIVE_ID = re.compile(r"(?:^drive:|/d/|[?&]id=)([A-Za-z0-9_-]{20,})")


def drive_file_id(ref: str) -> str:
    t = (ref or "").strip()
    m = _DRIVE_ID.search(t)
    if m:
        return m.group(1)
    return t if re.fullmatch(r"[A-Za-z0-9_-]{25,}", t) else ""


async def _drive_get(file_id: str, token: str, *, as_bytes: bool = False):
    async with httpx.AsyncClient(timeout=300, follow_redirects=True) as cx:
        r = await cx.get(f"{DRIVE_FILES_URL}/{file_id}",
                         headers={"Authorization": f"Bearer {token}"},
                         params={"alt": "media", "supportsAllDrives": "true"})
    if r.status_code != 200:
        raise HTTPException(
            status_code=502,
            detail=f"Drive returned {r.status_code} for that file. Check it is in the "
                   f"signed-in account's Drive and is not a Google-native document.")
    return r.content if as_bytes else r.text


async def _get(url: str, *, as_bytes: bool = False, timeout: int = 120,
               drive_token: str = ""):
    fid = drive_file_id(url)
    if fid and drive_token:
        return await _drive_get(fid, drive_token, as_bytes=as_bytes)
    if fid and not url.lower().startswith("http"):
        raise HTTPException(
            status_code=400,
            detail="That source is a Google Drive file, but Drive is not connected. "
                   "Sign in with Google, or point it at a public URL instead.")
    async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as cx:
        r = await cx.get(url, headers={"User-Agent": EDGAR_USER_AGENT or
                                       "FPA Lead Qualifier"})
    if r.status_code != 200:
        raise HTTPException(status_code=502,
                            detail=f"{url} returned {r.status_code}: {r.text[:200]}")
    return r.content if as_bytes else r.text


# ---- source cache ----
# Both loaders download and parse an entire file. Without this, every click of
# "Price the employers", every "Check for events", and every opportunities
# refresh re-fetched about 36MB and re-parsed a few hundred thousand rows —
# per user, per click. Against a government host that is also impolite.
#
# In process rather than in Firestore: the parsed index is large and cheap to
# rebuild, and Cloud Run scaling to zero simply means the next request pays for
# it once. A cold start re-fetching is the correct behaviour, not a miss.
_SRC_CACHE: dict[str, tuple] = {}
PLANS_TTL = int(os.environ.get("PLANS_CACHE_SECONDS", "86400"))   # filed monthly
WARN_TTL = int(os.environ.get("WARN_CACHE_SECONDS", "21600"))     # notices land daily


def _cache_get(key: str, ttl: int):
    hit = _SRC_CACHE.get(key)
    if hit and time.time() - hit[0] < ttl:
        return hit[1]
    return None


def _cache_put(key: str, value):
    # One entry per distinct source configuration, and only a handful can ever
    # exist, so nothing needs evicting except by age.
    _SRC_CACHE[key] = (time.time(), value)
    return value


async def _load_warn(drive_token: str = "", fresh: bool = False) -> dict:
    """Every configured WARN feed, normalised. Reports per-feed what happened."""
    key = "warn|" + (await _warn_feeds_value())[0]
    if not fresh:
        got = _cache_get(key, WARN_TTL)
        if got is not None:
            return {**got, "cached": True}
    events, report = [], []
    for feed in await _warn_feeds(drive_token):
        fid = feed.get("id") or feed.get("state") or feed["url"][:40]
        try:
            if (feed.get("format") or "csv").lower() == "json":
                raw = json.loads(await _get(feed["url"], drive_token=drive_token))
                if isinstance(raw, dict):
                    # Socrata sometimes wraps rows; take the first list it holds.
                    raw = next((v for v in raw.values() if isinstance(v, list)), [])
                parsed = prospecting.parse_warn_json(raw, feed.get("state", ""))
            else:
                parsed = prospecting.parse_warn_csv(
                    await _get(feed["url"], drive_token=drive_token), feed.get("state", ""))
        except HTTPException as e:
            report.append({"id": fid, "error": e.detail})
            continue
        except Exception as e:
            report.append({"id": fid, "error": f"{type(e).__name__}: {str(e)[:200]}"})
            continue
        events.extend(parsed["events"])
        report.append({"id": fid, "events": len(parsed["events"]),
                       "mapped": parsed["mapped"], "unmapped": parsed["unmapped"]})
    return _cache_put(key, {"events": events, "feeds": report})


async def _drive_token_for(request: Request) -> str:
    """The signed-in user's Google token, or "" — Drive-hosted sources need it,
    web-hosted ones do not, so a Microsoft-only session is not an error here."""
    try:
        return await _google_token(request.state.session) or ""
    except Exception:
        return ""


async def _fetch_source(ref: str, token: str, want: str) -> str:
    """One source file as text, from Drive or the open web, unzipped if needed."""
    blob = await _get(ref, as_bytes=True, drive_token=token)
    if ref.lower().endswith(".zip") or blob[:2] == b"PK":
        return prospecting.unzip_first_csv(blob, want)
    return blob.decode("utf-8", errors="replace")


async def _load_plans(drive_token: str = "", fresh: bool = False) -> dict:
    """Form 5500 sponsors, with assets joined in from Schedule H or I.

    The 5500 file itself carries no money — only participant counts — so
    without the schedule there is no average balance and nothing to rank on.
    That is reported rather than left to be discovered.
    """
    if not FORM5500_URL:
        return {"plans": {}, "note": "FORM5500_URL is not set."}
    key = "plans|" + FORM5500_URL + "|" + ",".join(FORM5500_SCHEDULE_URLS) + "|" + SOURCE_STATES
    if not fresh:
        got = _cache_get(key, PLANS_TTL)
        if got is not None:
            return {**got, "cached": True}
    text = await _fetch_source(FORM5500_URL, drive_token, FORM5500_CSV_IN_ZIP)
    parsed = prospecting.parse_5500_csv(text, states=_source_states() or None)
    out = {"plans": parsed["plans"], "rows": parsed.get("rows"),
           "kept": parsed.get("kept"), "mapped": parsed["mapped"],
           "unmapped": parsed["unmapped"]}

    priced = sum(1 for p in parsed["plans"].values() if p.get("avg_balance"))
    if priced:
        out["priced"] = priced
        return _cache_put(key, out)

    if not FORM5500_SCHEDULE_URLS:
        out["note"] = ("The 5500 file has participant counts but no assets — assets are "
                       "on Schedule H and Schedule I, which are separate downloads. Set "
                       "FORM5500_SCHEDULE_URLS or nothing can be priced.")
        return _cache_put(key, out)

    assets, notes = {}, []
    for ref in FORM5500_SCHEDULE_URLS:
        try:
            sched = prospecting.parse_schedule_assets(
                await _fetch_source(ref, drive_token, "sch"))
            if sched.get("note"):
                notes.append(sched["note"])
            # Later schedules fill gaps rather than overwrite: Schedule H is the
            # large-plan file and is the better number where both exist.
            for k, v in sched["assets"].items():
                assets.setdefault(k, v)
        except HTTPException as e:
            notes.append(f"schedule unavailable: {e.detail}"[:200])
        except Exception as e:
            notes.append(f"schedule unreadable: {type(e).__name__}")
    rep = prospecting.attach_assets(parsed["plans"], assets)
    out["priced"] = rep["filled"]
    out["schedule_rows"] = len(assets)
    if notes:
        out["note"] = " ".join(notes)
    elif not rep["filled"]:
        out["note"] = ("Schedule files loaded but nothing joined — the ACK_ID values do "
                       "not line up with the 5500 file. Check both are the same year.")
    return _cache_put(key, out)


@app.get("/api/sources/probe")
async def sources_probe(request: Request, which: str = "all"):
    """What the feeds actually return, and which columns were matched.

    None of these hosts is reachable from the environment this was written in,
    so this is how the column mapping gets pinned to reality instead of to a
    guess. Run it once after setting the URLs; anything in `unmapped` is a
    column alias that needs adding.
    """
    await _active_token(request)
    out = {"states": sorted(_source_states()), "counties": sorted(_source_counties()),
           "min_workers": SOURCE_MIN_WORKERS}
    if which in ("all", "warn"):
        feeds = await _warn_feeds(await _drive_token_for(request))
        out["warn"] = {"configured": len(feeds)}
        if feeds:
            got = await _load_warn(await _drive_token_for(request), fresh=True)
            out["warn"]["feeds"] = got["feeds"]
            out["warn"]["sample"] = got["events"][:3]
    if which in ("all", "5500"):
        if not FORM5500_URL:
            out["form5500"] = {"error": "FORM5500_URL is not set."}
        else:
            try:
                got = await _load_plans(await _drive_token_for(request), fresh=True)
                sample = list(got["plans"].items())[:3]
                out["form5500"] = {"rows_read": got.get("rows"), "kept": got.get("kept"),
                                   "mapped": got.get("mapped"), "unmapped": got.get("unmapped"),
                                   # The number that decides whether anything can be
                                   # ranked. Sponsors without it are counted, not hidden.
                                   "priced": got.get("priced"),
                                   "schedule_rows": got.get("schedule_rows"),
                                   "schedules_configured": len(FORM5500_SCHEDULE_URLS),
                                   "note": got.get("note"),
                                   "sample": [v for _, v in sample]}
            except HTTPException as e:
                out["form5500"] = {"error": e.detail}
            except Exception as e:
                out["form5500"] = {"error": f"{type(e).__name__}: {str(e)[:300]}"}
    return out


@app.post("/api/sources/refresh")
async def sources_refresh(request: Request):
    """Rebuild the opportunity list and store it.

    Cheap enough to run on a schedule — Cloud Scheduler hitting this weekly is
    what turns a set of buttons into something that watches.
    """
    await _active_token(request)
    tok = await _drive_token_for(request)
    warn = await _load_warn(tok, fresh=True)
    if not warn["events"]:
        return {"stored": 0, "feeds": warn["feeds"],
                "note": ((await _warn_feeds_checked(tok))[1]
                         or "No WARN events. Set WARN_FEEDS, then check /api/sources/probe.")}
    try:
        plans = await _load_plans(tok, fresh=True)
    except Exception as e:
        plans = {"plans": {}, "note": f"Form 5500 unavailable: {type(e).__name__}: {str(e)[:200]}"}
    opps = prospecting.build_opportunities(
        warn["events"], plans.get("plans") or {},
        min_workers=SOURCE_MIN_WORKERS, states=_source_states() or None,
        counties=_source_counties() or None)
    await _fs_set(FS_OPPS, "latest", {
        "built_at": int(time.time()),
        "count": len(opps),
        "matched": sum(1 for o in opps if o["plan_matched"]),
        "items": opps[:500],                 # the tail is noise, and documents have limits
    })
    return {"stored": len(opps), "matched": sum(1 for o in opps if o["plan_matched"]),
            "feeds": warn["feeds"], "plans": {k: v for k, v in plans.items() if k != "plans"}}


@app.get("/api/opportunities")
async def opportunities(request: Request, refresh: bool = False):
    """Employers with money about to come loose, biggest first."""
    await _active_token(request)
    if not refresh:
        doc = await _fs_get(FS_OPPS, "latest")
        if doc:
            return doc
    warn = await _load_warn(await _drive_token_for(request), fresh=refresh)
    if not warn["events"]:
        # Two different silences, and telling them apart is the whole point:
        # nothing is configured yet, or feeds ran and found nothing. Pointing
        # a user at a markdown file answers neither, and it was the first
        # thing a new account saw here.
        _feeds, complaint = await _warn_feeds_checked()
        if complaint:
            # Configured, and configured wrongly. Naming that is the difference
            # between a five-minute fix and a week of believing there are no
            # layoffs anywhere.
            note = complaint + " Nothing else in the app depends on this."
        elif not (await _warn_feeds_value())[0].strip():
            note = ("This needs at least one state's layoff-notice feed, which an admin "
                    "sets once for the whole firm (WARN_FEEDS). Nothing else in the app "
                    "depends on it — sourcing, enrichment and scoring all work without it.")
        else:
            # A feed reports either events+mapped, or an error. There is no
            # "ok" flag to read; the absence of an error is the success.
            ran = [f for f in warn["feeds"] if not f.get("error")]
            broke = [f for f in warn["feeds"] if f.get("error")]
            note = (f"{len(ran)} feed{'' if len(ran) == 1 else 's'} ran and reported no "
                    f"notices in the window.")
            if broke:
                note += (" " + ", ".join(f"{f.get('id') or 'a feed'} failed "
                                         f"({str(f.get('error') or 'no answer')[:60]})"
                                         for f in broke[:3])
                         + " — a feed that fails is a gap, not an all-clear.")
        return {"built_at": int(time.time()), "count": 0, "matched": 0, "items": [],
                "feeds": warn["feeds"],
                "configured": bool((await _warn_feeds_value())[0].strip()),
                "note": note}
    try:
        plans = (await _load_plans(await _drive_token_for(request), fresh=refresh)).get("plans") or {}
    except Exception:
        plans = {}
    opps = prospecting.build_opportunities(
        warn["events"], plans, min_workers=SOURCE_MIN_WORKERS,
        states=_source_states() or None, counties=_source_counties() or None)
    return {"built_at": int(time.time()), "count": len(opps),
            "matched": sum(1 for o in opps if o["plan_matched"]),
            "items": opps[:500], "feeds": warn["feeds"]}


class WarmOpportunitiesRequest(BaseModel):
    leads: list = []
    refresh: bool = False
    sort: str = "warmth"


@app.post("/api/opportunities/warmth")
async def opportunities_warmth(body: WarmOpportunitiesRequest, request: Request):
    """The same opportunities, ranked by the way in you already have.

    The leads are posted by the client rather than read server-side, for the
    same reason /api/signals does it: it then works on the list currently open
    whether or not it has been saved, and on a list shared by another advisor,
    without this route re-deriving who may see what.

    Nothing is fetched that the GET route would not fetch. This only asks a
    second question of a list the advisor already has.
    """
    await _active_token(request)
    warn = await _load_warn(await _drive_token_for(request), fresh=body.refresh)
    if not warn["events"]:
        return {"built_at": int(time.time()), "count": 0, "matched": 0, "items": [],
                "feeds": warn["feeds"], "warmth": {},
                "note": "No WARN feeds configured — see SETUP-prospecting.md."}
    try:
        plans = (await _load_plans(await _drive_token_for(request),
                                   fresh=body.refresh)).get("plans") or {}
    except Exception:
        plans = {}
    opps = prospecting.build_opportunities(
        warn["events"], plans, min_workers=SOURCE_MIN_WORKERS,
        states=_source_states() or None, counties=_source_counties() or None)
    prospecting.annotate_warmth(opps, body.leads)
    prospecting.sort_opportunities(opps, by=body.sort)
    tally = {}
    for o in opps:
        tally[o["warmth"]] = tally.get(o["warmth"], 0) + 1
    return {"built_at": int(time.time()), "count": len(opps),
            "matched": sum(1 for o in opps if o["plan_matched"]),
            "warmth": tally, "sort": body.sort,
            "items": opps[:500], "feeds": warn["feeds"]}


# ------------------- ZoomInfo through the MCP connector -------------------
# The REST integration above needs a DevPortal entitlement this subscription does
# not have. This route needs none: Anthropic's MCP connector opens the connection
# to ZoomInfo's MCP server server-side, carrying the user's own token, and the
# app only ever talks to Claude.
#
# Two halves are mandatory and the app previously sent only one. `mcp_servers`
# alone is rejected as a validation error — every server must also be referenced
# by an `mcp_toolset` entry in `tools`, under the mcp-client-2025-11-20 beta.

MCP_BETA = "mcp-client-2025-11-20"


def _zi_mcp_token(session: dict) -> str:
    return (session.get("zi_mcp") or {}).get("token") or ""


class ZIMcpRequest(BaseModel):
    prompt: str
    max_tokens: int = 4000


class ZIMcpToken(BaseModel):
    token: str = ""


@app.post("/api/zi/mcp-token")
async def zi_mcp_token_set(req: ZIMcpToken, request: Request):
    """Store (or clear) this user's ZoomInfo MCP token.

    It lives in the session document, which is KMS-encrypted when KMS_KEY_NAME is
    set — the same protection the Google and Microsoft refresh tokens get. It is
    never returned to the browser afterwards; /api/me reports only whether one
    exists.
    """
    await _active_token(request)
    tok = (req.token or "").strip()
    if tok:
        request.state.session["zi_mcp"] = {"token": tok, "saved_at": int(time.time())}
        await _save_all_attachments(request.state.session)
    else:
        request.state.session.pop("zi_mcp", None)
        await _del_attachment(_identity(request.state.session), "zimcp")
    return {"connected": bool(tok)}


@app.post("/api/zi/mcp")
async def zi_mcp(req: ZIMcpRequest, request: Request):
    """Ask Claude to run a ZoomInfo query with this user's MCP token.

    Returns the tool results raw. Like /api/zi/search this is deliberately not a
    parser — the client shapes it, so a change at ZoomInfo's end shows up as odd
    data rather than a silently empty list.
    """
    await _active_token(request)
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500,
                            detail="ANTHROPIC_API_KEY is not set on this service.")
    token = _zi_mcp_token(request.state.session)
    if not token:
        raise HTTPException(
            status_code=401,
            detail="No ZoomInfo MCP token saved. Add one under ICP settings, or "
                   "run the app inside Claude to use the connector there.")

    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = await client.beta.messages.create(
            model=ZI_MCP_MODEL or CLAUDE_MODEL,
            max_tokens=max(256, min(req.max_tokens, 16000)),
            betas=[MCP_BETA],
            mcp_servers=[{
                "type": "url",
                "url": ZI_MCP_URL,
                "name": "zoominfo",
                "authorization_token": token,
            }],
            # Required. Without it the request is rejected outright.
            tools=[{"type": "mcp_toolset", "mcp_server_name": "zoominfo"}],
            messages=[{"role": "user", "content": req.prompt}],
        )
    except anthropic.APIStatusError as e:
        raise HTTPException(status_code=502,
                            detail=f"Claude API error {e.status_code}: {str(e)[:300]}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {str(e)[:300]}")

    if getattr(msg, "stop_reason", "") == "refusal":
        raise HTTPException(status_code=502, detail="Claude declined this ZoomInfo request.")

    results, text = [], []
    for block in msg.content:
        btype = getattr(block, "type", "")
        if btype == "mcp_tool_result":
            for part in (getattr(block, "content", None) or []):
                t = getattr(part, "text", None)
                if t:
                    results.append(t)
        elif btype == "text":
            text.append(block.text)
    return {"results": results, "text": " ".join(text)[:2000],
            "stop_reason": getattr(msg, "stop_reason", "")}


@app.get("/api/zi/mcp-debug")
async def zi_mcp_debug(request: Request):
    """Smallest possible round-trip through the connector — does the token work?

    mcp.zoominfo.com is unreachable from the environment this was written in, so
    this is how the first real answer arrives.
    """
    await _active_token(request)
    if not ANTHROPIC_API_KEY:
        return {"error": "ANTHROPIC_API_KEY is not set on this service."}
    token = _zi_mcp_token(request.state.session)
    if not token:
        return {"error": "No ZoomInfo MCP token saved for this account.",
                "mcp_url": ZI_MCP_URL}
    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = await client.beta.messages.create(
            model=ZI_MCP_MODEL or CLAUDE_MODEL, max_tokens=600, betas=[MCP_BETA],
            mcp_servers=[{"type": "url", "url": ZI_MCP_URL, "name": "zoominfo",
                          "authorization_token": token}],
            tools=[{"type": "mcp_toolset", "mcp_server_name": "zoominfo"}],
            messages=[{"role": "user", "content":
                       "List the ZoomInfo tools you can call. Names only, no commentary."}],
        )
    except Exception as e:
        return {"mcp_url": ZI_MCP_URL, "error": f"{type(e).__name__}: {str(e)[:600]}"}
    return {
        "mcp_url": ZI_MCP_URL,
        "model": ZI_MCP_MODEL or CLAUDE_MODEL,
        "stop_reason": getattr(msg, "stop_reason", ""),
        "block_types": [getattr(b, "type", "?") for b in msg.content],
        "text": " ".join(b.text for b in msg.content if getattr(b, "type", "") == "text")[:2000],
    }


# ------------------------- SEC EDGAR -------------------------
# Why this exists: signal A (age) is the single largest hole in the scoring
# model, and for officers and directors of public companies age is a *required
# public disclosure* — Regulation S-K Item 401 obliges the proxy statement to
# list names, ages and positions. That is a free, exact answer for exactly the
# segment the ICP targets, from a source that permits automated access.
#
# It answers for nobody else. A private-company owner or a long-tenured engineer
# will not be in a proxy statement, and this returns "not found" for them rather
# than a guess.

_edgar_last = [0.0]
_edgar_lock = asyncio.Lock()


async def _edgar_fetch(url: str):
    """One rate-limited EDGAR request, with the User-Agent the SEC asks for.

    Returns the raw response, whatever its status — the judgement about what a
    status means lives in _edgar_get, so the debug endpoint can show a failure
    verbatim instead of converting it into an exception page.
    """
    if not EDGAR_USER_AGENT:
        raise HTTPException(
            status_code=500,
            detail="EDGAR_USER_AGENT is not set. The SEC requires a descriptive "
                   "User-Agent with a contact email; see SETUP-edgar.md.",
        )
    # Serialised and spaced rather than burst-and-apologise: exceeding 10/sec
    # gets the whole IP range throttled, which would take the app down with it.
    async with _edgar_lock:
        gap = 1.0 / EDGAR_MAX_RPS
        wait = gap - (time.monotonic() - _edgar_last[0])
        if wait > 0:
            await asyncio.sleep(wait)
        _edgar_last[0] = time.monotonic()
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as cx:
        return await cx.get(url, headers={
            "User-Agent": EDGAR_USER_AGENT,
            "Accept-Encoding": "gzip, deflate",
        })


async def _edgar_get(url: str, as_json: bool = True):
    r = await _edgar_fetch(url)
    # Seen live: the identical query answered 200 with full results, then 500
    # with {"message": "Internal server error"} minutes later. EFTS throws
    # transient 5xxs, and during a sweep each one would otherwise turn into a
    # coverage gap on some lead. One retry, spaced by the same throttle.
    if r.status_code >= 500:
        r = await _edgar_fetch(url)
    if r.status_code >= 500:
        raise HTTPException(status_code=502,
                            detail=f"SEC internal error ({r.status_code}) — transient, "
                                   "try again shortly.")
    if r.status_code == 403:
        raise HTTPException(status_code=502,
                            detail="SEC returned 403 — usually a missing or rejected "
                                   "User-Agent. Check EDGAR_USER_AGENT names your firm "
                                   "and carries a contact email.")
    if r.status_code == 429:
        raise HTTPException(status_code=502,
                            detail="SEC rate-limited this address. Lower EDGAR_MAX_RPS and retry.")
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"SEC {r.status_code} for {url}: {r.text[:200]}")
    if not as_json:
        return r.text
    try:
        return r.json()
    except ValueError:
        raise HTTPException(status_code=502, detail=f"SEC returned non-JSON for {url}: {r.text[:200]}")


def _norm_company(s: str) -> str:
    """Company names for comparison — 'The Boeing Company' and 'Boeing' match."""
    t = (s or "").lower()
    for junk in (" incorporated", " corporation", " company", " holdings", " group",
                 " inc.", " inc", " corp.", " corp", " co.", " llc", " lp", " plc", " ltd"):
        t = t.replace(junk, " ")
    return " ".join(t.replace("the ", " ").replace(",", " ").replace(".", " ").split())


_edgar_tickers: dict = {}
_edgar_exact: dict = {}


async def _edgar_company_cik(employer: str) -> Optional[dict]:
    """CIK for an employer name, from the SEC's own published company list."""
    global _edgar_tickers, _edgar_exact
    if not _edgar_tickers:
        data = await _edgar_get(f"{EDGAR_WWW}/files/company_tickers.json")
        # Shape is {"0": {"cik_str": 320193, "ticker": "AAPL", "title": "Apple Inc."}, ...}
        # Every key holds a *list*. Stripping legal suffixes collapses "Acme
        # Industrial Corp" and "Acme Industrial Holdings" onto the same key, and
        # keeping only the first would quietly return the wrong company — and so
        # the wrong person's age. Collisions have to stay visible to be refused.
        for row in (data.values() if isinstance(data, dict) else data):
            title = (row or {}).get("title") or ""
            if not title:
                continue
            entry = {
                "cik": str(row.get("cik_str") or "").zfill(10),
                "name": title,
                "ticker": row.get("ticker") or "",
            }
            _edgar_tickers.setdefault(_norm_company(title), []).append(entry)
            _edgar_exact.setdefault(" ".join(title.lower().split()), entry)
    raw = " ".join((employer or "").lower().split())
    want = _norm_company(employer)
    if not want:
        return None
    # The full legal name is unambiguous by definition, so it is tried first —
    # otherwise "Acme Industrial Corp" would be refused for colliding with
    # "Acme Industrial Holdings" once the suffixes are stripped.
    if raw in _edgar_exact:
        return _edgar_exact[raw]

    def _one(rows):
        """A match only counts when it points at a single company."""
        if not rows:
            return None
        ciks = {r["cik"] for r in rows}
        return rows[0] if len(ciks) == 1 else None

    if want in _edgar_tickers:
        return _one(_edgar_tickers[want])
    # One fallback: a name that extends or is extended by what we were given.
    hits = [r for k, rows in _edgar_tickers.items()
            if k.startswith(want + " ") or want.startswith(k + " ")
            for r in rows]
    return _one(hits)


async def _edgar_latest_proxy(cik: str) -> Optional[dict]:
    """The most recent DEF 14A for a CIK, as {url, filed, accession}."""
    subs = await _edgar_get(f"{EDGAR_DATA}/submissions/CIK{cik}.json")
    recent = ((subs or {}).get("filings") or {}).get("recent") or {}
    forms = recent.get("form") or []
    for i, form in enumerate(forms):
        if form != "DEF 14A":
            continue
        acc = (recent.get("accessionNumber") or [])[i].replace("-", "")
        doc = (recent.get("primaryDocument") or [])[i]
        if not acc or not doc:
            continue
        return {
            "url": f"{EDGAR_WWW}/Archives/edgar/data/{int(cik)}/{acc}/{doc}",
            "filed": (recent.get("filingDate") or [])[i],
            "accession": (recent.get("accessionNumber") or [])[i],
        }
    return None


async def _edgar_recent_8k(cik: str, within_days: int = 45) -> Optional[dict]:
    """A recent 8-K reporting an officer departure, as {url, filed, accession}.

    Item 5.02 — departure or election of directors and principal officers — is
    a required disclosure with a four-business-day deadline, so it is the
    fastest free notice that a senior person's employment is ending. The item
    number is not in the submissions index, so the filing is fetched and read;
    that is one extra request per employer per refresh, not per lead.
    """
    subs = await _edgar_get(f"{EDGAR_DATA}/submissions/CIK{cik}.json")
    recent = ((subs or {}).get("filings") or {}).get("recent") or {}
    forms = recent.get("form") or []
    cutoff = time.time() - within_days * 86400
    for i, form in enumerate(forms):
        if form != "8-K":
            continue
        filed = (recent.get("filingDate") or [])[i]
        try:
            if time.mktime(time.strptime(filed, "%Y-%m-%d")) < cutoff:
                break            # the index is newest first, so we are past the window
        except (ValueError, TypeError):
            continue
        items = ((recent.get("items") or [])[i] if recent.get("items") else "") or ""
        if "5.02" not in items:
            continue
        acc = (recent.get("accessionNumber") or [])[i].replace("-", "")
        doc = (recent.get("primaryDocument") or [])[i]
        if not acc or not doc:
            continue
        url = f"{EDGAR_WWW}/Archives/edgar/data/{int(cik)}/{acc}/{doc}"
        summary = ""
        try:
            text = _strip_html(await _edgar_get(url, as_json=False) or "")
            m = re.search(r"(?is)item\s*5\.02[^.]{0,120}\.(.{0,400})", text)
            summary = (m.group(1).strip() if m else text[:400])
        except Exception:
            pass
        days_ago = int((time.time() - time.mktime(time.strptime(filed, "%Y-%m-%d"))) / 86400)
        return {"url": url, "filed": filed, "days_ago": days_ago,
                "accession": (recent.get("accessionNumber") or [])[i],
                "summary": summary or "Item 5.02 — departure or election of directors "
                                      "and principal officers"}
    return None


def _strip_html(html: str) -> str:
    """Filing text without the markup. Proxy statements are laid out by dozens of
    different filing agents, so no structural assumption survives — the text is
    flattened and the reading is left to Claude rather than to a regex that would
    silently match the wrong table cell."""
    out = re.sub(r"(?is)<(script|style).*?</\1>", " ", html)
    out = re.sub(r"(?s)<[^>]+>", " ", out)
    out = (out.replace("&nbsp;", " ").replace("&amp;", "&")
              .replace("&lt;", "<").replace("&gt;", ">").replace("&#160;", " "))
    return re.sub(r"\s+", " ", out).strip()


class EdgarRequest(BaseModel):
    first_name: str = ""
    last_name: str = ""
    employer: str = ""


@app.post("/api/edgar")
async def edgar(req: EdgarRequest, request: Request):
    """Age and role for one lead, from their employer's latest proxy statement.

    Free and unlimited — no credits, no vendor. It only answers for officers and
    directors of public companies, and says so plainly when it cannot.
    """
    await _active_token(request)
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500,
                            detail="Reading a proxy statement needs ANTHROPIC_API_KEY "
                                   "(the same key the AI QC button uses).")
    name = f"{req.first_name} {req.last_name}".strip()
    if not name or not req.employer.strip():
        raise HTTPException(status_code=400,
                            detail="A name and an employer are both required — "
                                   "the employer is how the filing is found.")

    company = await _edgar_company_cik(req.employer)
    if not company:
        return {"found": False,
                "reason": f"No public company on file matching “{req.employer}”. "
                          f"Private employers do not file proxy statements."}

    proxy = await _edgar_latest_proxy(company["cik"])
    if not proxy:
        return {"found": False, "company": company,
                "reason": f"{company['name']} has no DEF 14A on file."}

    text = await _edgar_get(proxy["url"], as_json=False)
    flat = _strip_html(text)
    # Proxy statements run long; the officer and director tables sit in the first
    # part, and this keeps the request affordable.
    flat = flat[:180_000]

    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    prompt = (
        "You are reading a US proxy statement (SEC form DEF 14A). Find this person "
        "and report only what the document actually states.\n\n"
        f"Person: {name}\nCompany: {company['name']}\n\n"
        "Reply with one JSON object and nothing else:\n"
        '{"found": true|false, "age": <integer or null>, "title": "<as printed or null>", '
        '"as_of": "<year the age was stated, or null>", "quote": "<the sentence or table row '
        'the age came from, max 200 chars, or null>"}\n\n'
        "Rules: only report an age printed in the document for THIS person. Never "
        "estimate, infer from career length, or carry an age across from a similarly "
        "named person. If the name does not appear, or appears without an age, return "
        'found=false. A wrong age here silently mis-scores a lead, so prefer '
        "found=false over a guess.\n\nDOCUMENT:\n" + flat
    )
    try:
        msg = await client.messages.create(
            model=CLAUDE_MODEL, max_tokens=700,
            messages=[{"role": "user", "content": prompt}],
        )
    except anthropic.APIStatusError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error {e.status_code}: {str(e)[:200]}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {str(e)[:200]}")

    body = "".join(b.text for b in msg.content if b.type == "text")
    a, z = body.find("{"), body.rfind("}")
    if a == -1 or z == -1:
        raise HTTPException(status_code=502, detail="Claude returned no JSON object.")
    try:
        got = json.loads(body[a:z + 1])
    except ValueError:
        raise HTTPException(status_code=502, detail="Claude returned malformed JSON.")

    age = got.get("age")
    if not isinstance(age, int) or not (18 <= age <= 100):
        age = None                       # a number outside a working life is a misread
    return {
        "found": bool(got.get("found")) and age is not None,
        "age": age,
        "title": got.get("title") or "",
        "as_of": got.get("as_of") or proxy["filed"][:4],
        "quote": (got.get("quote") or "")[:200],
        "company": company,
        "filing": proxy,
        "reason": "" if age else f"{name} is not listed with an age in that proxy statement.",
    }


PROXY_TTL = int(os.environ.get("PROXY_CACHE_SECONDS", str(7 * 86400)))


class RosterName(BaseModel):
    i: int = 0
    first_name: str = ""
    last_name: str = ""


class RosterRequest(BaseModel):
    employer: str = ""
    people: list[RosterName] = []


async def _proxy_roster(employer: str) -> dict:
    """Every person a company's latest proxy states an age for, read once.

    /api/edgar asks a 45,000-token document one question per lead, which is why
    proxy ages were kept out of any list-scale action: ten leads at one employer
    meant ten reads of one file. But the document answers for the whole board at
    once — it is a table of directors and officers with their ages — so reading
    it per *employer* costs one AI call however many leads work there, and the
    result is cached for a week because a proxy is filed annually.

    Only structural answers are cached (a roster, no such company, no DEF 14A on
    file). A transient failure is raised, never stored: a cached outage would
    outlive the outage.
    """
    company = await _edgar_company_cik(employer)
    if not company:
        return {"found": False, "people": [],
                "reason": f"No public company on file matching \u201c{employer}\u201d. "
                          f"Private employers do not file proxy statements."}
    key = "proxy:" + company["cik"]
    hit = _cache_get(key, PROXY_TTL)
    if hit is not None:
        return hit

    proxy = await _edgar_latest_proxy(company["cik"])
    if not proxy:
        return _cache_put(key, {"found": False, "people": [], "company": company,
                                "reason": f"{company['name']} has no DEF 14A on file."})

    text = await _edgar_get(proxy["url"], as_json=False)
    flat = _strip_html(text)[:180_000]
    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    prompt = (
        "You are reading a US proxy statement (SEC form DEF 14A). List every "
        "person whose age the document states — directors, director nominees "
        "and named executive officers.\n\n"
        f"Company: {company['name']}\n\n"
        "Reply with one JSON object and nothing else:\n"
        '{"people": [{"name": "<as printed>", "age": <integer>, "title": "<as printed, or empty>"}]}\n\n'
        "Rules: include a person only when the document prints an age for them "
        "(commonly a table column, or \u201cAge: 57\u201d beside a biography). Never "
        "estimate an age, infer one from career length, or carry one across from "
        "a similarly named person. Return {\"people\": []} when the document "
        "states no ages. A wrong age here silently mis-scores a lead, so omit "
        "anyone you are unsure of.\n\nDOCUMENT:\n" + flat
    )
    try:
        msg = await client.messages.create(
            model=CLAUDE_MODEL, max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
    except anthropic.APIStatusError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error {e.status_code}: {str(e)[:200]}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {str(e)[:200]}")

    body = "".join(b.text for b in msg.content if b.type == "text")
    a, z = body.find("{"), body.rfind("}")
    if a == -1 or z == -1:
        raise HTTPException(status_code=502, detail="Claude returned no JSON object.")
    try:
        got = json.loads(body[a:z + 1])
    except ValueError:
        raise HTTPException(status_code=502, detail="Claude returned malformed JSON.")

    people = freesources.roster_people(got)
    return _cache_put(key, {
        "found": bool(people), "people": people, "company": company, "filing": proxy,
        "reason": "" if people
                  else f"{company['name']}\u2019s latest proxy statement states no ages.",
    })


@app.post("/api/edgar-roster")
async def edgar_roster(req: RosterRequest, request: Request):
    """Proxy-statement ages for everyone on a list who works at one employer.

    The batch shape is the whole point: this is what lets the one free-enrichment
    button include proxy ages at list scale, where the per-lead reader could not
    be afforded. Matching stays here rather than in the browser so the namesake
    guard is the same tested one the FEC and insider searches use — and so an
    ambiguous roster returns nothing rather than a plausible wrong age.
    """
    await _active_token(request)
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500,
                            detail="Reading a proxy statement needs ANTHROPIC_API_KEY "
                                   "(the same key the AI QC button uses).")
    if not req.employer.strip():
        raise HTTPException(status_code=400,
                            detail="An employer is required — it is how the filing is found.")

    roster = await _proxy_roster(req.employer)
    matches = {}
    for who in req.people:
        if not who.last_name.strip():
            continue
        hit = freesources.roster_match(roster.get("people") or [],
                                       who.last_name, who.first_name)
        if hit:
            matches[str(who.i)] = {
                "age": hit["age"], "title": hit["title"],
                "as_of": (roster.get("filing") or {}).get("filed", "")[:4],
                "name": hit["name"],
            }
    return {
        "found": roster.get("found", False),
        "company": roster.get("company") or {},
        "filing": roster.get("filing") or {},
        "roster_size": len(roster.get("people") or []),
        "matches": matches,
        "reason": roster.get("reason", ""),
    }


@app.get("/api/edgar-debug")
async def edgar_debug(request: Request, url: str = ""):
    """Raw EDGAR round-trip — the URL called, the status, the first of the body.

    EDGAR is unreachable from the environment this was written in, so no response
    from it has ever been seen here. This is the probe that will show what it
    actually returns.
    """
    await _active_token(request)
    if not EDGAR_USER_AGENT:
        return {"error": "EDGAR_USER_AGENT is not set on this service."}
    target = url or f"{EDGAR_WWW}/files/company_tickers.json"
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as cx:
        try:
            r = await cx.get(target, headers={"User-Agent": EDGAR_USER_AGENT})
        except Exception as e:
            return {"url": target, "error": f"{type(e).__name__}: {e}"}
    return {"url": target, "status": r.status_code,
            "content_type": r.headers.get("content-type", ""),
            "body": r.text[:4000]}


async def _credit_block(who: str = "") -> dict:
    """What this person has left, in the shape the browser renders.

    Returned by /api/me (once per page load, no extra request) and by the
    credits endpoint, so the number a user is looking at was true as of the
    last thing they did rather than as of the last time they thought to ask.

    WhitePages is an allowance because the app holds the key and pays the bill.
    ZoomInfo is not: every user brings their own subscription, so what is
    reported there is usage — what this app spent on their behalf, for
    reconciling against their own dashboard — with no ceiling of ours to hit.
    """
    if not (WHITEPAGES_API_KEY or ZI_CLIENT_ID or ANTHROPIC_API_KEY):
        return {}
    mine = await _ledger_read(who) if who else {"wp": 0, "zi": 0}
    room = await _wp_left(who)
    return {
        "month": _month(), "resets_on": _month_resets(),
        "whitepages": {"spent": mine["wp"], "budget": WP_PER_USER,
                       "left": room["left"], "yours_left": room["mine"],
                       "firm_left": room["firm"], "firm_budget": WP_MONTHLY,
                       "capped_by": room["capped_by"]},
        "zoominfo": {"used": mine["zi"], "own_subscription": True},
        "cache_entries": len(_WP_CACHE),
        "cache_days": round(WP_TTL / 86400),
    }


class SpendReport(BaseModel):
    kind: str = ""          # "zi"
    n: int = 0


@app.get("/api/credits")
async def credits(request: Request):
    """Both paid allowances: what is spent this month, and what is left.

    Counted at the point of spending and stored per calendar month, so it
    survives the instance recycles that used to reset the old counter to zero
    every time traffic paused.
    """
    await _active_token(request)
    return await _credit_block(await _who(request))


@app.post("/api/credits")
async def credits_spend(req: SpendReport, request: Request):
    """Record ZoomInfo enrichments this app ran on the user's own subscription.

    Two of the three routes to ZoomInfo go through the user's own Claude
    connector and never touch this server, so it cannot see them. This is not
    a budget — the subscription is theirs and its ceiling is theirs — it is
    usage, kept so they can reconcile what this app did against their own
    dashboard.

    WhitePages is not accepted here. That one the server spends itself, at a
    single choke point, and a client-reported number for it would be a number
    the server could have counted correctly.
    """
    await _active_token(request)
    if req.kind != "zi":
        raise HTTPException(
            status_code=400,
            detail="Only ZoomInfo usage is self-reported; WhitePages is counted "
                   "where it is spent.")
    who = await _who(request)
    n = max(0, min(int(req.n or 0), 10000))
    await _ledger_add("zi", n, who)
    return {"used": (await _ledger_read(who))["zi"], "own_subscription": True}


@app.get("/api/wp-spend")
async def wp_spend(request: Request):
    """What this service has actually spent, and what it has saved.

    Counted, not estimated: every number here is incremented at the point the
    call is made or avoided. Deliberately not a call to the account usage
    endpoint, which is itself billed — asking how much you have spent should
    not spend any.
    """
    await _active_token(request)
    total = WP_SPEND["calls"] + WP_SPEND["served_from_cache"]
    return {
        "billed_calls": WP_SPEND["calls"],
        "answered_from_cache": WP_SPEND["served_from_cache"],
        "refused_before_sending": WP_SPEND["refused"],
        "saved_pct": round(100 * WP_SPEND["served_from_cache"] / total) if total else 0,
        "cache_entries": len(_WP_CACHE),
        "cache_seconds": WP_TTL,
        # Since this process started. Cloud Run recycles instances, so it is a
        # recent picture rather than a lifetime bill — the account dashboard is
        # the authority on that.
        "since": "this server instance started",
    }


@app.get("/api/wp-debug")
async def wp_debug(request: Request, phone: str = "", name: str = "", path: str = ""):
    """Raw WhitePages round-trip, for diagnosing an integration that returns
    nothing. Shows the exact URL called, the status, and the untouched body.

    Signed-in only, same as every other lookup — it spends a credit when the
    call actually lands. Visit /api/wp-debug?phone=6313121293 in the browser.
    """
    await _active_token(request)
    if not WHITEPAGES_API_KEY:
        return {"error": "WHITEPAGES_API_KEY is not set on this service."}
    params = {}
    if phone:
        params["phone"] = _digits(phone)
    if name:
        params["name"] = name
    if not params:
        return {"error": "Pass ?phone= or ?name="}
    url = WHITEPAGES_BASE_URL + ("/" + path.lstrip("/") if path else _wp_path("phone"))
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as cx:
        r = await cx.get(url, params=params, headers={"X-Api-Key": WHITEPAGES_API_KEY})
    body = r.text
    try:
        parsed = r.json()
    except ValueError:
        parsed = None
    return {
        "url": str(r.url),
        "status": r.status_code,
        "key_len": len(WHITEPAGES_API_KEY),      # confirms a key is present, not its value
        # Which fields THIS account's responses actually carry. Four parsers in
        # this app were written from documentation and all four were wrong
        # about the real data, so the question "does my key return a date of
        # birth" deserves an answer that is one line long instead of four
        # thousand characters of JSON to read by eye.
        "fields": _key_census(parsed) if parsed is not None else [],
        "read": _person_facts(_best_person(parsed)) if parsed is not None else {},
        "body": body[:4000],
    }


def _key_census(node, prefix: str = "", out=None, depth: int = 0) -> list:
    """Every field path in a response, with its type and a short sample.

    Lists are described by their first element rather than every element, so
    a hundred-address record still summarises in a screenful. Samples are
    truncated: this is for identifying the shape, and the untouched body is
    right below it for anyone who wants the values.
    """
    out = [] if out is None else out
    if depth > 5 or len(out) > 300:
        return out
    if isinstance(node, dict):
        for k, v in node.items():
            _key_census(v, f"{prefix}.{k}" if prefix else str(k), out, depth + 1)
    elif isinstance(node, list):
        out.append(f"{prefix}[] — {len(node)} item(s)")
        if node:
            _key_census(node[0], f"{prefix}[0]", out, depth + 1)
    elif node is None:
        out.append(f"{prefix} = null")
    else:
        s = str(node)
        out.append(f"{prefix} = {s[:60]}" + ("…" if len(s) > 60 else ""))
    return out


class FreeEnrichRequest(BaseModel):
    first_name: str = ""
    last_name: str = ""
    employer: str = ""


async def _fec_get(params: dict):
    """One FEC call. The API is free; the failure modes are the key and the
    shared demo rate limit, and each gets its own sentence."""
    q = dict(params)
    q["api_key"] = FEC_API_KEY
    async with httpx.AsyncClient(timeout=30) as cx:
        r = await cx.get(f"{FEC_API_BASE}/schedules/schedule_a/", params=q)
    if r.status_code == 403:
        raise HTTPException(status_code=502,
                            detail="FEC rejected the API key — set FEC_API_KEY "
                                   "(free at api.open.fec.gov/developers).")
    if r.status_code == 429:
        raise HTTPException(status_code=502,
                            detail="FEC rate limit hit. The shared DEMO_KEY allows 40 "
                                   "lookups an hour across everyone using it — a free "
                                   "personal FEC_API_KEY raises that to 1,000.")
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"FEC error {r.status_code}: {r.text[:200]}")
    try:
        return r.json()
    except ValueError:
        raise HTTPException(status_code=502, detail="FEC returned non-JSON.")


@app.post("/api/free-enrich")
async def free_enrich(req: FreeEnrichRequest, request: Request):
    """Everything the free public records say about one person.

    Two sources, each optional, and the response says which ran — the same
    coverage honesty the signals panel got in §26's predecessor: an empty
    answer from a check that never ran must not read as an all-clear.

    Costs nothing. The FEC search itself is free, and the SEC search rides the
    same rate-limited client as every other EDGAR call in the app.
    """
    await _active_token(request)
    last = req.last_name.strip()
    if not last:
        raise HTTPException(status_code=400, detail="A last name is required.")
    name = f"{req.first_name.strip()} {last}".strip()

    sources = {}
    donations, filings, insiders = {}, [], []

    try:
        payload = await _fec_get({"contributor_name": name, "per_page": 100,
                                  "sort": "-contribution_receipt_date"})
        rows = freesources.match_rows(freesources.fec_rows(payload),
                                      last, req.first_name)
        donations = freesources.summarize_fec(rows, req.employer)
        sources["fec"] = {"ran": True,
                          "note": "" if rows else "no itemised contributions under this name"}
    except HTTPException as e:
        sources["fec"] = {"ran": False, "reason": e.detail}

    if EDGAR_USER_AGENT:
        try:
            # Unquoted on purpose. EDGAR writes people surname-first — Tim
            # Cook's filings say "COOK TIMOTHY D" — and a quoted phrase is
            # order-sensitive, so it misses real officers filed the other way
            # round. Unquoted means every word must appear, in any order; the
            # surname guard in match_filings does the disambiguation a phrase
            # was doing, without the false negatives.
            payload = await _edgar_get(
                EFTS_URL + "?" + urlencode({"q": name, "forms": "3,4,5"}))
            hits = freesources.efts_hits(payload)
            filings = freesources.match_filings(hits, last, req.first_name)[:8]
            insiders = freesources.match_entities(
                freesources.efts_entities(payload), last, req.first_name)[:3]
            sources["edgar"] = {"ran": True,
                                "note": "" if (filings or insiders)
                                        else "no insider filings under this name"}
        except HTTPException as e:
            sources["edgar"] = {"ran": False, "reason": e.detail}
    else:
        sources["edgar"] = {"ran": False,
                            "reason": "EDGAR_USER_AGENT not set — see SETUP-edgar.md"}

    return {
        "sources": sources,
        "donations": donations,
        "filings": filings,
        # The person-level verdict from the response's own aggregation — set
        # only when the EDGAR search ran; [] then means "searched, not an
        # insider", which is an answer, not an absence.
        "insider_entities": insiders if sources.get("edgar", {}).get("ran") else [],
        # Deep links to the same searches on the source's own site, so every
        # number above can be checked by a person in one click.
        "links": {
            "fec": "https://www.fec.gov/data/receipts/individual-contributions/"
                   f"?{urlencode({'contributor_name': name})}",
            "edgar": "https://www.sec.gov/edgar/search/#/"
                     f"q={name.replace(' ', '%20')}&forms=3,4,5",
        },
    }


@app.get("/api/free-debug")
async def free_debug(request: Request, source: str = "fec", name: str = ""):
    """Raw round-trip against one free source, with the field census.

    Both parsers were written without a live response — the build environment
    cannot reach either host — which is the condition this app has been wrong
    under five times. One call to this endpoint from the deployed app is the
    line of truth that settles it.
    """
    await _active_token(request)
    if not name:
        return {"error": "Pass ?name=First Last"}
    if source == "fec":
        try:
            payload = await _fec_get({"contributor_name": name, "per_page": 3})
        except HTTPException as e:
            return {"source": source, "error": e.detail,
                    "fec_key": "personal" if FEC_API_KEY != "DEMO_KEY" else "DEMO_KEY"}
        parsed = freesources.fec_rows(payload)
    elif source == "efts":
        # Raw round-trip on purpose: when the SEC refuses, the refusal itself —
        # status, headers' story, body — is the diagnostic, and an exception
        # page that says "does not work" hides all three.
        # Unquoted for the same word-order reason as the live route.
        url = EFTS_URL + "?" + urlencode({"q": name, "forms": "3,4,5"})
        if not EDGAR_USER_AGENT:
            return {"source": source, "url": url, "ua_set": False,
                    "error": "EDGAR_USER_AGENT is not set on this service. The SEC "
                             "requires a descriptive User-Agent with a contact email "
                             "(e.g. 'Financial Planners of America "
                             "dst@financialplannersofamerica.com'). Cloud Run -> Edit & "
                             "deploy new revision -> Variables. See SETUP-edgar.md."}
        r = await _edgar_fetch(url)
        try:
            payload = r.json()
        except ValueError:
            payload = None
        out = {"source": source, "url": url, "status": r.status_code,
               "ua_set": True, "ua_len": len(EDGAR_USER_AGENT)}
        if r.status_code != 200 or payload is None:
            out["error"] = ("The SEC refused this request — a 403 here usually means "
                            "the User-Agent lacks a firm name and contact email."
                            if r.status_code == 403 else
                            "Non-JSON or non-200 answer; the body below is the SEC's own words.")
            out["body"] = r.text[:2000]
            return out
        out["fields"] = _key_census(payload)
        # read is the raw parsed hits, BEFORE the who-is-it guard — prose
        # mentions of the name in other people's paperwork appear here on
        # purpose. entities is the aggregation the verdict comes from.
        out["read"] = freesources.efts_hits(payload)[:3]
        out["entities"] = freesources.efts_entities(payload)[:5]
        return out
    else:
        return {"error": "source must be fec or efts"}
    return {"source": source, "fields": _key_census(payload), "read": parsed[:3]}


class EnrichRequest(BaseModel):
    first_name: str = ""
    last_name: str = ""
    middle_name: str = ""
    city: str = ""
    state: str = ""
    street: str = ""
    zip: str = ""
    phone: str = ""
    email: str = ""
    # The property lookup is a second billed call and answers a different
    # question, so it is asked for rather than assumed. Off by default: most
    # presses want to know who someone is, not who holds the deed.
    want_property: bool = False
    fresh: bool = False


@app.post("/api/enrich")
async def enrich(req: EnrichRequest, request: Request):
    """Household facts for one lead: where they live, what they own, how to reach them.

    Costs one person lookup, plus one property lookup when an address comes
    back — so it stays a per-lead button rather than a bulk sweep.
    """
    await _active_token(request)
    name = f"{req.first_name} {req.last_name}".strip()
    if not name:
        raise HTTPException(status_code=400, detail="A name is required to enrich a lead.")

    want_state = _state_code(req.state)
    person, matched_by = {}, ""
    steps: list = []          # what was actually spent, in order

    # The ladder is ordered by how well each thing identifies a person, and it
    # stops at the first answer. Every rung is a billed call, so a rung is only
    # climbed when the one below it found nothing and the next one has
    # something real to search on.
    #
    # A phone number identifies a person; a name does not. "Daniel Treacy"
    # returns a 30-year-old in Kansas City and a realtor in Portland, and the
    # highest match_score is the realtor.
    last = req.last_name.strip().lower()

    def _mine(p) -> bool:
        """Whether a record plausibly belongs to this lead."""
        if not p:
            return False
        if not last:
            return True
        if last in (_first_str(p, "name", "full_name") or "").lower():
            return True
        return any(last in a.lower() for a in _aliases(p))

    if req.phone and WP_PHONE_RE.match(_digits(req.phone)):
        steps.append("phone")
        p, _hit = await _wp_phone(_digits(req.phone), req.fresh, who=await _who(request))
        if _mine(p):
            person, matched_by = p, "phone"

    # Not every lead has a mobile, and an email address identifies a person
    # nearly as well as a number does — better than a name, because nobody
    # shares one. This rung is why a lead with no phone is now worth a press.
    if not person and req.email and WP_EMAIL_RE.match(req.email.strip()):
        steps.append("email")
        p = _best_person(await _wp_get("person", {"email": req.email.strip()}, who=await _who(request),
                                       fresh=req.fresh),
                         require_last=req.last_name)
        if _mine(p):
            person, matched_by = p, "email"

    if not person and (req.last_name.strip() or name):
        # Individual name fields are documented as matching each part
        # specifically, where `name` matches loosely against the whole field.
        # They cannot be combined, so it is one or the other.
        params = {"strict_match": "true"}
        if req.last_name.strip():
            params["last_name"] = req.last_name.strip()
            if req.first_name.strip():
                params["first_name"] = req.first_name.strip()
            if req.middle_name.strip():
                params["middle_name"] = req.middle_name.strip()
        else:
            params["name"] = name
        if req.city:
            params["city"] = req.city
        if want_state:
            params["state_code"] = want_state
        if req.street:
            params["street"] = req.street
        if WP_ZIP_RE.match((req.zip or "").strip()):
            params["zipcode"] = req.zip.strip()
        # A name with no location at all is the query that returns the realtor
        # in Portland. Refuse it rather than pay for it.
        if not any(k in params for k in ("city", "state_code", "street", "zipcode")):
            return {"found": False, "steps": steps,
                    "rejected": "a name with no city, state or ZIP matches too many "
                                "people to be worth a lookup — add a location"}
        steps.append("name")
        person = _best_person(await _wp_get("person", params, fresh=req.fresh, who=await _who(request)),
                              require_last=req.last_name)
        matched_by = "name"

    if not person:
        return {"found": False, "steps": steps}

    # Namesakes are the failure mode here, so a record from a different state
    # than the lead is refused rather than silently attributed to them.
    got_state = _state_code(_home_address(person).get("state", ""))
    if matched_by == "name" and want_state and got_state and got_state != want_state:
        return {"found": False,
                "rejected": f"closest match lives in {got_state}, lead is in {want_state}"}

    addr = _home_address(person)
    mobiles = _mobiles(person)

    # Ownership stands in for net worth here: this API carries no home value,
    # but who holds the deed is itself a strong signal — a house in a trust or
    # an LLC means someone has already done estate or entity planning.
    owns_home = None
    owner_type = ""
    co_owners: list = []
    if req.want_property and addr["street"] and (addr["city"] or addr["zip"]):
        prop_params = {"street": addr["street"]}
        if addr["city"]:
            prop_params["city"] = addr["city"]
        if _state_code(addr["state"]):
            prop_params["state_code"] = _state_code(addr["state"])
        if addr["zip"]:
            prop_params["zipcode"] = addr["zip"]
        steps.append("property")
        prop = await _wp_get("property", prop_params, fresh=req.fresh, who=await _who(request))
        result = (prop or {}).get("result") if isinstance(prop, dict) else None
        if isinstance(result, dict):
            info = result.get("ownership_info") or {}
            owner_type = _first_str(info, "owner_type")
            names = [_first_str(o, "name") for o in (info.get("person_owners") or [])
                     if isinstance(o, dict)]
            biz = [_first_str(o, "name") for o in (info.get("business_owners") or [])
                   if isinstance(o, dict)]
            last = req.last_name.strip().lower()
            owns_home = bool(last) and any(last in n.lower() for n in names if n)
            if biz:
                owner_type = owner_type or "entity"
            co_owners = [n for n in names + biz if n]

    # Everything the record carries, read once and shared with the phone check
    # so the two buttons cannot disagree about the same person. The deeds are
    # the part worth reading twice: each address is a property to value, and a
    # second one says more about net worth than any title ever will.
    out = _person_facts(person)
    out.update({
        "found": True,
        "matched_by": matched_by,
        "match_score": person.get("match_score"),
        "mobile_count": len(mobiles),
        "owns_home": owns_home,
        "owner_type": owner_type,
        "co_owners": co_owners,
        # Which rungs were climbed. The UI shows this so a press that cost two
        # lookups says so instead of looking the same as one that cost none.
        "steps": steps,
        "property_checked": bool(req.want_property),
    })
    return out


def _property_addresses(person: dict) -> list:
    out = []
    for p in person.get("owned_properties") or []:
        if not isinstance(p, dict):
            continue
        a = _first_str(p, "address", "full_address")
        if not a:
            a = ", ".join(x for x in _addr(p).values() if x)
        if a and a not in out:
            out.append(a)
    return out[:6]


def _prior_places(person: dict) -> list:
    """Previous cities, deduplicated — a mobility picture, not a mailing list.

    Deliberately not used to widen matching: the Portland namesake in the live
    data had a historic address in the same state as the real lead, so a
    "lived there once" rule would have accepted exactly the record the
    current-state check exists to reject.
    """
    out = []
    for a in person.get("historic_addresses") or []:
        parts = _addr(a)
        place = ", ".join(x for x in (parts["city"], parts["state"]) if x)
        if place and place not in out:
            out.append(place)
    return out[:6]


class LeadState(BaseModel):
    settings: dict = {}
    leads: list = []


@app.get("/api/state")
async def get_state(request: Request, email: str = Depends(signed_in)):
    """The signed-in user's saved list. 401 keeps the browser on localStorage."""
    doc = await _fs_get(FS_STATE, email) or _MEM_STATE.get(email)
    if not doc:
        return {"found": False, "settings": {}, "leads": [], "backend": storage_backend()}
    try:
        payload = json.loads(doc.get("data") or "{}")
    except ValueError:
        raise HTTPException(status_code=500, detail="Saved list is corrupt.")
    return {"found": True, "settings": payload.get("settings") or {},
            "leads": payload.get("leads") or [], "saved_at": doc.get("saved_at"),
            "backend": storage_backend()}


@app.put("/api/state")
async def put_state(body: LeadState, request: Request, email: str = Depends(signed_in)):
    payload = {"data": json.dumps({"settings": body.settings, "leads": body.leads}),
               "saved_at": time.time(), "lead_count": len(body.leads)}
    if not await _fs_set(FS_STATE, email, payload):
        _MEM_STATE[email] = payload
    return {"ok": True, "leads": len(body.leads), "backend": storage_backend()}


# ------------------------- named lead lists -------------------------
# One user, several lists: a rollover campaign, an SCS campaign, the leads that
# came off one employer's WARN notice. Each is a separate document keyed
# email__listId, and the index of them lives on the user's state document, so
# opening the app reads a short index rather than every lead the user owns.

_MEM_LISTS: dict[str, dict] = {}
_MEM_SHARED: dict[str, dict] = {}
_MEM_STATS: dict[str, dict] = {}
_MEM_BATTLES: dict[str, dict] = {}
LIST_ID_RE = re.compile(r"^[A-Za-z0-9_-]{1,40}$")


def _list_key(email: str, list_id: str) -> str:
    return f"{email}__{list_id}"


def _check_list_id(list_id: str) -> str:
    if not LIST_ID_RE.match(list_id or ""):
        raise HTTPException(status_code=400, detail="Bad list id.")
    return list_id


async def _read_index(email: str) -> dict:
    doc = await _fs_get(FS_STATE, email) or _MEM_STATE.get(email) or {}
    try:
        payload = json.loads(doc.get("data") or "{}")
    except ValueError:
        payload = {}
    return {"settings": payload.get("settings") or {},
            "lists": payload.get("lists") or [],
            "legacy_leads": payload.get("leads") or [],
            "saved_at": doc.get("saved_at")}


async def _write_index(email: str, settings: dict, lists: list, legacy: list) -> None:
    payload = {"data": json.dumps({"settings": settings, "lists": lists, "leads": legacy}),
               "saved_at": time.time(), "lead_count": sum(l.get("count", 0) for l in lists)}
    if not await _fs_set(FS_STATE, email, payload):
        _MEM_STATE[email] = payload
    await _touch_directory(email, lists)


async def _touch_directory(email: str, lists: list) -> None:
    """Record that this advisor exists, and how much they are carrying.

    Written here rather than at sign-in because this is the moment the counts
    change, and the counts are what an admin actually reads. A first page load
    creates the default list, so everyone lands in the directory on day one.
    """
    if not email:
        return
    row = {"email": email.lower(), "lists": len(lists),
           "leads": sum(int(l.get("count") or 0) for l in lists),
           "last_seen": time.time()}
    prior = await _fs_get(FS_DIRECTORY, email.lower()) or _MEM_DIRECTORY.get(email.lower()) or {}
    row["first_seen"] = prior.get("first_seen") or row["last_seen"]
    if not await _fs_set(FS_DIRECTORY, email.lower(), row):
        _MEM_DIRECTORY[email.lower()] = row


async def _directory() -> list:
    """Every advisor, newest activity first. Firestore when it is there."""
    db = _firestore()
    rows = []
    if db is not None:
        try:
            async for snap in db.collection(FS_DIRECTORY).stream():
                d = snap.to_dict() or {}
                if d.get("email"):
                    rows.append(d)
        except Exception as e:
            print(f"[storage] directory read failed, using memory: {e}")
            rows = []
    if not rows:
        rows = list(_MEM_DIRECTORY.values())
    rows.sort(key=lambda r: r.get("last_seen") or 0, reverse=True)
    return rows


async def _read_list(email: str, list_id: str) -> list:
    key = _list_key(email, list_id)
    doc = await _fs_get(FS_LISTS, key) or _MEM_LISTS.get(key)
    if not doc:
        return []
    try:
        return json.loads(doc.get("data") or "[]")
    except ValueError:
        return []


async def _write_list(email: str, list_id: str, leads: list) -> None:
    key = _list_key(email, list_id)
    payload = {"data": json.dumps(leads), "saved_at": time.time(), "lead_count": len(leads)}
    if not await _fs_set(FS_LISTS, key, payload):
        _MEM_LISTS[key] = payload


async def _ensure_lists(email: str) -> dict:
    """The user's list index, migrating a pre-lists single list on first read."""
    idx = await _read_index(email)
    if idx["lists"]:
        return idx
    legacy = idx["legacy_leads"]
    # id "default" IS the master list. Every lead that enters the app lands on
    # it (the client appends imports made into campaign lists), it cannot be
    # deleted, and a restore merges into it instead of replacing it. Existing
    # users keep whatever they renamed theirs to; the id is the identity.
    first = {"id": "default", "name": "All leads", "created_at": time.time(),
             "count": len(legacy)}
    if legacy:
        await _write_list(email, "default", legacy)
    # The legacy array is left in place rather than deleted. If this migration
    # is wrong the original is still there to read; nothing writes to it again.
    await _write_index(email, idx["settings"], [first], legacy)
    idx["lists"] = [first]
    return idx


class ListCreate(BaseModel):
    name: str
    copy_from: str = ""


class ListPatch(BaseModel):
    name: str


class ListSave(BaseModel):
    leads: list


# ------------------------- sharing -------------------------
# A list belongs to the advisor who built it. Sharing grants a named colleague
# access to that one list, and nothing else — there is no firm-wide pool, and
# no way to enumerate what has not been shared with you.

def _domain(email: str) -> str:
    return email.split("@", 1)[1].lower() if "@" in email else ""


async def _shared_index(email: str) -> list:
    doc = await _fs_get(FS_SHARED, email) or _MEM_SHARED.get(email) or {}
    try:
        return json.loads(doc.get("data") or "[]")
    except ValueError:
        return []


async def _write_shared_index(email: str, rows: list) -> None:
    payload = {"data": json.dumps(rows), "saved_at": time.time()}
    if not await _fs_set(FS_SHARED, email, payload):
        _MEM_SHARED[email] = payload


async def _grant(owner: str, list_id: str, name: str, to: str, role: str) -> None:
    rows = [r for r in await _shared_index(to)
            if not (r.get("owner") == owner and r.get("id") == list_id)]
    rows.append({"owner": owner, "id": list_id, "name": name, "role": role,
                 "shared_at": time.time()})
    await _write_shared_index(to, rows)


async def _revoke(owner: str, list_id: str, frm: str) -> None:
    rows = [r for r in await _shared_index(frm)
            if not (r.get("owner") == owner and r.get("id") == list_id)]
    await _write_shared_index(frm, rows)


async def _access(email: str, owner: str, list_id: str) -> str:
    """'owner', 'editor', 'viewer' or '' for the caller on one list."""
    if owner == email:
        return "owner"
    # An admin can open and work any list in the firm. Deliberately "editor"
    # and not "owner": ownership carries the right to delete and to re-share,
    # and an admin who wants a list for good takes it explicitly, by reclaiming
    # it, which leaves a record on the list itself.
    if _is_admin(email):
        return "editor"
    for r in await _shared_index(email):
        if r.get("owner") == owner and r.get("id") == list_id:
            return r.get("role") or "viewer"
    return ""


def _split_ref(list_id: str) -> tuple[str, str]:
    """'someone@firm.com~listid' addresses a shared list; a bare id is your own."""
    if "~" in list_id:
        owner, _, rid = list_id.partition("~")
        return owner.lower(), rid
    return "", list_id


class ShareRequest(BaseModel):
    email: EmailStr
    role: str = "editor"


@app.get("/api/lists")
async def get_lists(request: Request, email: str = Depends(signed_in)):
    idx = await _ensure_lists(email)
    mine = [dict(l, owner="", role="owner", master=(l["id"] == "default"))
            for l in idx["lists"]]
    # Lists other advisors have shared, addressed as owner~id so the two kinds
    # can live in one switcher without their ids colliding.
    shared = []
    for r in await _shared_index(email):
        shared.append({"id": f"{r['owner']}~{r['id']}", "name": r.get("name") or "Shared list",
                       "owner": r["owner"], "role": r.get("role") or "viewer",
                       "count": r.get("count", 0), "shared_at": r.get("shared_at")})
    return {"lists": mine + shared, "settings": idx["settings"],
            "backend": storage_backend()}


@app.post("/api/lists")
async def create_list(body: ListCreate, request: Request, email: str = Depends(signed_in)):
    name = (body.name or "").strip()[:60]
    if not name:
        raise HTTPException(status_code=400, detail="A list needs a name.")
    idx = await _ensure_lists(email)
    if len(idx["lists"]) >= 40:
        raise HTTPException(status_code=400, detail="40 lists is the limit — rename or delete one.")
    new_id = "l" + secrets.token_urlsafe(9).replace("-", "_")
    leads = await _read_list(email, _check_list_id(body.copy_from)) if body.copy_from else []
    if leads:
        await _write_list(email, new_id, leads)
    entry = {"id": new_id, "name": name, "created_at": time.time(), "count": len(leads)}
    lists = idx["lists"] + [entry]
    await _write_index(email, idx["settings"], lists, idx["legacy_leads"])
    return {"list": entry, "lists": lists}


@app.get("/api/lists/{list_id}")
async def read_list(list_id: str, request: Request, email: str = Depends(signed_in)):
    owner, rid = _split_ref(list_id)
    _check_list_id(rid)
    if owner:
        role = await _access(email, owner, rid)
        if not role:
            raise HTTPException(status_code=403, detail="That list is not shared with you.")
        oidx = await _read_index(owner)
        entry = next((l for l in oidx["lists"] if l["id"] == rid), None)
        if not entry:
            raise HTTPException(status_code=404, detail="The owner has deleted that list.")
        return {"list": dict(entry, id=list_id, owner=owner, role=role),
                "leads": await _read_list(owner, rid),
                "settings": (await _ensure_lists(email))["settings"],
                "backend": storage_backend()}
    idx = await _ensure_lists(email)
    entry = next((l for l in idx["lists"] if l["id"] == rid), None)
    if not entry:
        raise HTTPException(status_code=404, detail="No such list.")
    return {"list": dict(entry, role="owner"), "leads": await _read_list(email, rid),
            "settings": idx["settings"], "backend": storage_backend()}


@app.put("/api/lists/{list_id}")
async def save_list(list_id: str, body: ListSave, request: Request, email: str = Depends(signed_in)):
    owner, rid = _split_ref(list_id)
    _check_list_id(rid)
    if owner:
        role = await _access(email, owner, rid)
        if role != "editor":
            raise HTTPException(
                status_code=403,
                detail="You have view-only access to that list — ask the owner for editing.")
        oidx = await _read_index(owner)
        if not any(l["id"] == rid for l in oidx["lists"]):
            raise HTTPException(status_code=404, detail="The owner has deleted that list.")
        await _write_list(owner, rid, body.leads)
        for l in oidx["lists"]:
            if l["id"] == rid:
                l["count"] = len(body.leads)
                l["saved_at"] = time.time()
        await _write_index(owner, oidx["settings"], oidx["lists"], oidx["legacy_leads"])
        return {"ok": True, "leads": len(body.leads),
                "lists": (await get_lists(request, email))["lists"]}
    idx = await _ensure_lists(email)
    lists = idx["lists"]
    if not any(l["id"] == rid for l in lists):
        raise HTTPException(status_code=404, detail="No such list.")
    await _write_list(email, rid, body.leads)
    for l in lists:
        if l["id"] == rid:
            l["count"] = len(body.leads)
            l["saved_at"] = time.time()
    await _write_index(email, idx["settings"], lists, idx["legacy_leads"])
    return {"ok": True, "leads": len(body.leads),
            "lists": (await get_lists(request, email))["lists"]}


@app.get("/api/lists/{list_id}/shares")
async def list_shares(list_id: str, request: Request, email: str = Depends(signed_in)):
    owner, rid = _split_ref(list_id)
    if owner and owner != email:
        raise HTTPException(status_code=403, detail="Only the owner can see who a list is shared with.")
    _check_list_id(rid)
    idx = await _ensure_lists(email)
    entry = next((l for l in idx["lists"] if l["id"] == rid), None)
    if not entry:
        raise HTTPException(status_code=404, detail="No such list.")
    return {"shares": entry.get("shares") or []}


@app.post("/api/lists/{list_id}/shares")
async def add_share(list_id: str, body: ShareRequest, request: Request, email: str = Depends(signed_in)):
    owner, rid = _split_ref(list_id)
    if owner and owner != email:
        raise HTTPException(status_code=403, detail="Only the owner can share a list.")
    _check_list_id(rid)
    to = str(body.email).lower().strip()
    if to == email:
        raise HTTPException(status_code=400, detail="That is your own address.")
    role = body.role if body.role in ("editor", "viewer") else "editor"
    idx = await _ensure_lists(email)
    entry = next((l for l in idx["lists"] if l["id"] == rid), None)
    if not entry:
        raise HTTPException(status_code=404, detail="No such list.")
    shares = [s for s in (entry.get("shares") or []) if s.get("email") != to]
    shares.append({"email": to, "role": role, "at": time.time()})
    entry["shares"] = shares
    await _write_index(email, idx["settings"], idx["lists"], idx["legacy_leads"])
    await _grant(email, rid, entry.get("name") or "Shared list", to, role)
    return {"shares": shares}


@app.delete("/api/lists/{list_id}/shares/{who}")
async def drop_share(list_id: str, who: str, request: Request, email: str = Depends(signed_in)):
    owner, rid = _split_ref(list_id)
    who = who.lower().strip()
    # Either the owner revokes, or the recipient walks away from a list they
    # did not ask for. Both are legitimate; nobody else may do either.
    if owner and owner != email:
        if who != email:
            raise HTTPException(status_code=403, detail="You can only remove your own access.")
        await _revoke(owner, rid, email)
        return {"ok": True, "shares": []}
    _check_list_id(rid)
    idx = await _ensure_lists(email)
    entry = next((l for l in idx["lists"] if l["id"] == rid), None)
    if not entry:
        raise HTTPException(status_code=404, detail="No such list.")
    entry["shares"] = [s for s in (entry.get("shares") or []) if s.get("email") != who]
    await _write_index(email, idx["settings"], idx["lists"], idx["legacy_leads"])
    await _revoke(email, rid, who)
    return {"ok": True, "shares": entry["shares"]}


@app.patch("/api/lists/{list_id}")
async def rename_list(list_id: str, body: ListPatch, request: Request, email: str = Depends(signed_in)):
    _check_list_id(list_id)
    name = (body.name or "").strip()[:60]
    if not name:
        raise HTTPException(status_code=400, detail="A list needs a name.")
    idx = await _ensure_lists(email)
    entry = next((l for l in idx["lists"] if l["id"] == list_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail="No such list.")
    entry["name"] = name
    await _write_index(email, idx["settings"], idx["lists"], idx["legacy_leads"])
    return {"list": entry, "lists": idx["lists"]}


@app.delete("/api/lists/{list_id}")
async def delete_list(list_id: str, request: Request, email: str = Depends(signed_in)):
    owner, list_id = _split_ref(list_id)
    if owner and owner != email:
        raise HTTPException(
            status_code=403,
            detail=f"That list belongs to {owner}. Remove your own access instead of deleting it.")
    _check_list_id(list_id)
    if list_id == "default":
        raise HTTPException(
            status_code=400,
            detail="That is your master list — every lead you have ever added lives "
                   "on it, which is exactly why it cannot be deleted. Campaign lists "
                   "are the deletable kind.")
    idx = await _ensure_lists(email)
    if len(idx["lists"]) <= 1:
        raise HTTPException(status_code=400, detail="This is your only list — rename it instead of deleting it.")
    if not any(l["id"] == list_id for l in idx["lists"]):
        raise HTTPException(status_code=404, detail="No such list.")
    gone = next(l for l in idx["lists"] if l["id"] == list_id)
    lists = [l for l in idx["lists"] if l["id"] != list_id]
    await _write_index(email, idx["settings"], lists, idx["legacy_leads"])
    key = _list_key(email, list_id)
    await _fs_del(FS_LISTS, key)
    _MEM_LISTS.pop(key, None)
    # Everyone it was shared with loses it too, or their switcher keeps
    # offering a list that no longer exists.
    for sh in (gone.get("shares") or []):
        await _revoke(email, list_id, sh.get("email") or "")
    return {"ok": True, "lists": (await get_lists(request, email))["lists"]}


# ------------------------- the admin view -------------------------
#
# "Both Advisors & Admin can add, upload, and share leads. The admin must have
# access to all leads and the ability to reclaim all leads."
#
# So an admin is an advisor with two extra powers, not a different application:
# they see the whole firm, and they can move any list to themselves. Everything
# else — building, importing, enriching, sharing — they do exactly as anyone
# else does, with their own lists.
#
# Reclaiming is a MOVE, not a copy. A copy would leave the advisor holding a
# second, diverging set of the same people, which is how two advisors end up
# calling one prospect. The list arrives whole: leads, enrichment, activity
# history, and a note on the list saying where it came from and when.

def _admin_only(email: str) -> None:
    if not _is_admin(email):
        raise HTTPException(status_code=403, detail="Admins only.")


class TransferRequest(BaseModel):
    owner: str = ""          # who holds it now ("" = the admin themselves)
    list_id: str = ""        # "" with an owner means every list they have
    to: str = ""             # who gets it ("" = the admin themselves)


@app.get("/api/admin/overview")
async def admin_overview(request: Request, email: str = Depends(signed_in)):
    """Every advisor and every list in the firm, with what each is carrying."""
    _admin_only(email)
    advisors, lists = [], []
    for row in await _directory():
        who = row["email"]
        idx = await _read_index(who)
        theirs = idx["lists"]
        advisors.append({"email": who, "lists": len(theirs),
                         "leads": sum(int(l.get("count") or 0) for l in theirs),
                         "first_seen": row.get("first_seen"),
                         "last_seen": row.get("last_seen"),
                         "is_admin": _is_admin(who)})
        for l in theirs:
            lists.append({"ref": f"{who}~{l['id']}", "owner": who,
                          "id": l["id"], "name": l.get("name") or "Untitled",
                          "count": int(l.get("count") or 0),
                          "saved_at": l.get("saved_at"),
                          "created_at": l.get("created_at"),
                          "shares": [sh.get("email") for sh in (l.get("shares") or [])],
                          "reclaimed_from": l.get("reclaimed_from") or ""})
    lists.sort(key=lambda l: l["count"], reverse=True)
    return {"you": email, "advisors": advisors, "lists": lists,
            "total_leads": sum(a["leads"] for a in advisors),
            "backend": storage_backend()}


async def _transfer_one(frm: str, list_id: str, to: str) -> dict:
    """Move one list, whole, from one advisor to another.

    The leads document is rewritten under the new owner and deleted from the
    old one, every share on it is revoked, and the entry records where it came
    from. Anything less than all of that leaves a copy somewhere, and a second
    copy of a call list is how two advisors phone the same person.
    """
    src = await _read_index(frm)
    entry = next((l for l in src["lists"] if l["id"] == list_id), None)
    if not entry:
        raise HTTPException(status_code=404, detail=f"{frm} has no list {list_id}.")
    leads = await _read_list(frm, list_id)

    dst = await _ensure_lists(to)
    new_id = list_id
    if any(l["id"] == new_id for l in dst["lists"]):
        new_id = "l" + secrets.token_urlsafe(9).replace("-", "_")
    moved = {"id": new_id, "name": entry.get("name") or "Reclaimed list",
             "created_at": entry.get("created_at") or time.time(),
             "count": len(leads), "saved_at": time.time(),
             "reclaimed_from": frm, "reclaimed_at": time.time()}
    await _write_list(to, new_id, leads)
    await _write_index(to, dst["settings"], dst["lists"] + [moved], dst["legacy_leads"])

    for sh in (entry.get("shares") or []):
        await _revoke(frm, list_id, sh.get("email") or "")
    remaining = [l for l in src["lists"] if l["id"] != list_id]
    await _write_index(frm, src["settings"], remaining, src["legacy_leads"])
    key = _list_key(frm, list_id)
    await _fs_del(FS_LISTS, key)
    _MEM_LISTS.pop(key, None)
    return {"from": frm, "to": to, "list": moved, "leads": len(leads)}


@app.post("/api/admin/transfer")
async def admin_transfer(body: TransferRequest, request: Request, email: str = Depends(signed_in)):
    """Reclaim a list to the admin, or hand one out to an advisor.

    One operation in two directions: reclaiming is a transfer to yourself,
    assigning is a transfer from yourself. Naming one endpoint for both is what
    keeps the two halves from drifting into different rules about shares and
    counts.
    """
    _admin_only(email)
    frm = (body.owner or email).lower().strip()
    to = (body.to or email).lower().strip()
    if frm == to:
        raise HTTPException(status_code=400,
                            detail="That list is already there — name a different owner.")
    if "@" not in frm or "@" not in to:
        raise HTTPException(status_code=400, detail="Both sides need an email address.")

    if body.list_id:
        _check_list_id(body.list_id)
        moved = [await _transfer_one(frm, body.list_id, to)]
    else:
        # Every list this advisor holds — the case that actually comes up, when
        # someone leaves and their pipeline has to keep being worked.
        src = await _read_index(frm)
        if not src["lists"]:
            raise HTTPException(status_code=404, detail=f"{frm} has no lists.")
        moved = []
        for l in list(src["lists"]):
            moved.append(await _transfer_one(frm, l["id"], to))
    return {"moved": moved, "leads": sum(m["leads"] for m in moved),
            "overview": await admin_overview(request, email)}


class SettingsSave(BaseModel):
    settings: dict


@app.put("/api/settings")
async def save_settings(body: SettingsSave, request: Request, email: str = Depends(signed_in)):
    """Settings belong to the user, not to any one list."""
    idx = await _ensure_lists(email)
    await _write_index(email, body.settings, idx["lists"], idx["legacy_leads"])
    return {"ok": True}


# ------------------------- free enrichment -------------------------
# Two things that cost nothing and were already half-built.
#
# The Form 5500 file was fetched to price WARN employers. The same file prices
# *any* employer: assets over participants is an average balance, and that puts
# a dollar figure on a lead whose only other data is a job title. It is an
# order of magnitude, not a quote, and it is labelled as one.
#
# And a fetcher for a page a person names — a company leadership page, a
# licensing register, an obituary. See webapp/harvest.py for why "legally
# accessible" is encoded in the tool rather than left to the user.

HARVEST_USER_AGENT = os.environ.get("HARVEST_USER_AGENT", "")


class PlanRequest(BaseModel):
    employers: list = []


@app.post("/api/plans")
async def employer_plans(body: PlanRequest, request: Request):
    """Retirement-plan size for named employers, from the DOL Form 5500 file.

    Bulk on purpose: the file is large and fetched once, so asking about forty
    employers costs exactly what asking about one does.
    """
    await _active_token(request)
    if not FORM5500_URL:
        return {"plans": {}, "note": "FORM5500_URL is not set — see SETUP-prospecting.md."}
    try:
        plans = (await _load_plans(await _drive_token_for(request))).get("plans") or {}
    except Exception as e:
        raise HTTPException(status_code=502,
                            detail=f"Form 5500 unavailable: {type(e).__name__}: {str(e)[:200]}")
    out = {}
    for name in (body.employers or [])[:500]:
        key = prospecting.norm_company(name or "")
        p = plans.get(key)
        if not p:
            continue
        # avg_balance is computed by the parser, so the WARN pricing and this
        # lookup cannot drift into two different definitions of the same number.
        out[name] = {k: p.get(k) for k in
                     ("plan_name", "participants", "assets", "avg_balance", "sponsor",
                      "state", "plan_year")}
    return {"plans": out, "asked": len(body.employers or []), "matched": len(out)}


class HarvestRequest(BaseModel):
    url: str


@app.post("/api/harvest")
async def harvest_page(body: HarvestRequest, request: Request):
    """One public page, fetched on the publisher's stated terms.

    Not a crawler: it fetches the URL given, once, and follows nothing.
    """
    await _active_token(request)
    if not HARVEST_USER_AGENT:
        raise HTTPException(
            status_code=400,
            detail="HARVEST_USER_AGENT is not set. Fetching a page without identifying "
                   "who is asking is exactly what this app will not do — see "
                   "SETUP-harvest.md.")
    async with httpx.AsyncClient() as cx:
        got = await harvest.fetch(cx, (body.url or "").strip(), HARVEST_USER_AGENT)
    if not got.get("ok"):
        raise HTTPException(status_code=422, detail=got.get("reason") or "Could not read that page.")
    # A whole page is more than any caller here needs and more than is polite to
    # keep; the useful part is what a person will read.
    got["text"] = got["text"][:40_000]
    return got


class SiteRequest(BaseModel):
    website: str
    company: str = ""
    extract: bool = True


@app.post("/api/harvest/site")
async def harvest_site(body: SiteRequest, request: Request):
    """Read one company's own site for who runs it and how long it has run.

    Aimed at the population a proxy statement cannot answer for: the owner of a
    private company. EDGAR prints an age for Section 16 officers and nobody
    else, and most lists of local business owners are private companies, so the
    free age signal is unavailable for exactly the leads it would matter most
    for. A firm's own About page frequently states what no vendor sells --
    "founded by Frank Delgado in 1987", "serving Long Island for over 35 years".

    Still not a crawler: a fixed list of conventional paths on the one host
    given, each through the same fetch() that honours robots.txt, the terms
    denylist and the rate limit. See webapp/harvest.py.

    Nothing here is written onto a lead. It returns findings with the sentence
    each came from and the page it was on, for a person to accept or discard --
    ADR 17 on why a derived value arriving dressed as an observed one is worse
    than a blank.
    """
    await _active_token(request)
    if not HARVEST_USER_AGENT:
        raise HTTPException(
            status_code=400,
            detail="HARVEST_USER_AGENT is not set. Fetching a page without identifying "
                   "who is asking is exactly what this app will not do — see "
                   "SETUP-harvest.md.")

    async with httpx.AsyncClient() as cx:
        got = await harvest.read_site(cx, (body.website or "").strip(), HARVEST_USER_AGENT)

    if not got.get("ok"):
        return {"ok": False, "reason": got.get("reason") or "Nothing readable on that site.",
                "rule": got.get("rule", ""), "root": got.get("root", ""),
                "pages": [], "tried": got.get("tried", []), "findings": None}

    pages = [{"url": p["url"], "title": p["title"], "chars": p["chars"]}
             for p in got["pages"]]
    out = {"ok": True, "root": got["root"], "pages": pages,
           "tried": got.get("tried", []), "findings": None}
    if not body.extract or not ANTHROPIC_API_KEY:
        # The text is returned unread when there is no key, so the feature
        # degrades to "fetched it, here it is" rather than failing.
        out["text"] = "\n\n".join(
            f"--- {p['url']} ---\n{p['text']}" for p in got["pages"])[:40_000]
        if not ANTHROPIC_API_KEY and body.extract:
            out["note"] = "ANTHROPIC_API_KEY is not set, so the pages were not read."
        return out

    doc = "\n\n".join(f"--- {p['url']} ---\n{p['text']}" for p in got["pages"])[:120_000]
    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    prompt = (
        "You are reading the public website of a company, to learn who owns or runs "
        "it and how long it has operated. Report only what the pages actually say.\n\n"
        f"Company: {body.company or got['root']}\n\n"
        "Reply with one JSON object and nothing else:\n"
        '{"owners": [{"name": "<as printed>", "title": "<as printed or null>", '
        '"quote": "<the sentence saying it, max 200 chars>"}], '
        '"year_founded": <integer or null>, "year_founded_quote": "<sentence or null>", '
        '"years_in_business": <integer or null>, "years_in_business_quote": "<sentence or null>", '
        '"ages": [{"name": "<person>", "age": <integer>, "quote": "<sentence>"}], '
        '"career_start_year": <integer or null>, "career_start_quote": "<sentence or null>"}\n\n'
        "Rules. Every value needs a quote from the page it came from; a value you "
        "cannot quote is one you inferred, and it does not belong here. Do not "
        "estimate an age from a founding year, a career length or a photograph — "
        "report age only where the page prints a number of years old for a named "
        "person. 'Serving Long Island since 1987' is year_founded, not an age. "
        "'Over 35 years in the industry' is years_in_business. Return empty lists "
        "and nulls rather than guesses: this feeds a call list, and a wrong owner "
        "or a wrong age is worse than a blank.\n\nPAGES:\n" + doc
    )
    try:
        msg = await client.messages.create(
            model=CLAUDE_MODEL, max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
    except anthropic.APIStatusError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error {e.status_code}: {str(e)[:200]}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {str(e)[:200]}")

    raw = "".join(b.text for b in msg.content if b.type == "text")
    a, z = raw.find("{"), raw.rfind("}")
    if a == -1 or z == -1:
        raise HTTPException(status_code=502, detail="Claude returned no JSON object.")
    try:
        found = json.loads(raw[a:z + 1])
    except ValueError:
        raise HTTPException(status_code=502, detail="Claude returned malformed JSON.")

    out["findings"] = _clean_site_findings(found)
    return out


# --- Attom property data (TRIAL) --------------------------------------------
# A 10-day trial key, added to see whether a home value earns its place next
# to the plan-balance proxy. Everything about this is built to be temporary:
# one env var gates it, answers cache durably (sealed, like WhitePages — a
# home value at a named address is PII), and if the trial lapses the buttons
# vanish while fetched values stay on their leads. The evaluation itself is
# the What's-converting panel: it reads Attom like any other enrichment and
# reports whether valued leads convert better.
ATTOM_API_KEY = os.environ.get("ATTOM_API_KEY", "")
ATTOM_BASE = os.environ.get("ATTOM_BASE", "https://api.gateway.attomdata.com")
ATTOM_TTL = 365 * 86400          # trial data does not go stale inside the trial

_ATTOM_CACHE: dict = {}


async def _attom_cached(key: str):
    hit = _ATTOM_CACHE.get(key)
    if hit and time.time() - hit[0] < ATTOM_TTL:
        return True, hit[1]
    doc = await _fs_get(FS_CACHE, "attom:" + hashlib.sha256(key.encode()).hexdigest())
    if not doc or time.time() - float(doc.get("at") or 0) >= ATTOM_TTL:
        return False, None
    raw = _unseal(doc)
    if raw is None:
        return False, None
    try:
        value = json.loads(raw)
    except ValueError:
        return False, None
    _ATTOM_CACHE[key] = (float(doc["at"]), value)
    return True, value


async def _attom_remember(key: str, value) -> None:
    _ATTOM_CACHE[key] = (time.time(), value)
    doc = _seal(json.dumps(value))
    doc["at"] = time.time()
    await _fs_set(FS_CACHE, "attom:" + hashlib.sha256(key.encode()).hexdigest(), doc)


class AttomRequest(BaseModel):
    street: str = ""
    city: str = ""
    state: str = ""
    zip: str = ""


def _attom_pick(js: dict) -> dict:
    """The fields worth keeping from either property endpoint, normalised."""
    props = js.get("property") or []
    if not props:
        return {}
    p = props[0] or {}
    avm = (p.get("avm") or {}).get("amount") or {}
    assessment = p.get("assessment") or {}
    market = (assessment.get("market") or {})
    sale = p.get("sale") or {}
    sale_amt = ((sale.get("amount") or {}).get("saleamt")
                or (sale.get("amount") or {}).get("saleAmt"))
    summary = p.get("summary") or {}
    owner = p.get("owner") or {}
    owners = []
    for k in ("owner1", "owner2", "owner3", "owner4"):
        o = owner.get(k) or {}
        nm = " ".join(x for x in (o.get("firstnameandmi") or o.get("firstNameAndMi"),
                                  o.get("lastname") or o.get("lastName")) if x).strip()
        if nm:
            owners.append(nm[:120])
    out = {
        "avm": avm.get("value"),
        "avm_high": avm.get("high"), "avm_low": avm.get("low"),
        "avm_score": avm.get("scr"),
        "market_value": (market.get("mktttlvalue") or market.get("mktTtlValue")),
        "assessed_value": ((assessment.get("assessed") or {}).get("assdttlvalue")
                           or (assessment.get("assessed") or {}).get("assdTtlValue")),
        "sale_amount": sale_amt,
        "sale_date": (sale.get("salesearchdate") or sale.get("saleSearchDate")
                      or sale.get("saleTransDate")),
        "year_built": (summary.get("yearbuilt") or summary.get("yearBuilt")),
        "owners": owners,
    }
    return {k: v for k, v in out.items() if v not in (None, "", [], 0)}


@app.post("/api/attom")
async def attom_property(body: AttomRequest, request: Request):
    """One address, priced. AVM first (an estimate with a confidence score),
    the expanded profile as fallback and for owner/sale/year-built. Both
    answers — including "no property found" — are remembered durably, so the
    trial's spend is never repeated on the same door."""
    await _active_token(request)
    if not ATTOM_API_KEY:
        raise HTTPException(status_code=400, detail="ATTOM_API_KEY is not set — the trial feature is off.")
    street = body.street.strip()
    locality = ", ".join(x for x in (body.city.strip(),
                                     " ".join(y for y in (body.state.strip(), body.zip.strip()) if y)) if x)
    if not street or not locality:
        raise HTTPException(status_code=400,
                            detail="Attom needs a street address and a city/state — enrich the "
                                   "household first, or accept a web-research location.")
    key = ("attom|" + street + "|" + locality).lower()
    found_cached, cached = await _attom_cached(key)
    if found_cached:
        return {**cached, "cached": True}

    headers = {"apikey": ATTOM_API_KEY, "accept": "application/json"}
    params = {"address1": street, "address2": locality}
    picked, tried, err = {}, [], ""
    async with httpx.AsyncClient(timeout=20) as cx:
        for path in ("/propertyapi/v1.0.0/attomavm/detail",
                     "/propertyapi/v1.0.0/property/expandedprofile"):
            tried.append(path.rsplit("/", 1)[-1])
            try:
                r = await cx.get(ATTOM_BASE + path, params=params, headers=headers)
                if r.status_code == 401:
                    raise HTTPException(status_code=502,
                                        detail="Attom rejected the key (401) — the trial may have ended.")
                if r.status_code == 400:
                    # Attom answers "no such property" as a 400 SuccessWithoutResult.
                    continue
                r.raise_for_status()
                got = _attom_pick(r.json())
                # AVM answers value; the profile fills what AVM lacks.
                for k, v in got.items():
                    picked.setdefault(k, v)
                if picked.get("avm") and picked.get("owners"):
                    break
            except HTTPException:
                raise
            except Exception as e:                       # noqa: BLE001 — reported, not swallowed
                err = f"{type(e).__name__}: {str(e)[:120]}"
    if not picked:
        answer = {"found": False,
                  "reason": err or "No property on file at that address.", "tried": tried}
        await _attom_remember(key, answer)
        return answer
    answer = {"found": True, "address": f"{street}, {locality}", **picked}
    await _attom_remember(key, answer)
    return answer


class WebResearchRequest(BaseModel):
    first_name: str = ""
    last_name: str = ""
    employer: str = ""
    title: str = ""
    state: str = ""


@app.post("/api/web-research")
async def web_research(body: WebResearchRequest, request: Request):
    """One lead, searched the way a person would.

    "If I was doing this manually... Search google for Tim Shaughnessy +
    Preferred Construction + Business Owner." This does that search through
    the Claude API's web-search tool — a licensed API, so the boundary in
    CLAUDE.md holds — and returns only what pages actually say, each finding
    carrying its quote and its source URL. Nothing lands on the lead from
    here: the client shows findings for the operator to accept one by one,
    the same discipline as the site reader and for the same reason (ADR 17).

    Per-lead and operator-initiated only. This is a token cost (Tier 3
    class); it is deliberately NOT part of the free sweep, because a search
    per lead across a 342-lead list is a real spend nobody asked for.
    """
    await _active_token(request)
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=400, detail="ANTHROPIC_API_KEY is not set on this service.")
    name = f"{body.first_name} {body.last_name}".strip()
    if not body.last_name.strip():
        raise HTTPException(status_code=400, detail="Web research needs at least a last name.")
    q = " ".join(x for x in (
        f'"{name}"',
        f'"{body.employer.strip()}"' if body.employer.strip() else "",
        body.title.strip(), body.state.strip()) if x)
    # The dynamic-filter search tool ships on the current model generation;
    # older models keep the basic variant.
    new_gen = any(k in CLAUDE_MODEL for k in ("opus-5", "sonnet-5", "fable", "4-6", "4-7", "4-8"))
    tool_type = "web_search_20260209" if new_gen else "web_search_20250305"
    prompt = (
        "You are researching one sales lead the way a careful person would, "
        f"to fill in what a data vendor could not. Search the web for: {q}\n\n"
        "Report ONLY what pages actually say about THIS person at THIS employer. "
        "If you cannot tie a page to both the name and the employer, it is a "
        "namesake and does not belong in the answer. Reply with one JSON object "
        "and nothing else:\n"
        '{"summary": "<2 sentences on who this appears to be, or null>", '
        '"location": {"city": "<city or null>", "state": "<2-letter state or null>", '
        '"quote": "<sentence saying it>", "url": "<page>"}, '
        '"age_hints": [{"hint": "<what the page says, e.g. over two decades at the helm>", '
        '"quote": "<sentence>", "url": "<page>"}], '
        '"ages": [{"name": "<person>", "age": <integer>, "quote": "<sentence printing the age>", "url": "<page>"}], '
        '"spouse": {"name": "<name or null>", "quote": "<sentence>", "url": "<page>"}, '
        '"office_phone": {"number": "<as printed or null>", "quote": "<context>", "url": "<page>"}, '
        '"email_pattern": {"pattern": "<e.g. first@domain.com or null>", "quote": "<what showed it>", "url": "<page>"}, '
        '"links": {"company_site": "<url or null>", "linkedin": "<url or null>", '
        '"instagram": "<url or null>", "facebook": "<url or null>"}}\n\n'
        "Rules. Every value needs a quote and the URL it came from; a value you "
        "cannot quote is one you inferred, and it does not belong here. Do not "
        "compute an age from a founding year or a photograph — age_hints carries "
        "the raw wording instead, and ages[] only what a page prints as a number "
        "for a named person. Return nulls and empty lists rather than guesses: "
        "this feeds a call list, and a wrong person is worse than a blank."
    )
    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = await client.messages.create(
            model=CLAUDE_MODEL, max_tokens=2000,
            tools=[{"type": tool_type, "name": "web_search", "max_uses": 4}],
            messages=[{"role": "user", "content": prompt}],
        )
    except anthropic.APIStatusError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error {e.status_code}: {str(e)[:200]}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {str(e)[:200]}")
    raw = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text")
    a, z = raw.find("{"), raw.rfind("}")
    if a == -1 or z == -1:
        raise HTTPException(status_code=502, detail="Claude returned no JSON object.")
    try:
        found = json.loads(raw[a:z + 1])
    except ValueError:
        raise HTTPException(status_code=502, detail="Claude returned malformed JSON.")
    return {"ok": True, "searched": q, "found": _clean_web_findings(found)}


def _clean_web_findings(f: dict) -> dict:
    """Drop anything unquoted before it reaches a screen, same as the site
    reader and for the same reason: an unsourced value dressed as a sourced
    one is worse than a blank."""
    def txt(v, n=200):
        return str(v)[:n] if v else None

    def quoted(d, *fields):
        if not isinstance(d, dict) or not d.get("quote") or not d.get("url"):
            return None
        if not any(d.get(x) for x in fields):
            return None
        return d

    loc = quoted(f.get("location"), "city", "state")
    if loc:
        st = txt(loc.get("state"), 2)
        loc = {"city": txt(loc.get("city"), 80),
               "state": st.upper() if st and len(st) == 2 and st.isalpha() else None,
               "quote": txt(loc.get("quote")), "url": txt(loc.get("url"), 300)}
        if not (loc["city"] or loc["state"]):
            loc = None
    hints = [{"hint": txt(h.get("hint"), 120), "quote": txt(h.get("quote")), "url": txt(h.get("url"), 300)}
             for h in (f.get("age_hints") or [])
             if isinstance(h, dict) and h.get("hint") and h.get("quote") and h.get("url")][:5]
    ages = [{"name": txt(a.get("name"), 120), "age": a.get("age"),
             "quote": txt(a.get("quote")), "url": txt(a.get("url"), 300)}
            for a in (f.get("ages") or [])
            if isinstance(a, dict) and a.get("name") and a.get("quote") and a.get("url")
            and isinstance(a.get("age"), int) and 18 <= a["age"] <= 100][:3]
    spouse = quoted(f.get("spouse"), "name")
    if spouse:
        spouse = {"name": txt(spouse.get("name"), 120), "quote": txt(spouse.get("quote")),
                  "url": txt(spouse.get("url"), 300)}
    phone = quoted(f.get("office_phone"), "number")
    if phone:
        digits = re.sub(r"\D", "", str(phone.get("number") or ""))
        phone = ({"number": txt(phone.get("number"), 30), "quote": txt(phone.get("quote")),
                  "url": txt(phone.get("url"), 300)} if len(digits) >= 10 else None)
    pat = quoted(f.get("email_pattern"), "pattern")
    if pat:
        pat = {"pattern": txt(pat.get("pattern"), 80), "quote": txt(pat.get("quote")),
               "url": txt(pat.get("url"), 300)}
    links = {k: txt(v, 300) for k, v in (f.get("links") or {}).items()
             if k in ("company_site", "linkedin", "instagram", "facebook")
             and isinstance(v, str) and v.startswith("http")}
    return {"summary": txt(f.get("summary"), 400), "location": loc, "age_hints": hints,
            "ages": ages, "spouse": spouse, "office_phone": phone,
            "email_pattern": pat, "links": links}


def _clean_site_findings(f: dict) -> dict:
    """Drop anything unquoted or out of range before it reaches the caller.

    The prompt asks for a quote behind every value; this enforces it, because a
    model that ignores that instruction once would otherwise put an unsourced
    age on screen looking exactly like a sourced one.
    """
    now = time.gmtime().tm_year

    def year(v):
        return v if isinstance(v, int) and 1800 <= v <= now else None

    owners = [{"name": str(o.get("name"))[:120],
               "title": (str(o["title"])[:120] if o.get("title") else None),
               "quote": str(o.get("quote"))[:200]}
              for o in (f.get("owners") or [])
              if isinstance(o, dict) and o.get("name") and o.get("quote")]

    ages = [{"name": str(a.get("name"))[:120], "age": a.get("age"),
             "quote": str(a.get("quote"))[:200]}
            for a in (f.get("ages") or [])
            if isinstance(a, dict) and a.get("quote")
            and isinstance(a.get("age"), int) and 18 <= a["age"] <= 100]

    yf, yq = year(f.get("year_founded")), f.get("year_founded_quote")
    yib = f.get("years_in_business")
    yibq = f.get("years_in_business_quote")
    cs, csq = year(f.get("career_start_year")), f.get("career_start_quote")
    return {
        "owners": owners[:12],
        "ages": ages[:12],
        "year_founded": yf if yq else None,
        "year_founded_quote": str(yq)[:200] if (yf and yq) else None,
        "years_in_business": (yib if (isinstance(yib, int) and 0 < yib <= 150 and yibq) else None),
        "years_in_business_quote": str(yibq)[:200] if (yibq and isinstance(yib, int)) else None,
        "career_start_year": cs if csq else None,
        "career_start_quote": str(csq)[:200] if (cs and csq) else None,
    }


# ------------------------- money-in-motion signals -------------------------
# The rest of the app finds people. This watches the ones already found and
# says when something makes their retirement money movable. Everything here is
# free: an age is arithmetic, a WARN notice and an 8-K are public filings.

FS_SEEN = os.environ.get("FIRESTORE_SEEN_COLLECTION", "signals_seen")
_MEM_SEEN: dict[str, dict] = {}


async def _seen_ids(email: str) -> set:
    doc = await _fs_get(FS_SEEN, email) or _MEM_SEEN.get(email) or {}
    try:
        return set(json.loads(doc.get("data") or "[]"))
    except ValueError:
        return set()


async def _mark_seen(email: str, ids: list) -> None:
    # Bounded, newest-last: an advisor who works a list for a year should not
    # accumulate an unbounded document of ids they will never see again.
    keep = list(dict.fromkeys(list(await _seen_ids(email)) + list(ids)))[-4000:]
    payload = {"data": json.dumps(keep), "saved_at": time.time()}
    if not await _fs_set(FS_SEEN, email, payload):
        _MEM_SEEN[email] = payload


class SignalsRequest(BaseModel):
    leads: list = []
    min_tenure: float = 18.0
    mark_seen: bool = False


@app.post("/api/signals")
async def get_signals(body: SignalsRequest, request: Request, email: str = Depends(signed_in)):
    """Every money-in-motion event across the leads posted.

    The leads come from the client rather than being read server-side, so this
    works on the open list whether or not it has been saved, and on a list
    shared from another advisor without this route needing to re-derive access.
    """

    warn_index, warn_note = {}, ""
    warn_ran = False
    _feeds_val = (await _warn_feeds_value())[0].strip()
    if not _feeds_val:
        # This used to say nothing at all, so the panel reported that every lead
        # had been "checked against the WARN feeds" on a deployment with no WARN
        # feeds configured. Zero events then reads as "nothing is happening"
        # when it means "nothing was looked at".
        warn_note = ("No WARN feed source is set, so mass-separation notices are not "
                     "checked. An admin can set it in Money in motion.")
    if _feeds_val:
        try:
            got = await _load_warn(await _drive_token_for(request))
            plans = {}
            try:
                plans = (await _load_plans(await _drive_token_for(request))).get("plans") or {}
            except Exception:
                pass
            opps = prospecting.build_opportunities(
                got["events"], plans, min_workers=1,
                states=_source_states() or None, counties=_source_counties() or None)
            warn_index = signals.index_warn(opps)
            warn_ran = True
        except Exception as e:
            warn_note = f"WARN feeds unavailable: {type(e).__name__}"

    # One EDGAR round-trip per distinct employer, not per lead: a list of forty
    # people at four companies costs four lookups.
    filings, filing_note, filings_ran = {}, "", False
    if EDGAR_USER_AGENT:
        filings_ran = True
        employers = {}
        for L in body.leads:
            name = (L.get("employer") or "").strip()
            if name:
                employers.setdefault(signals.norm_company(name), name)
        for key, name in list(employers.items())[:25]:
            try:
                hit = await _edgar_company_cik(name)
                if not hit:
                    continue
                f = await _edgar_recent_8k(hit["cik"])
                if f:
                    filings[key] = f
            except Exception:
                continue
    else:
        filing_note = "EDGAR_USER_AGENT is not set, so officer-departure filings are not checked."

    seen = await _seen_ids(email)
    out = signals.build_signals(body.leads, warn_index, filings,
                                min_tenure=body.min_tenure, seen=seen)
    if body.mark_seen and out:
        await _mark_seen(email, [s["id"] for s in out])
    # Three detectors run here and each can be silent for two very different
    # reasons: nothing happened, or nothing was looked at. Zero events is only
    # good news when the checks actually ran, so the coverage says which.
    #
    # The age count is the one that matters most. It is the only detector that
    # needs nothing configured — but it needs an age, and a list with no ages
    # produces the same confident zero as a list where nobody is near 59 1/2.
    with_age = sum(1 for L in body.leads if signals._age_now(L)[0] is not None)
    dated = sum(1 for L in body.leads if signals.half_month(L) is not None)
    return {"signals": out, "new": sum(1 for s in out if s["new"]),
            "checked": len(body.leads),
            "coverage": {
                "leads": len(body.leads),
                "with_age": with_age,
                "with_birth_date": dated,
                "warn": warn_ran,
                "warn_events": len(warn_index),
                "filings": filings_ran,
                "employers_checked": len(filings),
            },
            "notes": [n for n in (warn_note, filing_note) if n]}


# ------------------------- team: stats, leaderboard, battles -------------------------
# What an advisor did today, as four integers. The alternative — deriving the
# leaderboard from everyone's lead documents — means reading every lead in the
# firm to draw one table, and it would expose lists that were never shared.
# A counter document leaks nothing but the count.

STAT_KEYS = ("calls", "emails", "invites", "meetings")


def _stat_key(email: str, day: str) -> str:
    return f"{email}__{day}"


def _today(offset_days: int = 0) -> str:
    return time.strftime("%Y-%m-%d", time.gmtime(time.time() - offset_days * 86400))


class StatsPost(BaseModel):
    day: str = ""
    calls: int = 0
    emails: int = 0
    invites: int = 0
    meetings: int = 0


@app.put("/api/stats")
async def put_stats(body: StatsPost, request: Request, email: str = Depends(signed_in)):
    """Today's counters for the signed-in advisor. Idempotent: the client sends
    totals for the day, not increments, so a replay cannot inflate a score."""
    day = body.day if re.match(r"^\d{4}-\d{2}-\d{2}$", body.day or "") else _today()
    payload = {"email": email, "day": day, "domain": _domain(email),
               "saved_at": time.time()}
    for k in STAT_KEYS:
        payload[k] = max(0, int(getattr(body, k, 0) or 0))
    key = _stat_key(email, day)
    if not await _fs_set(FS_STATS, key, payload):
        _MEM_STATS[key] = payload
    return {"ok": True, "day": day}


async def _team_members(email: str) -> list:
    """Colleagues: same email domain, plus anyone in a shared-list relationship.

    Domain is the rule a single firm actually wants and needs no administering.
    Sharing adds the cases it misses — an advisor at another firm you work a
    list with.
    """
    people = {email}
    if TEAM_BY_DOMAIN and _domain(email):
        dom = _domain(email)
        for key, doc in list(_MEM_STATS.items()):
            if doc.get("domain") == dom:
                people.add(doc["email"])
        db = _firestore()
        if db is not None:
            try:
                q = db.collection(FS_STATS).where("domain", "==", dom).limit(500)
                async for d in q.stream():
                    v = d.to_dict() or {}
                    if v.get("email"):
                        people.add(v["email"])
            except Exception as e:
                print(f"[team] domain query failed: {e}")
    for r in await _shared_index(email):
        people.add(r["owner"])
    idx = await _read_index(email)
    for l in idx["lists"]:
        for sh in (l.get("shares") or []):
            if sh.get("email"):
                people.add(sh["email"])
    return sorted(people)


async def _stats_for(people: list, days: int) -> dict:
    wanted = {_today(i) for i in range(max(1, days))}
    out = {p: {k: 0 for k in STAT_KEYS} for p in people}
    for p in people:
        for day in wanted:
            key = _stat_key(p, day)
            doc = await _fs_get(FS_STATS, key) or _MEM_STATS.get(key)
            if not doc:
                continue
            for k in STAT_KEYS:
                out[p][k] += int(doc.get(k) or 0)
    return out


def _score_row(row: dict) -> int:
    """One number to rank on. A meeting is the product; a call is the input.

    The weights are deliberately lopsided so nobody wins a contest by dialling
    numbers they never intended to talk to.
    """
    return (row["calls"] * 1) + (row["emails"] * 1) + (row["invites"] * 3) + (row["meetings"] * 10)


@app.get("/api/leaderboard")
async def leaderboard(request: Request, days: int = 7, email: str = Depends(signed_in)):
    days = max(1, min(int(days or 7), 90))
    people = await _team_members(email)
    stats = await _stats_for(people, days)
    rows = [dict(stats[p], email=p, you=(p == email), points=_score_row(stats[p]))
            for p in people]
    rows.sort(key=lambda r: (-r["points"], r["email"]))
    for i, r in enumerate(rows, 1):
        r["rank"] = i
    return {"days": days, "rows": rows, "team_size": len(people)}


# --- battles ---------------------------------------------------------------
# A contest is a leaderboard with a start, an end and a named field. Everything
# scoring-related is the same code; only the window and the roster differ.

class BattleCreate(BaseModel):
    name: str
    days: int = 1
    opponents: list = []
    metric: str = "points"


@app.get("/api/battles")
async def get_battles(request: Request, email: str = Depends(signed_in)):
    out = []
    db = _firestore()
    docs = []
    if db is not None:
        try:
            async for d in db.collection(FS_BATTLES).limit(200).stream():
                docs.append(d.to_dict() or {})
        except Exception as e:
            print(f"[battles] read failed: {e}")
    docs += list(_MEM_BATTLES.values())
    seen = set()
    for b in docs:
        if b.get("id") in seen:
            continue
        seen.add(b.get("id"))
        roster = b.get("players") or []
        if email not in roster:
            continue
        stats = {}
        for p in roster:
            row = {k: 0 for k in STAT_KEYS}
            day = b.get("start_day") or _today()
            for i in range(max(1, int(b.get("days") or 1))):
                d = time.strftime("%Y-%m-%d",
                                  time.gmtime(time.mktime(time.strptime(day, "%Y-%m-%d")) + i * 86400))
                key = _stat_key(p, d)
                doc = await _fs_get(FS_STATS, key) or _MEM_STATS.get(key)
                if doc:
                    for k in STAT_KEYS:
                        row[k] += int(doc.get(k) or 0)
            stats[p] = row
        rows = [dict(stats[p], email=p, you=(p == email),
                     points=stats[p][b["metric"]] if b.get("metric") in STAT_KEYS
                     else _score_row(stats[p])) for p in roster]
        rows.sort(key=lambda r: (-r["points"], r["email"]))
        for i, r in enumerate(rows, 1):
            r["rank"] = i
        out.append({**{k: b.get(k) for k in ("id", "name", "metric", "days", "start_day", "created_by")},
                    "rows": rows, "over": _battle_over(b)})
    out.sort(key=lambda x: (x["over"], x.get("start_day") or ""), reverse=False)
    return {"battles": out}


def _battle_over(b: dict) -> bool:
    try:
        start = time.mktime(time.strptime(b.get("start_day") or _today(), "%Y-%m-%d"))
    except Exception:
        return False
    return time.time() > start + max(1, int(b.get("days") or 1)) * 86400


@app.post("/api/battles")
async def create_battle(body: BattleCreate, request: Request, email: str = Depends(signed_in)):
    name = (body.name or "").strip()[:60]
    if not name:
        raise HTTPException(status_code=400, detail="A contest needs a name.")
    players = sorted({email} | {str(o).lower().strip() for o in (body.opponents or []) if str(o).strip()})
    if len(players) < 2:
        raise HTTPException(status_code=400, detail="Pick at least one person to go up against.")
    if len(players) > 25:
        raise HTTPException(status_code=400, detail="25 people is the limit for one contest.")
    metric = body.metric if body.metric in STAT_KEYS or body.metric == "points" else "points"
    bid = "b" + secrets.token_urlsafe(9).replace("-", "_")
    doc = {"id": bid, "name": name, "metric": metric,
           "days": max(1, min(int(body.days or 1), 31)),
           "start_day": _today(), "created_by": email, "players": players,
           "created_at": time.time()}
    if not await _fs_set(FS_BATTLES, bid, doc):
        _MEM_BATTLES[bid] = doc
    return {"battle": {k: doc[k] for k in ("id", "name", "metric", "days", "start_day")}}


@app.delete("/api/battles/{bid}")
async def delete_battle(bid: str, request: Request, email: str = Depends(signed_in)):
    doc = await _fs_get(FS_BATTLES, bid) or _MEM_BATTLES.get(bid)
    if not doc:
        raise HTTPException(status_code=404, detail="No such contest.")
    if doc.get("created_by") != email:
        raise HTTPException(status_code=403, detail="Only whoever started it can end it.")
    await _fs_del(FS_BATTLES, bid)
    _MEM_BATTLES.pop(bid, None)
    return {"ok": True}


# ------------------------- Google Drive import -------------------------

# The sheet the research tool writes into. Named rather than picked, because the
# whole point is that the advisor never goes looking for it: the tool appends
# rows on its own schedule and the app notices.
DRIVE_DEFAULT_NAME = os.environ.get("DRIVE_LEADS_FILE", "Wealth Management Lead Prospecting")


async def _google_only(request: Request) -> str:
    """Drive needs the Google token specifically, not whichever is signed in."""
    provider, token = await _active_token(request)
    if provider != "google":
        raise HTTPException(
            status_code=400,
            detail="Drive import needs a Google sign-in. Sign out and sign in with Google.",
        )
    return token


def _rows_from_csv(text: str) -> list:
    import csv, io
    return [r for r in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(blob: bytes) -> list:
    """First worksheet, as a list of string rows.

    read_only keeps a large export from being held in memory twice, and
    data_only takes computed values rather than formula text.
    """
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                rows.append(cells)
        return rows
    finally:
        wb.close()


@app.get("/api/drive/find")
async def drive_find(request: Request, name: str = ""):
    """Spreadsheets in the user's Drive matching a name. Newest first."""
    token = await _google_only(request)
    wanted = (name or DRIVE_DEFAULT_NAME).replace("'", "\\'")
    query = (
        f"name contains '{wanted}' and trashed = false and ("
        f"mimeType = '{SHEET_MIME}' or mimeType = '{XLSX_MIME}' or mimeType = 'text/csv')"
    )
    async with httpx.AsyncClient(timeout=30) as cx:
        r = await cx.get(DRIVE_FILES_URL, headers={"Authorization": f"Bearer {token}"},
                         params={"q": query, "orderBy": "modifiedTime desc", "pageSize": 20,
                                 "fields": "files(id,name,mimeType,modifiedTime,size)",
                                 "supportsAllDrives": "true",
                                 "includeItemsFromAllDrives": "true"})
    if r.status_code == 403:
        raise HTTPException(
            status_code=502,
            detail="Google refused the Drive request. Sign out and back in to grant "
                   "Drive access — the permission was added after your last sign-in.",
        )
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Drive error {r.status_code}: {r.text[:300]}")
    return {"files": r.json().get("files", []), "searched": name or DRIVE_DEFAULT_NAME}


@app.get("/api/drive/rows")
async def drive_rows(request: Request, id: str):
    """A Drive spreadsheet as raw rows, for the existing column mapper."""
    token = await _google_only(request)
    headers = {"Authorization": f"Bearer {token}"}
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as cx:
        meta = await cx.get(f"{DRIVE_FILES_URL}/{id}", headers=headers,
                            params={"fields": "id,name,mimeType", "supportsAllDrives": "true"})
        if meta.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Drive error {meta.status_code}: {meta.text[:200]}")
        info = meta.json()
        mime = info.get("mimeType", "")

        if mime == SHEET_MIME:      # native Sheets must be exported, not downloaded
            resp = await cx.get(f"{DRIVE_FILES_URL}/{id}/export", headers=headers,
                                params={"mimeType": "text/csv"})
        else:
            resp = await cx.get(f"{DRIVE_FILES_URL}/{id}", headers=headers,
                                params={"alt": "media", "supportsAllDrives": "true"})
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Drive download failed {resp.status_code}: {resp.text[:200]}")

    try:
        if mime == XLSX_MIME or info.get("name", "").lower().endswith(".xlsx"):
            rows = _rows_from_xlsx(resp.content)
        else:
            rows = _rows_from_csv(resp.text)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Could not read {info.get('name')}: {e}")

    if not rows:
        raise HTTPException(status_code=400, detail=f"{info.get('name')} is empty.")
    return {"name": info.get("name"), "rows": rows[:5001], "truncated": len(rows) > 5001}


# ------------------------- AI lead QC (Claude) -------------------------

QC_RULES = {
    "base_age_min": 25,
    "base_age_max": 75,
    "net_worth_min": 2_000_000,
    "young_age_max": 45,
    "young_income_min": 250_000,
    "old_401k_min": 250_000,
    "wl_age_max": 70,
    "intent_assets_min": 250_000,
}

def qc_prompt(base_age_min: float) -> str:
    """The grading rule, built per request.

    The age floor comes from the caller so the AI grades against the same
    definition of a qualified lead as the scorer does. Hard-coding it here is
    how the two drifted apart: the app moved to a pre-retiree rule while this
    prompt still passed anyone from 25 up.
    """
    return f"""You are the quality-control engine for a wealth-management lead generation company. Evaluate each lead against this qualification rule:

    BASE REQUIREMENT: Age between {base_age_min:g} and {QC_RULES['base_age_max']} (estimate age from graduation year + 22, or total career length if grad year absent).

    Then at least ONE gate must hold:
    - NW: net worth > ${QC_RULES['net_worth_min']:,}
    - YHE: age < {QC_RULES['young_age_max']} AND income > ${QC_RULES['young_income_min']:,}
    - 401K: orphaned 401(k) balance > ${QC_RULES['old_401k_min']:,} (proxy: changed jobs in the last 1-5 years AFTER a long prior tenure at a company likely to offer a 401(k))
    - WL: age < {QC_RULES['wl_age_max']} AND holds whole life insurance (NEVER inferable from prospect data — always UNKNOWN unless the record explicitly confirms it)
    - INT: actively seeking financial help AND investable assets > ${QC_RULES['intent_assets_min']:,} (intent requires an explicit signal in the record)

    If a lead record includes an explicit "age" value, use it verbatim with ageStatus "CONFIRMED" instead of estimating.
    Status vocabulary per gate: "CONFIRMED" (record explicitly states it), "INFERRED" (strong proxy: seniority, tenure, company size, title), "UNKNOWN" (no signal), "FAIL" (evidence contradicts it).
    Inference guides: senior titles (VP/SVP/C-suite/Partner/Principal/Owner/MD) at mid-size+ companies => income likely >$250K. Long tenure in high-income roles or equity titles => higher net-worth likelihood. Many stints under 3 years => job hopper (small 401k balances, penalize).
    "yearsExperience" is total career length and "yearsAtEmployer" is time in the current seat; the difference is years spent at previous employers, which is the tenure the 401K gate turns on. Prefer them over estimating from a start date.
    Property fields, when present, come from public records and are CONFIRMED rather than inferred: "ownsHome" true means the deed carries their name, "propertiesOwned" counts deeded properties (two or more is a strong net-worth signal), and "deedHeldBy" naming a trust or an entity means estate or entity planning has already happened — treat that as a strong NW signal and note the existing planning in the checklist.

    Return ONLY a JSON array, one object per lead, same order, no prose:
    [{{"i":0,"ageEst":57,"ageStatus":"INFERRED","gates":{{"NW":{{"s":"INFERRED","ev":"short evidence"}},"YHE":{{"s":"FAIL","ev":""}},"401K":{{"s":"INFERRED","ev":""}},"WL":{{"s":"UNKNOWN","ev":""}},"INT":{{"s":"UNKNOWN","ev":""}}}},"jobHopper":false,"grade":"A","checklist":["Confirm approximate net worth","Confirm old 401(k) balance"],"note":"one-line QC summary"}}]

    Grading: A = base age passes + 2 or more gates at INFERRED-or-better, or 1 CONFIRMED gate, and not a job hopper. B = base age passes + exactly 1 INFERRED gate. C = base age passes but only UNKNOWNs, or job hopper with otherwise decent signals. If base age FAILS, grade "X".
    "checklist" = the specific facts a junior advisor must verify on the first call before this lead counts as fully qualified.

    LEADS:
    """


class QCLead(BaseModel):
    i: int
    name: str = ""
    title: str = ""
    jobLevel: str = ""
    company: str = ""
    state: str = ""
    gradYear: str = ""
    positionStart: str = ""
    employees: str = ""
    notes: str = ""
    yearsExperience: str = ""
    yearsAtEmployer: str = ""
    jobHopper: bool = False
    # Confirmed facts from a WhitePages enrichment, when one has been run.
    age: Optional[int] = None
    ownsHome: Optional[bool] = None
    propertiesOwned: Optional[int] = None
    deedHeldBy: str = ""


class QCRequest(BaseModel):
    leads: list[QCLead]
    # The age floor the app is currently scoring on, so the grader and the
    # scorer agree on who is even eligible. Falls back to the stored rule.
    base_age_min: Optional[float] = None


@app.post("/api/qc")
async def qc(req: QCRequest, request: Request):
    # Each batch is a paid Claude API call — signed-in users only.
    await _active_token(request)
    if not ANTHROPIC_API_KEY:
        raise HTTPException(
            status_code=500,
            detail="AI QC not configured — set ANTHROPIC_API_KEY "
                   "(get one at console.anthropic.com).",
        )
    if not req.leads:
        return {"verdicts": []}
    if len(req.leads) > 12:
        raise HTTPException(status_code=400, detail="Max 12 leads per QC batch.")

    age_floor = req.base_age_min if req.base_age_min else QC_RULES["base_age_min"]
    payload = json.dumps([lead.model_dump() for lead in req.leads], separators=(",", ":"))
    client = anthropic.AsyncAnthropic(api_key=ANTHROPIC_API_KEY)
    try:
        msg = await client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=4000,
            messages=[{"role": "user", "content": qc_prompt(age_floor) + payload}],
        )
    except anthropic.APIStatusError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error {e.status_code}: {str(e)[:300]}")
    except anthropic.APIError as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {str(e)[:300]}")

    if msg.stop_reason == "refusal":
        raise HTTPException(status_code=502, detail="Claude declined to process this batch.")

    text = "".join(b.text for b in msg.content if b.type == "text")
    start, end = text.find("["), text.rfind("]")
    if start == -1 or end == -1:
        raise HTTPException(status_code=502, detail="Claude returned no JSON array.")
    try:
        verdicts = json.loads(text[start:end + 1])
    except ValueError:
        raise HTTPException(status_code=502, detail="Claude returned malformed JSON.")
    return {"verdicts": verdicts}


# ------------------------- Pages -------------------------

@app.api_route("/healthz", methods=["GET", "HEAD"])
async def healthz():
    return {"ok": True}


@app.api_route("/", methods=["GET", "HEAD"])
async def index():
    return FileResponse(STATIC_DIR / "index.html")
